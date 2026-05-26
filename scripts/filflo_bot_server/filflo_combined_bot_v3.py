"""
╔══════════════════════════════════════════════════════════════════╗
║     Filflo B2B Combined Bot v3 — Anveshan Farm Tech             ║
║     Delivery Marking + POD Upload in ONE form submission        ║
║                                                                  ║
║  Date entry uses React-compatible JS (native value setter).     ║
║  POD upload is OPTIONAL — uploaded in the same form when a      ║
║  matching file exists, otherwise just the date is saved.        ║
║  Post-save verification confirms status actually changed.       ║
╚══════════════════════════════════════════════════════════════════╝

Excel format:
    PO Number | Order Type | Delivery Date | Tracking ID | Status

POD files go in: C:\\Users\\lenovo\\Desktop\\Filflo_Bot\\POD_FILES\\
    Named by Tracking ID, e.g.: TRACK123.jpg, TRACK123.pdf

Usage:
    python filflo_combined_bot_v3.py --once
    python filflo_combined_bot_v3.py
    python filflo_combined_bot_v3.py --excel "path\\to\\file.xlsx"
"""

import os
import sys
import time
import csv
import glob as glob_module
import shutil
import logging
import argparse
import threading
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from datetime import datetime, timedelta
try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*_args, **_kwargs):
        return False

load_dotenv(Path(__file__).resolve().parent / ".env")

import openpyxl
from openpyxl.utils import get_column_letter
try:
    from filelock import FileLock
except ImportError:
    class FileLock:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

from po_status import POStatus, is_row_done, categorize_result
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    StaleElementReferenceException,
    WebDriverException,
    JavascriptException,
    NoAlertPresentException,
)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

FILFLO_URL     = "https://anveshan.filflo.in/b2b/b2b_order_team"
LOGIN_EMAIL    = os.getenv("FILFLO_LOGIN_EMAIL", "")
LOGIN_PASSWORD = os.getenv("FILFLO_LOGIN_PASSWORD", "")

BOT_FOLDER          = Path(os.getenv("FILFLO_BOT_FOLDER", str(Path(__file__).resolve().parent)))
DEFAULT_EXCEL_PATH  = BOT_FOLDER / "Filflo_Tasks.xlsx"
POD_FOLDER          = BOT_FOLDER / "POD_FILES"
POD_DONE_FOLDER     = BOT_FOLDER / "POD_FILES" / "_uploaded"
LOG_DIR             = BOT_FOLDER / "logs"
EXCEL_LOCK_PATH     = BOT_FOLDER / ".excel.lock"

POLL_INTERVAL_SECONDS = 60

# ── Validate required environment variables at import time ──
_REQUIRED_ENV_VARS = ["FILFLO_LOGIN_EMAIL", "FILFLO_LOGIN_PASSWORD"]
_missing = [v for v in _REQUIRED_ENV_VARS if not os.getenv(v)]
if _missing:
    import warnings
    warnings.warn(
        f"Missing required environment variables: {', '.join(_missing)}. "
        f"Create a .env file in {BOT_FOLDER} — see .env.example for reference.",
        stacklevel=2,
    )

# Download & Email config
DOWNLOAD_FOLDER     = BOT_FOLDER   # Downloaded file saved here
GMAIL_SENDER        = os.getenv("FILFLO_GMAIL_SENDER", "")
GMAIL_APP_PASSWORD  = os.getenv("FILFLO_GMAIL_APP_PASSWORD", "")
EMAIL_RECIPIENT     = os.getenv("FILFLO_EMAIL_RECIPIENT", "")
GOOGLE_SHEET_ID     = os.getenv("FILFLO_GOOGLE_SHEET_ID", "").strip()
GOOGLE_WORKSHEET_NAME = os.getenv("FILFLO_GOOGLE_WORKSHEET_NAME", "Sheet1").strip() or "Sheet1"
GOOGLE_SHEET_ID_2   = os.getenv("FILFLO_GOOGLE_SHEET_ID_2", "").strip()
GOOGLE_WORKSHEET_NAME_2 = os.getenv("FILFLO_GOOGLE_WORKSHEET_NAME_2", "Sheet1").strip() or "Sheet1"
GOOGLE_SYNC_ENABLED = os.getenv("FILFLO_SYNC_DOWNLOAD_TO_GSHEET", "true").strip().lower() not in {"0", "false", "no", "off"}
GOOGLE_CLIENT_SECRET = BOT_FOLDER / "client_secret.json"
GOOGLE_AUTH_TOKEN    = BOT_FOLDER / "authorized_user.json"
GSHEET_BACKUP_DIR    = BOT_FOLDER / "google_sheet_backups"

# Excel columns (1-indexed)
COL_PO_NUMBER     = 1
COL_ORDER_TYPE    = 2
COL_DELIVERY_DATE = 3
COL_TRACKING_ID   = 4
COL_STATUS        = 5
HEADER_ROW        = 1

# Timeouts
PAGE_LOAD_TIMEOUT    = 30
ELEMENT_WAIT_TIMEOUT = 20
MAX_PO_RETRIES       = 3
SEARCH_POLL_SECONDS  = 18


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

import uuid
from logging.handlers import RotatingFileHandler
from filflo_monitor_bus import attach_monitor_handler

# ── Correlation ID for grouping log lines per PO processing run ──
_correlation_id = threading.local() if "threading" in dir() else type("", (), {"id": "MAIN"})()


class CorrelationFormatter(logging.Formatter):
    """Adds a correlation ID to every log line for tracing PO processing."""
    def format(self, record):
        record.correlation_id = getattr(_correlation_id, "id", "MAIN")
        return super().format(record)


def set_correlation_id(po_number: str = ""):
    """Set a correlation ID for the current thread (call at start of each PO)."""
    short_id = uuid.uuid4().hex[:8]
    _correlation_id.id = f"{po_number}:{short_id}" if po_number else short_id


def setup_logging(log_dir: Path) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"combined_bot_{datetime.now():%Y%m%d}.log"

    logger = logging.getLogger("FilfloCombinedBot")
    logger.setLevel(logging.DEBUG)

    if logger.handlers:
        attach_monitor_handler(logger, source="filflo")
        return logger

    # File handler with rotation (10 MB max, keep 5 backups)
    fh = RotatingFileHandler(
        log_file, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8"
    )
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(CorrelationFormatter(
        "%(asctime)s.%(msecs)03d | %(levelname)-8s | [%(correlation_id)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    ))

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(CorrelationFormatter(
        "%(asctime)s | %(levelname)-8s | [%(correlation_id)s] %(message)s",
        datefmt="%H:%M:%S"
    ))

    logger.addHandler(fh)
    logger.addHandler(ch)
    attach_monitor_handler(logger, source="filflo")
    return logger


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def ensure_excel_headers(path: Path, logger):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = {
        COL_PO_NUMBER: "PO Number",
        COL_ORDER_TYPE: "Order Type",
        COL_DELIVERY_DATE: "Delivery Date",
        COL_TRACKING_ID: "Tracking ID",
        COL_STATUS: "Status",
    }
    changed = False
    for col, name in headers.items():
        if ws.cell(row=HEADER_ROW, column=col).value != name:
            ws.cell(row=HEADER_ROW, column=col, value=name)
            changed = True
    if changed:
        wb.save(path)
    wb.close()


def read_pending_entries(path: Path, logger):
    lock = FileLock(str(EXCEL_LOCK_PATH), timeout=30)
    with lock:
        wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    entries = []
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        po_number = ws.cell(row=row_idx, column=COL_PO_NUMBER).value
        if not po_number:
            continue

        status = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()

        # Skip fully done rows
        if is_row_done(status):
            continue

        tracking_raw = ws.cell(row=row_idx, column=COL_TRACKING_ID).value
        tracking_id = ""

        if tracking_raw is not None:
            tracking_id = str(tracking_raw).strip()
            # Remove hidden characters
            tracking_id = tracking_id.replace("\u200b", "").replace("\ufeff", "").replace("\t", "")

            # Skip formula-based tracking IDs
            if tracking_id.startswith("="):
                logger.warning(f"Row {row_idx}: Tracking ID is a formula -- paste as plain values!")
                tracking_id = ""

            # Handle numeric tracking IDs (Excel may read as float like 6001029489.0)
            if tracking_id and "." in tracking_id:
                try:
                    tracking_id = str(int(float(tracking_id)))
                except (ValueError, OverflowError):
                    pass

        po_raw = str(po_number).strip()
        po_str = normalize_po_number(po_raw)
        if po_str != po_raw:
            logger.info(f"Row {row_idx}: Normalized PO '{po_raw}' -> '{po_str}'")
        logger.debug(f"Row {row_idx}: PO='{po_str}', Tracking ID='{tracking_id}' (raw={repr(tracking_raw)})")

        # Input validation
        if not validate_po_number(po_str):
            logger.warning(f"Row {row_idx}: Invalid PO format '{po_str}' -- skipping row.")
            continue
        if tracking_id and not validate_tracking_id(tracking_id):
            logger.warning(f"Row {row_idx}: Invalid tracking ID format '{tracking_id}' -- clearing it.")
            tracking_id = ""

        entries.append({
            "row": row_idx,
            "po_number": po_str,
            "order_type": str(ws.cell(row=row_idx, column=COL_ORDER_TYPE).value or "").strip(),
            "delivery_date": ws.cell(row=row_idx, column=COL_DELIVERY_DATE).value,
            "tracking_id": tracking_id,
            "status": status,
        })

    wb.close()
    return entries


def update_excel_status(path: Path, row_idx: int, status: str, logger):
    lock = FileLock(str(EXCEL_LOCK_PATH), timeout=30)
    try:
        with lock:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.cell(row=row_idx, column=COL_STATUS,
                    value=f"{status} ({datetime.now():%Y-%m-%d})")
            wb.save(path)
            wb.close()
    except PermissionError:
        logger.error(f"Row {row_idx}: Excel file is open. Please close it.")
    except (PermissionError, OSError) as e:
        logger.error(f"Row {row_idx}: Excel update failed -- {e}")


def validate_po_number(po_str: str) -> bool:
    """Validate PO number format. Returns True if valid."""
    import re
    if not po_str or len(po_str) < 3 or len(po_str) > 60:
        return False
    # Allow alphanumeric, hyphens, underscores, slashes
    return bool(re.match(r"^[A-Za-z0-9\-_/]+$", po_str))


def validate_tracking_id(tracking_id: str) -> bool:
    """Validate tracking ID format. Returns True if valid."""
    if not tracking_id:
        return True  # Empty is OK (means no tracking)
    if len(tracking_id) < 3 or len(tracking_id) > 50:
        return False
    # Allow alphanumeric and hyphens only
    import re
    return bool(re.match(r"^[A-Za-z0-9\-]+$", tracking_id))


ALLOWED_POD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf", ".tiff", ".tif", ".bmp", ".gif"}


def validate_pod_file(file_path) -> bool:
    """Validate POD file has an allowed extension."""
    from pathlib import Path
    ext = Path(file_path).suffix.lower()
    return ext in ALLOWED_POD_EXTENSIONS


def normalize_po_number(raw_po) -> str:
    po_str = str(raw_po or "").strip()
    po_str = po_str.replace("\u200b", "").replace("\ufeff", "").replace("\t", "")
    # Excel sometimes wraps text PO numbers with apostrophes/quotes to preserve formatting.
    return po_str.strip("'").strip('"').strip()


def parse_delivery_date(raw_date) -> str:
    """Convert delivery date to DD-MM-YYYY format."""
    if isinstance(raw_date, datetime):
        return raw_date.strftime("%d-%m-%Y")

    date_str = str(raw_date).strip()
    formats = [
        "%d-%b-%y", "%d-%b-%Y", "%d-%m-%Y", "%d/%m/%Y",
        "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            if dt.year < 100:
                dt = dt.replace(year=dt.year + 2000)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            continue

    raise ValueError(f"Cannot parse delivery date: '{raw_date}'")


def should_prefer_all_time(raw_date) -> bool:
    """
    Older deliveries are usually outside Filflo's default 'Last 30 Days' filter.
    Use the Excel date as a fast heuristic so we don't waste one search cycle first.
    """
    if not raw_date:
        return False

    try:
        normalized = parse_delivery_date(raw_date)
        delivery_dt = datetime.strptime(normalized, "%d-%m-%Y").date()
        return (datetime.now().date() - delivery_dt).days > 25
    except (ValueError, TypeError):
        return False


def find_pod_file(folder: Path, tracking_id: str, logger=None) -> Path | None:
    """
    Search for POD file by tracking ID — EXACT match only.
    File must be named exactly as the Tracking ID (e.g., TRACK123.jpg for Tracking ID 'TRACK123').
    No partial matching to prevent wrong POD being attached.
    Handles whitespace, hidden characters, and case differences.
    """
    if not tracking_id or not folder.exists():
        return None

    # Aggressive cleanup: remove ALL whitespace, hidden chars, BOM, zero-width chars
    tid = tracking_id.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
    if not tid:
        return None

    supported_ext = (".jpg", ".jpeg", ".jfif", ".png", ".pdf", ".csv", ".xlsx", ".xls")

    # EXACT stem match — also clean up filenames the same way
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
        if clean_stem == tid:
            if logger:
                logger.info(f"POD file matched: Tracking ID '{tid}' -> {f.name}")
            return f

    # Case-insensitive fallback (same exact ID but different case)
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
        if clean_stem.lower() == tid.lower():
            if logger:
                logger.info(f"POD file matched (case-insensitive): Tracking ID '{tid}' -> {f.name}")
            return f

    # Prefix match fallback: file stem starts with the tracking ID
    # Handles files like "301203752 a.pdf" matching tracking ID "301203752"
    # Only matches if the character after the ID is a space, hyphen, underscore, or letter suffix
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "")
        if clean_stem.lower().startswith(tid.lower()) and len(clean_stem) > len(tid):
            # Ensure the extra part is a harmless suffix (space, letter, hyphen, underscore)
            remainder = clean_stem[len(tid):]
            if remainder.strip().replace("-", "").replace("_", "").isalpha() or remainder.startswith(" "):
                if logger:
                    logger.info(f"POD file matched (prefix match): Tracking ID '{tid}' -> {f.name}")
                return f

    if logger:
        # Log available files to help debug
        available = [f.stem for f in folder.iterdir()
                     if f.is_file() and f.suffix.lower() in supported_ext][:10]
        logger.info(f"No POD file for Tracking ID '{tid}' (repr: {repr(tracking_id)}). "
                     f"Sample files in folder: {available}")
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  SELENIUM HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def create_driver(logger) -> webdriver.Chrome:
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-features=CalculateNativeWinOcclusion")
    options.page_load_strategy = "normal"
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "password_manager.leak_detection": False,
    })
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    logger.info("Launching Chrome browser...")
    try:
        driver = webdriver.Chrome(options=options)
    except WebDriverException as e:
        logger.warning(f"Standard launch failed ({e}), trying with Service...")
        driver = webdriver.Chrome(service=Service(), options=options)

    try:
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    except WebDriverException:
        pass
    return driver


def dismiss_popups(driver, logger):
    restore_browser_if_needed(driver, logger, "popup dismissal")
    try:
        alert = driver.switch_to.alert
        alert_text = alert.text
        if logger:
            logger.info(f"Dismissing browser alert: '{alert_text[:80]}'")
        alert.accept()  # Click OK (works for both OK-only and OK/Cancel alerts)
    except (NoAlertPresentException, WebDriverException):
        pass

    for xpath in [
        "//button[text()='OK'] | //button[text()='Close'] | //button[text()='Not now']"
        " | //button[contains(text(), 'Close')] | //button[contains(text(), 'Dismiss')]",
        "//*[@aria-label='Close'] | //*[@aria-label='close'] | //button[contains(@class, 'close')]"
    ]:
        try:
            for btn in driver.find_elements(By.XPATH, xpath):
                if btn.is_displayed():
                    btn.click()
                    time.sleep(0.3)
                    break
        except (NoSuchElementException, StaleElementReferenceException, ElementNotInteractableException):
            pass


def wait_and_find(driver, by, value, timeout=ELEMENT_WAIT_TIMEOUT, clickable=False, logger=None):
    condition_factory = (EC.element_to_be_clickable if clickable else EC.presence_of_element_located)
    condition = condition_factory((by, value))
    locator_desc = f"{by}={value}"

    def guarded_condition(drv):
        restore_browser_if_needed(drv, logger, f"wait for {locator_desc[:80]}")
        return condition(drv)

    return WebDriverWait(driver, timeout).until(guarded_condition)


def _visible_descendants(root, xpath: str, include_file_inputs=False):
    matches = []
    for el in root.find_elements(By.XPATH, xpath):
        try:
            if el.is_displayed():
                matches.append(el)
            elif include_file_inputs and el.get_attribute("type") == "file":
                matches.append(el)
        except StaleElementReferenceException:
            continue
    return matches


def find_panel_root_by_heading(driver, heading_text: str, logger, required_xpaths=None, screenshot_prefix=None):
    """
    Starting from a visible heading, walk up the DOM to find the smallest visible
    ancestor that contains the required controls. This keeps file inputs/buttons
    scoped to the active panel instead of the whole page.
    """
    restore_browser_if_needed(driver, logger, f"locating panel '{heading_text}'")
    required_xpaths = required_xpaths or []
    headings = [
        el for el in driver.find_elements(By.XPATH, f"//*[contains(normalize-space(text()), '{heading_text}')]")
        if el.is_displayed()
    ]
    if len(headings) != 1:
        logger.error(f"Expected 1 visible heading '{heading_text}', found {len(headings)}.")
        if screenshot_prefix:
            capture_debug_screenshot(driver, logger, screenshot_prefix)
        return None

    heading = headings[0]
    ancestors = heading.find_elements(By.XPATH, "./ancestor::*")
    for candidate in reversed(ancestors):
        try:
            if not candidate.is_displayed():
                continue
        except StaleElementReferenceException:
            continue

        ok = True
        for xpath in required_xpaths:
            include_file = "@type='file'" in xpath or '@type="file"' in xpath
            if not _visible_descendants(candidate, xpath, include_file_inputs=include_file):
                ok = False
                break
        if ok:
            return candidate

    logger.error(f"Could not find a scoped panel root for heading '{heading_text}'.")
    if screenshot_prefix:
        capture_debug_screenshot(driver, logger, screenshot_prefix)
    return None


def capture_debug_screenshot(driver, logger, prefix: str) -> Path | None:
    """Save a timestamped screenshot to help diagnose wrong-row or validation issues."""
    try:
        shots_dir = LOG_DIR / "screenshots"
        shots_dir.mkdir(parents=True, exist_ok=True)
        shot_path = shots_dir / f"{prefix}_{datetime.now():%Y%m%d_%H%M%S}.png"
        driver.save_screenshot(str(shot_path))
        logger.info(f"Saved debug screenshot: {shot_path}")
        return shot_path
    except (WebDriverException, JavascriptException) as e:
        logger.warning(f"Could not save debug screenshot ({prefix}): {e}")
        return None


def restore_browser_if_needed(driver, logger=None, reason: str = "interaction") -> bool:
    """
    Restore/focus Chrome when it gets minimized or occluded.
    POD upload is especially sensitive to Windows background throttling.
    """
    visibility_state = None
    rect = {}
    try:
        visibility_state = driver.execute_script("return document.visibilityState || 'visible';")
    except (WebDriverException, JavascriptException):
        pass

    try:
        rect = driver.get_window_rect() or {}
    except (WebDriverException, JavascriptException):
        rect = {}

    width = rect.get("width") or 0
    height = rect.get("height") or 0
    needs_restore = visibility_state == "hidden" or width < 500 or height < 400
    if not needs_restore:
        return False

    if logger:
        logger.info(f"Chrome window appears minimized/backgrounded; restoring before {reason}.")
    try:
        driver.switch_to.window(driver.current_window_handle)
    except WebDriverException:
        pass

    restored = False
    try:
        driver.maximize_window()
        restored = True
    except WebDriverException:
        pass

    if not restored:
        try:
            driver.set_window_rect(x=0, y=0, width=1440, height=1000)
            restored = True
        except WebDriverException:
            pass

    try:
        driver.execute_cdp_cmd("Page.bringToFront", {})
    except (WebDriverException, JavascriptException):
        pass

    try:
        driver.execute_script("window.focus();")
    except (WebDriverException, JavascriptException):
        pass

    if restored:
        time.sleep(1)
    return restored


def safe_click(driver, element, logger, retries=3):
    for attempt in range(retries):
        try:
            restore_browser_if_needed(driver, logger, "click")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.3)
            element.click()
            return True
        except ElementClickInterceptedException:
            time.sleep(0.5)
            if attempt == retries - 1:
                try:
                    driver.execute_script("arguments[0].click();", element)
                    return True
                except (JavascriptException, StaleElementReferenceException):
                    pass
        except StaleElementReferenceException:
            time.sleep(0.5)
    return False


def close_modal_if_open(driver, logger):
    restore_browser_if_needed(driver, logger, "modal cleanup")
    for attempt in range(3):
        try:
            backdrops = [b for b in driver.find_elements(By.XPATH,
                "//div[contains(@class, 'MuiBackdrop') or contains(@class, 'MuiModal-backdrop')]")
                if b.is_displayed()]
            if not backdrops:
                return

            try:
                cancel = driver.find_element(By.XPATH, "//button[contains(text(), 'Cancel')]")
                if cancel.is_displayed():
                    driver.execute_script("arguments[0].click();", cancel)
                    time.sleep(1.5)
                    continue
            except NoSuchElementException:
                pass

            try:
                driver.execute_script("arguments[0].click();", backdrops[0])
                time.sleep(0.8)
                continue
            except (JavascriptException, StaleElementReferenceException):
                pass

            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(0.8)
        except WebDriverException:
            pass

    try:
        backdrops = [b for b in driver.find_elements(By.XPATH,
            "//div[contains(@class, 'MuiBackdrop') or contains(@class, 'MuiModal-backdrop')]")
            if b.is_displayed()]
        if backdrops:
            logger.warning("Modal stuck -- refreshing page.")
            driver.get(FILFLO_URL)
            time.sleep(2)  # reduced from 4s
    except WebDriverException:
        pass


def is_session_valid(driver, logger) -> bool:
    try:
        driver.find_element(By.XPATH, "//input[@placeholder='Search orders...']")
        return True
    except (NoSuchElementException, WebDriverException):
        logger.warning("Session expired.")
        return False


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGIN & NAVIGATION
# ═══════════════════════════════════════════════════════════════════════════════

def login(driver, logger) -> bool:
    logger.info("Navigating to Filflo portal...")
    restore_browser_if_needed(driver, logger, "portal login")
    driver.get(FILFLO_URL)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    except TimeoutException:
        time.sleep(3)  # fallback

    try:
        driver.find_element(By.XPATH, "//input[@placeholder='Search orders...']")
        logger.info("Already logged in.")
        return True
    except NoSuchElementException:
        pass

    logger.info("Logging in...")
    try:
        email_input = wait_and_find(driver, By.XPATH,
            "//input[@type='email' or @name='email' or @placeholder='Email' "
            "or @placeholder='Enter your email' or @placeholder='Email address']", timeout=10, logger=logger)
        email_input.clear()
        email_input.send_keys(LOGIN_EMAIL)
        time.sleep(0.3)

        password_input = wait_and_find(driver, By.XPATH,
            "//input[@type='password' or @name='password' or @placeholder='Password']", timeout=5, logger=logger)
        password_input.clear()
        password_input.send_keys(LOGIN_PASSWORD)
        time.sleep(0.3)

        login_btn = wait_and_find(driver, By.XPATH,
            "//button[contains(text(),'Login') or contains(text(),'Sign In') "
            "or contains(text(),'Submit') or @type='submit']", timeout=5, clickable=True, logger=logger)
        safe_click(driver, login_btn, logger)

        time.sleep(2)  # reduced from 4s
        dismiss_popups(driver, logger)
        time.sleep(1)
        dismiss_popups(driver, logger)
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20, logger=logger)
        logger.info("Login successful.")
        return True
    except TimeoutException:
        logger.error("Login failed.")
        return False


def navigate_to_orders(driver, logger):
    if "b2b_order_team" not in driver.current_url:
        logger.info("Navigating to B2B Orders page...")
        restore_browser_if_needed(driver, logger, "orders navigation")
        driver.get(FILFLO_URL)
        time.sleep(2)  # reduced from 4s
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20, logger=logger)


def apply_all_time_filter(driver, logger) -> bool:
    try:
        restore_browser_if_needed(driver, logger, "date filter change")
        date_filter = wait_and_find(driver, By.XPATH,
            "//*[contains(text(), 'Last 30 Days') or contains(text(), 'Last 30 days')]",
            timeout=5, clickable=True, logger=logger)
        safe_click(driver, date_filter, logger)
        time.sleep(1)

        all_time = wait_and_find(driver, By.XPATH,
            "//*[contains(text(), 'All Time') or contains(text(), 'All time')]",
            timeout=5, clickable=True, logger=logger)
        safe_click(driver, all_time, logger)
        time.sleep(3)

        wait_for_table_rows(driver, timeout=12)
        logger.info("Applied 'All Time' filter.")
        return True
    except TimeoutException:
        logger.warning("Could not apply 'All Time' filter.")
        return False


def ensure_all_time_filter(driver, logger):
    """Only re-apply the filter if it has visibly reverted to 'Last 30 Days'."""
    try:
        restore_browser_if_needed(driver, logger, "filter verification")
        last30 = driver.find_elements(By.XPATH,
            "//*[contains(text(), 'Last 30 Days') or contains(text(), 'Last 30 days')]")
        if last30 and any(el.is_displayed() for el in last30):
            logger.info("Filter reverted to 'Last 30 Days' -- re-applying 'All Time'...")
            apply_all_time_filter(driver, logger)
        # If 'All Time' is visible or no filter text found, do nothing — filter is fine
    except (NoSuchElementException, StaleElementReferenceException):
        pass


# ═══════════════════════════════════════════════════════════════════════════════
#  SEARCH & ROW FINDING
# ═══════════════════════════════════════════════════════════════════════════════

def find_po_row(driver, po_number: str, logger, retries=3):
    """
    Find the table row with EXACT cell-level PO match only.
    No contains/partial matching — prevents clicking wrong PO row.
    e.g., 'REP-MUM-28022026-16' will NOT match 'SALE-REP-MUM-28022026-16' or 'REP-MUM-28022026-18'.
    """
    for attempt in range(retries):
        try:
            restore_browser_if_needed(driver, logger, f"finding row for {po_number}")
            rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
            stale = False
            matches = []

            for row in rows:
                try:
                    if not row.is_displayed():
                        continue
                    cells = row.find_elements(By.TAG_NAME, "td")
                    cell_texts = [cell.text.strip() for cell in cells[:4]]
                    if any(text == po_number for text in cell_texts):
                        matches.append((row, cell_texts))
                except StaleElementReferenceException:
                    stale = True
                    break

            if stale:
                time.sleep(1)
                continue

            if len(matches) == 1:
                logger.debug(f"Unique exact cell match for '{po_number}' in row: {matches[0][1]}")
                return matches[0][0]

            if len(matches) > 1:
                logger.error(f"SAFETY ABORT: Multiple visible rows matched PO '{po_number}': "
                             f"{[cells for _, cells in matches]}")
                capture_debug_screenshot(driver, logger, f"multiple_po_matches_{po_number}")
                return None

            if attempt < retries - 1:
                time.sleep(1)
        except (NoSuchElementException, StaleElementReferenceException):
            time.sleep(1)

    return None


def select_option_on_select(select_el, option_texts: tuple[str, ...], logger, label: str) -> bool:
    """Select an option on a specific <select> and dispatch React-friendly events."""
    try:
        restore_browser_if_needed(select_el.parent, logger, label)
    except WebDriverException:
        pass
    try:
        select = Select(select_el)
    except (WebDriverException, TypeError) as e:
        logger.error(f"{label}: element is not a usable <select>: {e}")
        return False

    selected_text = None
    for text in option_texts:
        try:
            select.select_by_visible_text(text)
            selected_text = text
            break
        except NoSuchElementException:
            continue

    if not selected_text:
        target_words = [t.lower() for t in option_texts]
        for idx, opt in enumerate(select.options):
            opt_text = opt.text.strip()
            lower = opt_text.lower()
            if any(target.lower() == lower for target in option_texts):
                select.select_by_index(idx)
                selected_text = opt_text
                break
            if len(option_texts) == 1 and all(word in lower for word in target_words[0].split()):
                select.select_by_index(idx)
                selected_text = opt_text
                break

    if not selected_text:
        available = [opt.text.strip() for opt in select.options]
        logger.error(f"{label}: option not found. Available options: {available}")
        return False

    try:
        driver = select_el.parent
    except AttributeError:
        driver = None

    try:
        select_el.parent.execute_script("""
            var el = arguments[0];
            var nativeSetter = Object.getOwnPropertyDescriptor(
                window.HTMLSelectElement.prototype, 'value').set;
            nativeSetter.call(el, el.value);
            el.dispatchEvent(new Event('change', { bubbles: true }));
            el.dispatchEvent(new Event('input', { bubbles: true }));
        """, select_el)
    except (WebDriverException, JavascriptException) as e:
        logger.warning(f"{label}: selected '{selected_text}' but event dispatch failed: {e}")
    else:
        logger.info(f"{label}: selected '{selected_text}'.")

    return True


def clear_search_field(driver, logger):
    search_input = wait_and_find(driver, By.XPATH,
        "//input[@placeholder='Search orders...']", timeout=10, logger=logger)

    restore_browser_if_needed(driver, logger, "search clear")
    search_input.click()
    time.sleep(0.2)
    search_input.send_keys(Keys.CONTROL + "a")
    search_input.send_keys(Keys.DELETE)
    time.sleep(0.3)

    if search_input.get_attribute("value"):
        driver.execute_script("""
            var el = arguments[0];
            var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            nativeInputValueSetter.call(el, '');
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        """, search_input)
        time.sleep(0.5)

    return search_input


def type_into_search(driver, po_number: str, logger) -> bool:
    for attempt in range(3):
        restore_browser_if_needed(driver, logger, f"typing search for {po_number}")
        search_input = clear_search_field(driver, logger)
        search_input.click()
        time.sleep(0.1)
        search_input.send_keys(po_number)
        time.sleep(0.5)

        actual = (search_input.get_attribute("value") or "").strip()
        if actual == po_number:
            return True

        # JS fallback
        driver.execute_script("""
            var el = arguments[0];
            var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            nativeInputValueSetter.call(el, arguments[1]);
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        """, search_input, po_number)
        time.sleep(0.5)

        actual = (search_input.get_attribute("value") or "").strip()
        if actual == po_number:
            return True
        time.sleep(0.5)

    logger.error(f"Could not type PO '{po_number}' into search.")
    return False


def wait_for_table_rows(driver, timeout=8):
    for _ in range(timeout):
        restore_browser_if_needed(driver, None, "table row wait")
        rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
        if rows:
            return len(rows)
        time.sleep(1)
    return 0


def _wait_for_table_stable(driver, logger, stable_seconds=2, max_wait=10):
    """
    Wait for the table to stop changing (React finished rendering).
    Returns True if table stabilized.
    """
    prev_count = -1
    prev_text = ""
    stable_for = 0

    for _ in range(max_wait):
        restore_browser_if_needed(driver, logger, "table stabilization")
        rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
        count = len(rows)
        # Get text of first row to detect content changes
        try:
            text = rows[0].text[:50] if rows else ""
        except (StaleElementReferenceException, IndexError):
            text = ""

        if count == prev_count and text == prev_text:
            stable_for += 1
            if stable_for >= stable_seconds:
                return True
        else:
            stable_for = 0

        prev_count = count
        prev_text = text
        time.sleep(1)

    return False


def _wait_for_search_to_filter(driver, pre_search_row_count: int, logger, max_wait=12):
    """
    Wait for the search filter to actually take effect.
    Detects filtering by waiting for the row count to CHANGE from before the search,
    OR for the table to show very few rows (1-3), which means filtering happened.
    """
    for waited in range(max_wait):
        restore_browser_if_needed(driver, logger, "search filtering")
        rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
        current_count = len(rows)

        # If row count decreased, the filter has kicked in
        if current_count < pre_search_row_count and current_count > 0:
            logger.debug(f"Search filter active: {pre_search_row_count} -> {current_count} rows (after {waited+1}s)")
            # Wait one more second for React to finish rendering
            time.sleep(1)
            return True

        # If we see very few rows (1-3), filter is likely done even if we didn't know the before count
        if 0 < current_count <= 3 and waited >= 2:
            logger.debug(f"Search filter active: {current_count} rows shown (after {waited+1}s)")
            time.sleep(1)
            return True

        time.sleep(1)

    logger.debug(f"Search filter wait timed out after {max_wait}s (rows: {pre_search_row_count} -> {len(driver.find_elements(By.XPATH, '//table//tbody//tr'))})")
    return False


def _do_search(driver, po_number: str, logger, poll_seconds=10) -> bool:
    """Type PO into search, wait for search filter to take effect, then check for match."""
    restore_browser_if_needed(driver, logger, f"search start for {po_number}")
    clear_search_field(driver, logger)
    time.sleep(2)

    # Count rows BEFORE searching (so we can detect when filtering happens)
    pre_search_rows = len(driver.find_elements(By.XPATH, "//table//tbody//tr"))
    logger.debug(f"Table has {pre_search_rows} rows before search.")

    if not type_into_search(driver, po_number, logger):
        return False

    # Wait for the search filter to actually take effect
    _wait_for_search_to_filter(driver, pre_search_rows, logger, max_wait=12)

    # Now wait for table to fully stabilize
    _wait_for_table_stable(driver, logger, stable_seconds=2, max_wait=6)

    # Check for the PO
    row = find_po_row(driver, po_number, logger, retries=3)
    if row:
        return True

    # Retry: retype and wait again
    table_rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
    try:
        si = wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=6, logger=logger)
        val = (si.get_attribute("value") or "").strip()
    except TimeoutException:
        logger.warning("Search input disappeared after filtering; letting the caller retry with a fresh page.")
        return False
    if not table_rows or val != po_number:
        logger.debug(f"Search retry: table has {len(table_rows)} rows, field='{val}'")
        clear_search_field(driver, logger)
        time.sleep(2)
        pre_search_rows = len(driver.find_elements(By.XPATH, "//table//tbody//tr"))
        wait_for_table_rows(driver, timeout=6)
        type_into_search(driver, po_number, logger)

        _wait_for_search_to_filter(driver, pre_search_rows, logger, max_wait=12)
        _wait_for_table_stable(driver, logger, stable_seconds=2, max_wait=6)

        row = find_po_row(driver, po_number, logger, retries=3)
        if row:
            return True

    return False


def search_order(driver, po_number: str, logger, prefer_all_time=False) -> bool:
    """
    Smart search:
      1. Search with the smartest initial filter for this PO
      2. If not found, switch to 'All Time' and retry — slower but catches older POs
      3. If still not found, full page refresh and final attempt
    """
    logger.info(f"Searching for PO: {po_number}")

    if prefer_all_time:
        logger.info(f"PO {po_number}: delivery date suggests an older order, applying 'All Time' first.")
        apply_all_time_filter(driver, logger)
        if _do_search(driver, po_number, logger, poll_seconds=12):
            logger.info(f"PO {po_number} found (All Time preferred).")
            return True
        logger.info(f"PO {po_number} not found with preferred 'All Time' filter -- falling back to normal flow...")

    # ── Attempt 1: Search with current filter (fast) ─────────────────────────
    if _do_search(driver, po_number, logger, poll_seconds=10):
        logger.info(f"PO {po_number} found (current filter).")
        return True

    # ── Attempt 2: Switch to 'All Time' and retry ───────────────────────────
    logger.info(f"PO {po_number} not found in current filter -- trying 'All Time'...")
    apply_all_time_filter(driver, logger)

    if _do_search(driver, po_number, logger, poll_seconds=12):
        logger.info(f"PO {po_number} found (All Time filter).")
        return True

    # ── Attempt 3: Full page refresh ─────────────────────────────────────────
    logger.info(f"PO {po_number} not found -- refreshing page...")
    try:
        driver.get(FILFLO_URL)
        time.sleep(2)  # reduced from 4s
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20)
        apply_all_time_filter(driver, logger)

        if _do_search(driver, po_number, logger, poll_seconds=10):
            logger.info(f"PO {po_number} found after page refresh.")
            return True
    except (TimeoutException, NoSuchElementException):
        pass

    logger.error(f"PO {po_number} not found after all attempts.")
    return False


def fresh_start_and_search(driver, po_number: str, logger, prefer_all_time=False) -> bool:
    """
    SAFETY: Full page reload + search for a specific PO.
    Eliminates ALL stale state from previous PO processing.
    This prevents wrong POD being attached to wrong PO.
    """
    logger.info(f"Fresh page load for PO: {po_number}")

    # Full page reload to clear all stale modals, panels, search results
    restore_browser_if_needed(driver, logger, f"fresh navigation for {po_number}")
    driver.get(FILFLO_URL)
    time.sleep(2)  # reduced from 4s

    try:
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20, logger=logger)
    except TimeoutException:
        logger.error("Page didn't load after fresh navigation.")
        return False

    dismiss_popups(driver, logger)
    time.sleep(1)

    # Now search
    return search_order(driver, po_number, logger, prefer_all_time=prefer_all_time)


def check_po_status(driver, po_number: str, row, logger) -> str:
    for attempt in range(3):
        try:
            restore_browser_if_needed(driver, logger, f"status check for {po_number}")
            if attempt > 0:
                row = find_po_row(driver, po_number, logger)
                if not row:
                    return "unknown"

            row_text = row.text.lower()
            for status in ("in transit", "grn entered", "open", "invoiced",
                           "dispatched", "delivered", "picked", "pending", "approved", "rto"):
                if status in row_text:
                    return status
            return "unknown"
        except StaleElementReferenceException:
            time.sleep(0.5)
    return "unknown"


def get_row_action_options(row) -> list[str]:
    """Return the visible option texts from the exact row action select, if present."""
    try:
        action_select = row.find_element(By.XPATH, ".//td[last()]//select")
        return [opt.text.strip() for opt in Select(action_select).options]
    except (NoSuchElementException, TimeoutException):
        return []


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN DELIVERY FORM — DATE + POD + SAVE (ALL IN ONE)
# ═══════════════════════════════════════════════════════════════════════════════

def click_enter_delivery_info(driver, po_number: str, logger) -> bool:
    """Click 'Choose an option' > 'Enter Delivery Info' to open the form."""
    try:
        restore_browser_if_needed(driver, logger, f"delivery form open for {po_number}")
        # Wait for table to stabilize before interacting
        time.sleep(2)
        _wait_for_table_stable(driver, logger, stable_seconds=2, max_wait=6)

        try:
            search_input = driver.find_element(By.XPATH, "//input[@placeholder='Search orders...']")
            search_value = (search_input.get_attribute("value") or "").strip()
            if search_value != po_number:
                logger.error(f"SAFETY ABORT: search field shows '{search_value}' but expected PO '{po_number}'.")
                capture_debug_screenshot(driver, logger, f"search_mismatch_{po_number}")
                return False
        except (NoSuchElementException, StaleElementReferenceException):
            pass

        row = find_po_row(driver, po_number, logger)
        if not row:
            logger.error(f"Row for {po_number} not found.")
            return False

        # VERIFY: confirm a cell in this row has the EXACT PO number
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            exact_match = any(c.text.strip() == po_number for c in cells[:4])
            if not exact_match:
                cell_texts = [c.text.strip() for c in cells[:4]]
                logger.error(f"Row mismatch! Looking for exact '{po_number}' but cells contain: {cell_texts}")
                return False
            logger.debug(f"Row verified (exact cell match) for PO {po_number}.")
        except StaleElementReferenceException:
            row = find_po_row(driver, po_number, logger)
            if not row:
                return False

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
        time.sleep(0.5)

        action_xpath = (".//td[last()]//button | .//td[last()]//*[contains(text(), 'Choose an option')]"
                        " | .//td[last()]//select | .//td[last()]//div[contains(@class, 'dropdown')]"
                        " | .//td[last()]//a[contains(text(), 'Choose')]")

        for find_attempt in range(3):
            try:
                action_btn = row.find_element(By.XPATH, action_xpath)
                break
            except (NoSuchElementException, StaleElementReferenceException):
                row = find_po_row(driver, po_number, logger)
                if not row:
                    return False
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
                time.sleep(0.5)
                if find_attempt == 2:
                    return False

        # SAFETY FIX: the row action is a <select> on Filflo. Select on the exact matched row only.
        if action_btn.tag_name.lower() == "select":
            if not select_option_on_select(
                action_btn,
                ("Enter Delivery Info",),
                logger,
                f"Row action for PO {po_number}",
            ):
                capture_debug_screenshot(driver, logger, f"delivery_action_missing_{po_number}")
                return False
            time.sleep(2)
        else:
            safe_click(driver, action_btn, logger)
            time.sleep(2)

            delivery_options = [
                el for el in driver.find_elements(By.XPATH, "//*[contains(text(), 'Enter Delivery Info')]")
                if el.is_displayed()
            ]
            if len(delivery_options) != 1:
                logger.error(f"SAFETY ABORT: expected 1 visible 'Enter Delivery Info' option for {po_number}, "
                             f"found {len(delivery_options)}.")
                capture_debug_screenshot(driver, logger, f"delivery_option_ambiguous_{po_number}")
                return False

            safe_click(driver, delivery_options[0], logger)
            time.sleep(2)

        logger.info(f"Opened delivery form for {po_number}.")
        return True

    except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
        logger.error(f"Could not open delivery form for {po_number}: {e}")
        return False


def select_enter_delivery_details(driver, logger) -> bool:
    """Select 'Enter Delivery Details' from the Choose Action dropdown."""
    try:
        restore_browser_if_needed(driver, logger, "delivery details selection")
        wait_and_find(driver, By.XPATH, "//*[contains(text(), 'Delivery Information')]", timeout=20, logger=logger)
        for attempt in range(3):
            time.sleep(2 if attempt == 0 else 3)

            action_selects = [
                s for s in driver.find_elements(By.XPATH, "//select[@name='actionType']")
                if s.is_displayed()
            ]
            if len(action_selects) == 1:
                if select_option_on_select(
                    action_selects[0],
                    ("Enter Delivery Details", "Enter Delivery Detail",
                     "Delivery Details", "Enter delivery details"),
                    logger,
                    "Delivery form action",
                ):
                    time.sleep(2)
                    return True
            elif len(action_selects) > 1:
                logger.error(f"SAFETY ABORT: expected 1 visible actionType select, found {len(action_selects)}.")
                capture_debug_screenshot(driver, logger, "action_type_ambiguous")
                return False

            logger.warning(f"Delivery form action select not ready yet (attempt {attempt + 1}/3).")

        # Click-based fallback
        try:
            action_el = wait_and_find(driver, By.XPATH,
                "//*[contains(text(), 'Select an Action') or contains(text(), 'Select an action')]",
                timeout=5, clickable=True)
            safe_click(driver, action_el, logger)
            time.sleep(1)
            detail_opt = wait_and_find(driver, By.XPATH,
                "//*[contains(text(), 'Enter Delivery Details') or contains(text(), 'Delivery Details')]",
                timeout=5, clickable=True)
            safe_click(driver, detail_opt, logger)
            time.sleep(2)
            return True
        except (TimeoutException, NoSuchElementException):
            pass

        logger.error("Could not select 'Enter Delivery Details'.")
        capture_debug_screenshot(driver, logger, "delivery_action_select_failed")
        return False

    except TimeoutException:
        logger.error("Delivery Information panel did not appear.")
        return False


def enter_delivery_date(driver, date_str: str, logger) -> bool:
    """Enter the delivery date using React-compatible JavaScript."""
    try:
        restore_browser_if_needed(driver, logger, "delivery date entry")
        time.sleep(2)

        date_input = None
        try:
            date_input = WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.XPATH, "//input[@name='deliveryDate']")))
        except TimeoutException:
            pass

        if not date_input:
            try:
                date_input = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='date']")))
            except TimeoutException:
                pass

        if not date_input:
            try:
                date_input = wait_and_find(driver, By.XPATH,
                    "//input[@type='text'][ancestor::*[contains(., 'Delivery Date')]]", timeout=5, logger=logger)
            except TimeoutException:
                pass

        if not date_input:
            logger.error("Could not find the Delivery Date field.")
            capture_debug_screenshot(driver, logger, "delivery_date_missing")
            return False

        day, month, year = date_str.split("-")

        if date_input.get_attribute("type") == "date":
            iso_date = f"{year}-{month}-{day}"

            # React-compatible: use native value setter + dispatch events
            driver.execute_script("""
                var el = arguments[0];
                var dateValue = arguments[1];
                var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype, 'value').set;
                nativeInputValueSetter.call(el, dateValue);
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                el.dispatchEvent(new Event('focus', { bubbles: true }));
                el.dispatchEvent(new Event('blur', { bubbles: true }));
            """, date_input, iso_date)

            time.sleep(1)
            entered = date_input.get_attribute("value")
            logger.info(f"Date entered via JS: {date_str} (field: {entered})")

            if entered != iso_date:
                # Fallback: send_keys
                logger.warning("JS date mismatch -- trying send_keys...")
                driver.execute_script("arguments[0].click();", date_input)
                time.sleep(0.3)
                date_input.send_keys(Keys.CONTROL + "a")
                time.sleep(0.1)
                date_input.send_keys(day + month + year)
                time.sleep(0.5)
                entered = date_input.get_attribute("value")
                logger.info(f"send_keys fallback: {entered}")
        else:
            date_input.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            date_input.send_keys(date_str)
            time.sleep(0.5)
            driver.execute_script("""
                var el = arguments[0];
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            """, date_input)

        time.sleep(0.5)
        return True

    except (NoSuchElementException, ElementNotInteractableException, TimeoutException) as e:
        logger.error(f"Could not fill Delivery Date: {e}")
        return False


def upload_pod_in_form(driver, pod_file: Path, logger) -> bool:
    """
    v3: Upload the POD file in the SAME delivery form.
    The form has a 'Choose File' input for Proof of Delivery.
    """
    try:
        time.sleep(1)
        logger.info(f">>> UPLOADING in delivery form: file '{pod_file.name}' <<<")
        restore_browser_if_needed(driver, logger, "delivery-form POD upload")

        delivery_panel = find_panel_root_by_heading(
            driver,
            "Delivery Information",
            logger,
            required_xpaths=[".//input[@type='file' and (@name='proofOfDelivery' or not(@name))]"],
            screenshot_prefix="delivery_panel_scope_failed",
        )
        scoped_root = delivery_panel or driver

        # Find file input on the active delivery form only
        file_inputs = _visible_descendants(
            scoped_root,
            ".//input[@type='file' and (@name='proofOfDelivery' or not(@name))]",
            include_file_inputs=True,
        )

        if file_inputs:
            file_input = next((fi for fi in file_inputs if fi.is_displayed()), file_inputs[0])
            # Make visible if hidden
            driver.execute_script(
                "arguments[0].style.display='block'; arguments[0].style.opacity=1; "
                "arguments[0].style.height='auto'; arguments[0].style.width='auto';",
                file_input)
            time.sleep(0.5)

            file_input.send_keys(str(pod_file.resolve()))
            time.sleep(1)

            try:
                selected_value = (file_input.get_attribute("value") or "").strip().lower()
                if pod_file.name.lower() in selected_value:
                    logger.info(f"POD file accepted by browser input: {pod_file.name}")
                    return True
            except (StaleElementReferenceException, WebDriverException):
                pass

            time.sleep(1)

            # Verify file was selected (check if "No file chosen" text is gone)
            try:
                no_file = scoped_root.find_elements(By.XPATH,
                    "//*[contains(text(), 'No file chosen')]")
                if no_file and any(el.is_displayed() for el in no_file):
                    logger.warning("File may not have been selected properly, retrying...")
                    file_input.send_keys(str(pod_file.resolve()))
                    time.sleep(2)
            except (NoSuchElementException, StaleElementReferenceException):
                pass

            logger.info(f"POD file selected: {pod_file.name}")
            return True
        else:
            logger.error("No file input found on the delivery form.")
            capture_debug_screenshot(driver, logger, "pod_input_missing")
            return False

    except (WebDriverException, TimeoutException, OSError) as e:
        logger.error(f"POD upload in form failed: {e}")
        return False


def click_save_changes(driver, logger) -> bool:
    """Click 'Save Changes' and verify."""
    try:
        restore_browser_if_needed(driver, logger, "save changes")
        save_buttons = [
            btn for btn in driver.find_elements(
                By.XPATH, "//button[contains(text(), 'Save Changes') or contains(text(), 'Save changes')]"
            )
            if btn.is_displayed()
        ]
        if not save_buttons:
            raise TimeoutException()

        if len(save_buttons) > 1:
            logger.warning(f"Multiple visible 'Save Changes' buttons found ({len(save_buttons)}); using the last visible one.")

        save_btn = save_buttons[-1]

        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
            time.sleep(0.5)
            save_btn.click()
        except (ElementClickInterceptedException, StaleElementReferenceException):
            driver.execute_script("arguments[0].click();", save_btn)

        logger.info("Clicked 'Save Changes' -- waiting for response...")
        time.sleep(3)

        # Check for success
        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH,
                "//*[contains(text(), 'successfully') or contains(text(), 'Success') "
                "or contains(text(), 'saved') or contains(text(), 'Updated')]")))
            logger.info("Save confirmed -- success message received!")
            time.sleep(1)
            close_modal_if_open(driver, logger)
            return True
        except TimeoutException:
            logger.info("No success toast found after save; checking if the form closed or if a real validation message appeared.")

        # If the delivery drawer/panel closed on its own, treat this as success and let verification decide the truth.
        try:
            for _ in range(5):
                panels = [
                    el for el in driver.find_elements(By.XPATH, "//*[contains(text(), 'Delivery Information')]")
                    if el.is_displayed()
                ]
                if not panels:
                    logger.info("Delivery form closed after save; proceeding to verification.")
                    return True
                time.sleep(1)
        except (TimeoutException, NoSuchElementException):
            pass

        # Only treat clearly-styled validation errors as blocking.
        try:
            error_els = driver.find_elements(By.XPATH,
                "//*[contains(@class, 'error') or contains(@class, 'Error') "
                "or contains(@class, 'text-red') or contains(@class, 'text-danger')]"
                "[contains(text(), 'error') or contains(text(), 'Error') "
                "or contains(text(), 'required') or contains(text(), 'Required') "
                "or contains(text(), 'Please upload') or contains(text(), 'Please select') "
                "or contains(text(), 'Please choose') or contains(text(), 'failed') "
                "or contains(text(), 'Failed')]")
            for el in error_els:
                if el.is_displayed() and el.text.strip():
                    logger.error(f"Save validation/error: {el.text[:200]}")
                    capture_debug_screenshot(driver, logger, "save_validation_error")
                    return False
        except (NoSuchElementException, StaleElementReferenceException):
            pass

        invalid_fields = []
        try:
            invalid_fields = driver.find_elements(By.XPATH, "//input[@aria-invalid='true'] | //select[@aria-invalid='true']")
        except (NoSuchElementException, StaleElementReferenceException):
            pass
        if invalid_fields:
            logger.error(f"Save blocked: {len(invalid_fields)} invalid field(s) still marked on the form.")
            capture_debug_screenshot(driver, logger, "save_invalid_fields")
            return False

        time.sleep(1)
        close_modal_if_open(driver, logger)
        logger.info("No blocking validation found after save; proceeding to verification.")
        return True

    except TimeoutException:
        logger.error("Could not find 'Save Changes' button.")
        return False


def verify_pod_uploaded(driver, po_number: str, logger, allow_missing_attach_as_success=False) -> bool:
    """
    Re-search the exact PO and confirm the target row now exposes a POD-view state.
    This protects against cases where a file got uploaded somewhere else on the page.
    Some Filflo states do not show a POD view action; in those cases, disappearance
    of 'Attach Proof of Delivery' from the exact row is still a strong success signal.
    """
    logger.info(f"Verifying POD upload for PO {po_number}...")
    restore_browser_if_needed(driver, logger, f"POD verification for {po_number}")
    time.sleep(2)
    close_modal_if_open(driver, logger)
    dismiss_popups(driver, logger)

    driver.get(FILFLO_URL)
    time.sleep(2)  # reduced from 4s

    try:
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20, logger=logger)
    except TimeoutException:
        logger.warning("Page didn't load after refresh for POD verification.")
        return False

    ensure_all_time_filter(driver, logger)
    time.sleep(2)

    if not type_into_search(driver, po_number, logger):
        logger.warning(f"Could not re-search PO {po_number} for POD verification.")
        return False

    time.sleep(2)  # reduced from 4s
    row = find_po_row(driver, po_number, logger, retries=5)
    if not row:
        logger.warning(f"PO {po_number} not found during POD verification.")
        return False

    options = get_row_action_options(row)
    logger.info(f"POD verification options for {po_number}: {options}")
    option_blob = " | ".join(options).lower()

    if any(key in option_blob for key in ("view pod", "view proof", "pod uploaded")):
        logger.info(f"VERIFIED -- PO {po_number} now shows POD view state.")
        return True

    if "attach proof of delivery" in option_blob:
        logger.error(f"POD verification failed -- PO {po_number} still offers 'Attach Proof of Delivery'.")
        return False

    if allow_missing_attach_as_success and options:
        logger.info(
            f"VERIFIED -- PO {po_number} no longer offers 'Attach Proof of Delivery' on the exact row."
        )
        return True

    logger.warning(f"POD verification inconclusive for {po_number}; expected POD view state not found.")
    return False


def verify_delivery_saved(driver, po_number: str, logger) -> bool:
    """
    Re-search the PO and check if status ACTUALLY changed from 'In Transit'.
    v3 FIX: Strict verification — 'unknown' status = NOT verified.
    Also does a full page refresh to get fresh data from the server.
    """
    logger.info(f"Verifying delivery for PO {po_number}...")
    restore_browser_if_needed(driver, logger, f"delivery verification for {po_number}")

    time.sleep(3)
    close_modal_if_open(driver, logger)
    dismiss_popups(driver, logger)

    # v3: Full page refresh to get fresh server data (not cached React state)
    driver.get(FILFLO_URL)
    time.sleep(2)  # reduced from 4s

    try:
        wait_and_find(driver, By.XPATH, "//input[@placeholder='Search orders...']", timeout=20, logger=logger)
    except TimeoutException:
        logger.warning("Page didn't load after refresh for verification.")
        return False

    ensure_all_time_filter(driver, logger)
    time.sleep(2)

    clear_search_field(driver, logger)
    time.sleep(2)

    if not type_into_search(driver, po_number, logger):
        logger.warning(f"Could not re-search PO {po_number} for verification.")
        return False

    time.sleep(3)  # reduced from 5s — portal refresh wait

    row = find_po_row(driver, po_number, logger, retries=5)
    if not row:
        logger.warning(f"PO {po_number} not found during verification.")
        return False

    status = check_po_status(driver, po_number, row, logger)
    logger.info(f"Verification -- PO {po_number} status: '{status}'")

    # v3 FIX: STRICT check — only pass if status is a KNOWN non-transit status
    if "in transit" in status.lower():
        logger.error(f"VERIFICATION FAILED -- PO {po_number} is STILL 'In Transit'.")
        return False

    if status == "unknown":
        logger.warning(f"VERIFICATION INCONCLUSIVE -- PO {po_number} status is 'unknown'. Treating as NOT verified.")
        return False

    # Status is a known value AND it's not "in transit" — truly verified
    logger.info(f"VERIFIED -- PO {po_number} status changed to '{status}'.")
    return True


# ═══════════════════════════════════════════════════════════════════════════════
#  POD-ONLY UPLOAD (for POs where delivery is already marked)
# ═══════════════════════════════════════════════════════════════════════════════

def do_pod_only_upload(driver, po: str, pod_file: Path, logger) -> bool:
    """
    Upload POD via the 'Attach Proof of Delivery' dropdown option.
    Used when delivery is already marked but POD hasn't been uploaded yet.
    """
    logger.info(f"[POD ONLY] Uploading POD for PO {po} -- file: {pod_file.name}")

    try:
        restore_browser_if_needed(driver, logger, f"POD flow for {po}")

        # Wait for table to stabilize before interacting
        time.sleep(2)
        _wait_for_table_stable(driver, logger, stable_seconds=2, max_wait=6)

        row = find_po_row(driver, po, logger)
        if not row:
            logger.error(f"Row for {po} not found.")
            return False

        # VERIFY: confirm a cell in this row has the EXACT PO number
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            exact_match = any(c.text.strip() == po for c in cells[:4])
            if not exact_match:
                cell_texts = [c.text.strip() for c in cells[:4]]
                logger.error(f"Row mismatch! Looking for exact '{po}' but cells contain: {cell_texts}")
                return False
            logger.debug(f"Row verified (exact cell match) for PO {po}.")
        except StaleElementReferenceException:
            row = find_po_row(driver, po, logger)
            if not row:
                return False

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
        time.sleep(0.5)

        # Click the action button on the row
        action_xpath = (".//td[last()]//button | .//td[last()]//*[contains(text(), 'Choose an option')]"
                        " | .//td[last()]//select | .//td[last()]//div[contains(@class, 'dropdown')]"
                        " | .//td[last()]//a[contains(text(), 'Choose')]")

        for find_attempt in range(3):
            try:
                action_btn = row.find_element(By.XPATH, action_xpath)
                break
            except (NoSuchElementException, StaleElementReferenceException):
                row = find_po_row(driver, po, logger)
                if not row:
                    return False
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
                time.sleep(0.5)
                if find_attempt == 2:
                    logger.error(f"Could not find action button for {po}.")
                    return False

        # Step 1 & 2: choose "Attach Proof of Delivery" ONLY on the exact matched row
        # Retry up to 3 times because the dropdown/select can sometimes fail to register.
        attach_success = False
        for dropdown_attempt in range(3):
            logger.info(f"[POD ONLY] Step 1: Opening row action for {po} (attempt {dropdown_attempt+1}/3)...")

            # Re-find the row and button if retrying
            if dropdown_attempt > 0:
                time.sleep(2)
                # Click elsewhere to close any stale dropdown
                try:
                    driver.find_element(By.TAG_NAME, "body").click()
                except (NoSuchElementException, ElementNotInteractableException):
                    pass
                time.sleep(1)

                row = find_po_row(driver, po, logger)
                if not row:
                    logger.error(f"Row for {po} not found on retry.")
                    return False
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
                time.sleep(0.5)
                try:
                    action_btn = row.find_element(By.XPATH, action_xpath)
                except (NoSuchElementException, StaleElementReferenceException):
                    continue

            # SAFETY FIX: choose the option on this row's own <select>, not from a page-global text match.
            if action_btn.tag_name.lower() == "select":
                if select_option_on_select(
                    action_btn,
                    ("Attach Proof of Delivery",),
                    logger,
                    f"[POD ONLY] Row action for {po}",
                ):
                    time.sleep(3)
                    logger.info(f"[POD ONLY] Step 2: Selected 'Attach Proof of Delivery' for {po}.")
                    attach_success = True
                    break
            else:
                safe_click(driver, action_btn, logger)
                time.sleep(2)

                attach_options = [
                    el for el in driver.find_elements(By.XPATH, "//*[contains(text(), 'Attach Proof of Delivery')]")
                    if el.is_displayed()
                ]
                if len(attach_options) == 1:
                    safe_click(driver, attach_options[0], logger)
                    time.sleep(3)
                    logger.info(f"[POD ONLY] Step 2: Clicked 'Attach Proof of Delivery' for {po}.")
                    attach_success = True
                    break

                logger.warning(f"[POD ONLY] Attach POD option ambiguous/not found (attempt {dropdown_attempt+1}/3, "
                               f"visible options={len(attach_options)}).")

                # Check if POD already uploaded
                try:
                    view_pod = driver.find_elements(By.XPATH,
                        "//*[contains(text(), 'View POD') or contains(text(), 'View Proof') "
                        "or contains(text(), 'POD Uploaded')]")
                    if view_pod and any(el.is_displayed() for el in view_pod):
                        logger.info(f"POD already uploaded for {po}.")
                        try:
                            driver.find_element(By.TAG_NAME, "body").click()
                        except (NoSuchElementException, StaleElementReferenceException):
                            pass
                        return True
                except (NoSuchElementException, ElementNotInteractableException):
                    pass

        if not attach_success:
            # Log visible texts for debugging
            try:
                all_visible = driver.find_elements(By.XPATH, "//*[text()]")
                visible_texts = [el.text.strip() for el in all_visible
                                if el.is_displayed() and el.text.strip()
                                and len(el.text.strip()) < 50][:15]
                logger.error(f"[POD ONLY] FAILED: 'Attach Proof of Delivery' never appeared for {po}. "
                            f"Visible texts: {visible_texts}")
            except (NoSuchElementException, StaleElementReferenceException):
                logger.error(f"[POD ONLY] FAILED: 'Attach Proof of Delivery' never appeared for {po}.")
            try:
                driver.find_element(By.TAG_NAME, "body").click()
            except (NoSuchElementException, ElementNotInteractableException):
                pass
            time.sleep(1)
            capture_debug_screenshot(driver, logger, f"attach_pod_failed_{po}")
            close_modal_if_open(driver, logger)
            return False

        # Wait for "Upload Proof of Delivery" panel to fully load
        try:
            wait_and_find(driver, By.XPATH,
                "//*[contains(text(), 'Upload Proof of Delivery')]", timeout=10)
            logger.info("Upload POD panel loaded.")
            time.sleep(2)
        except TimeoutException:
            logger.warning("Upload POD panel title not found, proceeding anyway...")
            time.sleep(2)

        # SAFETY: Verify search bar still shows the correct PO before uploading
        try:
            si = driver.find_element(By.XPATH, "//input[@placeholder='Search orders...']")
            search_val = (si.get_attribute("value") or "").strip()
            if search_val and po not in search_val and search_val not in po:
                logger.error(f"SAFETY ABORT: Search bar shows '{search_val}' but we need PO '{po}'. Page may have changed!")
                close_modal_if_open(driver, logger)
                return False
        except (NoSuchElementException, StaleElementReferenceException):
            pass

        upload_panel = find_panel_root_by_heading(
            driver,
            "Upload Proof of Delivery",
            logger,
            required_xpaths=[
                ".//input[@type='file' and (@name='proofOfDelivery' or not(@name))]",
                ".//button[contains(translate(normalize-space(text()), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'upload')]",
            ],
            screenshot_prefix=f"pod_panel_scope_failed_{po}",
        )
        if not upload_panel:
            close_modal_if_open(driver, logger)
            return False

        # Upload the file
        logger.info(f"[POD ONLY] Step 3: Uploading file '{pod_file.name}' for PO '{po}'...")
        restore_browser_if_needed(driver, logger, f"file upload for {po}")

        file_inputs = _visible_descendants(
            upload_panel,
            ".//input[@type='file' and (@name='proofOfDelivery' or not(@name))]",
            include_file_inputs=True,
        )
        logger.info(f"[POD ONLY] Step 3: Found {len(file_inputs)} file input(s) on page.")
        if file_inputs:
            if len(file_inputs) > 1:
                logger.warning(f"[POD ONLY] Found {len(file_inputs)} file inputs inside the upload panel; using the first scoped one.")
            file_input = file_inputs[0]
            driver.execute_script(
                "arguments[0].style.display='block'; arguments[0].style.opacity=1; "
                "arguments[0].style.height='auto'; arguments[0].style.width='auto';",
                file_input)
            time.sleep(0.5)
            file_input.send_keys(str(pod_file.resolve()))

            # Wait for file to be confirmed by the UI (up to 15 seconds)
            file_confirmed = False
            try:
                selected_value = (file_input.get_attribute("value") or "").strip().lower()
                if pod_file.name.lower() in selected_value:
                    logger.info(f"[POD ONLY] Browser input accepted file: {pod_file.name}")
                    file_confirmed = True
                else:
                    logger.debug(f"[POD ONLY] File input value after send_keys: '{selected_value}'")
            except (StaleElementReferenceException, WebDriverException):
                pass

            for wait_attempt in range(15):
                if file_confirmed:
                    break
                time.sleep(1)
                try:
                    selected_els = _visible_descendants(
                        upload_panel,
                        ".//*[contains(text(), 'Selected:') or contains(text(), 'selected') "
                        "or contains(text(), 'chosen') or contains(text(), 'Chosen')]",
                    )
                    for el in selected_els:
                        logger.info(f"POD file CONFIRMED selected: {el.text.strip()}")
                        file_confirmed = True
                        break
                except (NoSuchElementException, StaleElementReferenceException):
                    pass
                if file_confirmed:
                    break

                # Also check if a filename/thumbnail appeared (alternative confirmation)
                try:
                    name_els = _visible_descendants(
                        upload_panel,
                        f".//*[contains(text(), '{pod_file.stem}')]",
                    )
                    for el in name_els:
                        if el.tag_name not in ("input",):
                            logger.info(f"POD file name appeared in UI: {el.text.strip()}")
                            file_confirmed = True
                            break
                except (NoSuchElementException, StaleElementReferenceException):
                    pass
                if file_confirmed:
                    break

            if not file_confirmed:
                logger.warning(f"POD file sent but no UI confirmation after 15s: {pod_file.name}")
                # Give one final extra wait before proceeding
                time.sleep(3)
            else:
                # Small extra wait after confirmation for button to enable
                time.sleep(1)
        else:
            logger.error("No file input found for POD upload.")
            return False

        # Click the "Upload" button
        logger.info(f"[POD ONLY] Step 4: Clicking 'Upload' button for PO '{po}'...")
        time.sleep(1)
        restore_browser_if_needed(driver, logger, f"upload confirmation click for {po}")
        upload_clicked = False

        # First try: find the specific blue "Upload" button on the panel
        try:
            upload_btns = _visible_descendants(
                upload_panel,
                ".//button[normalize-space(text())='Upload']",
            )
            upload_btn = upload_btns[0] if upload_btns else None
            if upload_btn.is_displayed():
                safe_click(driver, upload_btn, logger)
                logger.info("Clicked 'Upload' button.")
                upload_clicked = True
                time.sleep(3)
        except (NoSuchElementException, WebDriverException):
            pass

        # Fallback: try other button labels
        if not upload_clicked:
            for btn_text in ["Upload", "Submit", "Save", "Confirm", "OK", "Done"]:
                btns = _visible_descendants(
                    upload_panel,
                    f".//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{btn_text.lower()}')]",
                )
                if btns:
                    if btns:
                        safe_click(driver, btns[0], logger)
                        logger.info(f"Clicked '{btn_text}' button (fallback).")
                        upload_clicked = True
                        time.sleep(3)
                        break

        if not upload_clicked:
            logger.error("Could not find Upload button.")
            return False

        # Check for success
        logger.info(f"[POD ONLY] Step 5: Checking for success confirmation for PO '{po}'...")
        time.sleep(3)
        success_found = False
        explicit_error_found = False
        try:
            for keyword in ["uploaded successfully", "proof of delivery uploaded",
                             "successfully", "uploaded", "saved", "attached", "updated"]:
                els = _visible_descendants(
                    upload_panel,
                    f".//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{keyword}')]",
                )
                visible_els = [el for el in els if el.is_displayed()]
                if visible_els:
                    logger.info(f"[POD ONLY] Upload confirmed for {po} -- indicator: '{keyword}'")
                    success_found = True
                    break
        except StaleElementReferenceException:
            logger.info(f"[POD ONLY] Upload panel refreshed for {po} after clicking upload; continuing to verification.")
            success_found = True

        if not success_found:
            # Toasts sometimes render outside the modal; allow a scoped-global fallback.
            toast_els = [
                el for el in driver.find_elements(
                    By.XPATH,
                    "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'uploaded successfully') "
                    "or contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'proof of delivery uploaded') "
                    "or contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'attached')]"
                )
                if el.is_displayed()
            ]
            if toast_els:
                logger.info(f"[POD ONLY] Upload confirmed for {po} via toast/global success message.")
                success_found = True

        # Fallback: if the upload dialog/modal closed on its own, that usually means success
        if not success_found:
            time.sleep(2)
            backdrops = driver.find_elements(By.XPATH,
                "//div[contains(@class, 'MuiBackdrop') or contains(@class, 'MuiModal-backdrop')]")
            modal_still_open = any(b.is_displayed() for b in backdrops) if backdrops else False

            if not modal_still_open:
                logger.info(f"[POD ONLY] Upload dialog closed for {po} -- treating as success.")
                success_found = True
            else:
                # Check for error messages in the modal
                try:
                    error_els = driver.find_elements(By.XPATH,
                        "//*[contains(text(), 'error') or contains(text(), 'Error') "
                        "or contains(text(), 'failed') or contains(text(), 'Failed')]")
                    for el in error_els:
                        if el.is_displayed():
                            logger.error(f"[POD ONLY] Error found for {po}: {el.text[:200]}")
                            explicit_error_found = True
                            close_modal_if_open(driver, logger)
                            return False
                except (NoSuchElementException, StaleElementReferenceException):
                    pass

        close_modal_if_open(driver, logger)

        if explicit_error_found:
            return False

        verified = verify_pod_uploaded(driver, po, logger, allow_missing_attach_as_success=True)
        if verified:
            return True

        if not success_found:
            logger.warning(f"[POD ONLY] No success confirmation for {po}, and post-upload verification also failed.")
            return False

        capture_debug_screenshot(driver, logger, f"pod_verify_failed_{po}")
        return False

    except (WebDriverException, TimeoutException) as e:
        logger.error(f"[POD ONLY] Failed for {po}: {e}")
        close_modal_if_open(driver, logger)
        return False


# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESS ONE PO — AUTO-DETECTS WHAT TASK TO DO
# ═══════════════════════════════════════════════════════════════════════════════

def process_single_po(driver, entry: dict, excel_path: Path, logger) -> str:
    """
    v3: Smart processing — automatically decides what to do:

    Scenario A: Delivery not done + date available → Mark delivery (+ POD if available)
    Scenario B: Delivery done + POD not done + POD file available → Upload POD only
    Scenario C: Everything done → Skip
    """
    set_correlation_id(entry.get("po_number", "UNKNOWN"))
    po = entry["po_number"]
    row_idx = entry["row"]
    current_status = entry["status"]

    # ── Determine what's already done ────────────────────────────────────────
    delivery_already_done = any(k in current_status for k in (
        "VERIFIED - Delivery", "Delivery Done", "Delivery + POD Done"
    ))
    pod_already_done = any(k in current_status for k in (
        "POD Done", "Delivery + POD Done", "POD Uploaded"
    ))

    # Fully done? Skip.
    if delivery_already_done and pod_already_done:
        logger.info(f"PO {po}: Both delivery and POD done -- skipping.")
        return POStatus.ALREADY_DONE.value

    # Parse delivery date
    delivery_date = None
    if entry["delivery_date"]:
        try:
            delivery_date = parse_delivery_date(entry["delivery_date"])
        except (ValueError, TypeError) as e:
            logger.warning(f"PO {po}: Bad delivery date -- {e}")
    prefer_all_time = should_prefer_all_time(entry["delivery_date"])

    # Find POD file
    pod_file = find_pod_file(POD_FOLDER, entry["tracking_id"], logger) if entry["tracking_id"] else None

    # ══════════════════════════════════════════════════════════════════════════
    #  SCENARIO P: POD-only mode — Excel has PO + Tracking ID but NO date
    #  Directly upload POD using PO and Tracking ID from the Excel file,
    #  regardless of portal status or prior bot status.
    # ══════════════════════════════════════════════════════════════════════════
    if not delivery_date and entry["tracking_id"] and pod_file:
        logger.info(f"[POD DIRECT] PO {po}: No delivery date, but Tracking ID "
                     f"'{entry['tracking_id']}' and POD file '{pod_file.name}' found. "
                     f"Proceeding with direct POD upload.")

        # SAFETY: Fresh page load to eliminate ALL stale state
        if not is_session_valid(driver, logger):
            if not login(driver, logger):
                update_excel_status(excel_path, row_idx, "FAILED - Login", logger)
                return "FAILED - Login"

        if not fresh_start_and_search(driver, po, logger, prefer_all_time=prefer_all_time):
            final = POStatus.FAILED_PO_NOT_FOUND_POD.value
            logger.warning(f"PO {po}: Not found on portal for POD upload.")
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        if do_pod_only_upload(driver, po, pod_file, logger):
            # Move file to _uploaded
            try:
                POD_DONE_FOLDER.mkdir(parents=True, exist_ok=True)
                dest = POD_DONE_FOLDER / pod_file.name
                if not dest.exists():
                    shutil.move(str(pod_file), str(dest))
                    logger.info(f"Moved {pod_file.name} -> _uploaded/")
            except (OSError, PermissionError) as e:
                logger.warning(f"Could not move file: {e}")

            final = POStatus.VERIFIED_POD_UPLOADED.value
            logger.info(f"=== PO {po} -- {final} ===")
            update_excel_status(excel_path, row_idx, final, logger)
            return final
        else:
            final = POStatus.FAILED_POD_UPLOAD.value
            logger.warning(f"=== PO {po} -- {final} ===")
            update_excel_status(excel_path, row_idx, final, logger)
            return final

    # ══════════════════════════════════════════════════════════════════════════
    #  SCENARIO B: Delivery done, POD pending — do POD-only upload
    # ══════════════════════════════════════════════════════════════════════════
    if delivery_already_done and not pod_already_done:
        if not pod_file:
            logger.info(f"PO {po}: Delivery done, but no POD file available yet. Skipping POD.")
            return POStatus.ALREADY_DONE.value  # Don't overwrite the delivery status

        # SAFETY: Fresh page load to eliminate ALL stale state
        if not is_session_valid(driver, logger):
            if not login(driver, logger):
                return "FAILED - Login"

        if not fresh_start_and_search(driver, po, logger, prefer_all_time=prefer_all_time):
            logger.warning(f"PO {po}: Not found for POD upload.")
            return "FAILED - PO Not Found for POD"

        if do_pod_only_upload(driver, po, pod_file, logger):
            # Move file
            try:
                POD_DONE_FOLDER.mkdir(parents=True, exist_ok=True)
                dest = POD_DONE_FOLDER / pod_file.name
                if not dest.exists():
                    shutil.move(str(pod_file), str(dest))
                    logger.info(f"Moved {pod_file.name} -> _uploaded/")
            except (OSError, PermissionError) as e:
                logger.warning(f"Could not move file: {e}")

            final = POStatus.VERIFIED_DELIVERY_AND_POD.value
            logger.info(f"=== PO {po} -- {final} ===")
            update_excel_status(excel_path, row_idx, final, logger)
            return final
        else:
            final = POStatus.VERIFIED_DELIVERY_POD_FAIL.value
            logger.warning(f"=== PO {po} -- {final} ===")
            update_excel_status(excel_path, row_idx, final, logger)
            return final

    # ══════════════════════════════════════════════════════════════════════════
    #  SCENARIO D: Delivery-only mode — Excel has PO + Delivery Date but NO Tracking ID
    #  Directly fill the delivery form using PO and Date from the Excel file,
    #  regardless of portal status.
    # ══════════════════════════════════════════════════════════════════════════
    if delivery_date and not entry["tracking_id"]:
        logger.info(f"[DELIVERY DIRECT] PO {po}: Delivery date '{delivery_date}' provided, "
                     f"no Tracking ID. Proceeding with direct delivery entry.")

        # SAFETY: Fresh page load
        if not is_session_valid(driver, logger):
            if not login(driver, logger):
                update_excel_status(excel_path, row_idx, "FAILED - Login", logger)
                return "FAILED - Login"

        if not fresh_start_and_search(driver, po, logger, prefer_all_time=prefer_all_time):
            final = POStatus.FAILED_PO_NOT_FOUND.value
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        current_row = find_po_row(driver, po, logger)
        current_portal_status = check_po_status(driver, po, current_row, logger) if current_row else "unknown"
        if current_portal_status in ("delivered", "grn entered"):
            final = POStatus.VERIFIED_DELIVERY_PORTAL.value
            logger.info(f"PO {po}: Portal already shows status '{current_portal_status}'.")
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        # Open delivery form directly — no portal status check
        if not click_enter_delivery_info(driver, po, logger):
            close_modal_if_open(driver, logger)
            final = POStatus.FAILED_DELIVERY_FORM.value
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        if not select_enter_delivery_details(driver, logger):
            close_modal_if_open(driver, logger)
            final = POStatus.FAILED_ACTION_SELECT.value
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        if not enter_delivery_date(driver, delivery_date, logger):
            close_modal_if_open(driver, logger)
            final = POStatus.FAILED_DATE_ENTRY.value
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        if not click_save_changes(driver, logger):
            close_modal_if_open(driver, logger)
            final = POStatus.FAILED_SAVE.value
            update_excel_status(excel_path, row_idx, final, logger)
            return final

        if verify_delivery_saved(driver, po, logger):
            final = POStatus.VERIFIED_DELIVERY.value
            logger.info(f"=== PO {po} -- {final} ===")
        else:
            final = POStatus.UNVERIFIED_STATUS_UNCHANGED.value
            logger.warning(f"=== PO {po} -- {final} ===")

        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # ══════════════════════════════════════════════════════════════════════════
    #  SCENARIO A: Delivery not done — mark delivery (+ POD if available)
    #  Used when BOTH delivery date AND tracking ID are provided.
    # ══════════════════════════════════════════════════════════════════════════
    if not delivery_date:
        final = "SKIPPED - No delivery date and no tracking ID"
        logger.warning(f"PO {po}: Missing both delivery date and tracking ID.")
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    if pod_file:
        logger.info(f"[DELIVERY + POD] PO {po}: date={delivery_date}, "
                     f"Tracking ID='{entry['tracking_id']}', POD file='{pod_file.name}'")
    else:
        logger.info(f"[DELIVERY ONLY] PO {po}: date={delivery_date}, "
                     f"Tracking ID='{entry['tracking_id']}', no matching POD file in POD_FILES")

    # SAFETY: Fresh page load to eliminate ALL stale state from previous PO
    if not is_session_valid(driver, logger):
        if not login(driver, logger):
            update_excel_status(excel_path, row_idx, "FAILED - Login", logger)
            return "FAILED - Login"

    if not fresh_start_and_search(driver, po, logger, prefer_all_time=prefer_all_time):
        final = "SKIPPED - PO Not Found"
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    current_row = find_po_row(driver, po, logger)
    current_portal_status = check_po_status(driver, po, current_row, logger) if current_row else "unknown"
    if current_portal_status in ("delivered", "grn entered"):
        final = POStatus.VERIFIED_DELIVERY_PORTAL.value
        logger.info(f"PO {po}: Portal already shows status '{current_portal_status}'.")
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # Open delivery form
    if not click_enter_delivery_info(driver, po, logger):
        close_modal_if_open(driver, logger)
        final = POStatus.FAILED_DELIVERY_FORM.value
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # Select "Enter Delivery Details"
    if not select_enter_delivery_details(driver, logger):
        close_modal_if_open(driver, logger)
        final = POStatus.FAILED_ACTION_SELECT.value
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # Fill delivery date
    if not enter_delivery_date(driver, delivery_date, logger):
        close_modal_if_open(driver, logger)
        final = POStatus.FAILED_DATE_ENTRY.value
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # Upload POD file (optional)
    pod_uploaded = False
    if pod_file:
        if upload_pod_in_form(driver, pod_file, logger):
            pod_uploaded = True
        else:
            logger.warning(f"PO {po}: POD upload failed, proceeding with date only.")
    else:
        logger.info(f"PO {po}: No POD file -- submitting delivery date only.")

    # Click Save
    if not click_save_changes(driver, logger):
        close_modal_if_open(driver, logger)
        final = POStatus.FAILED_SAVE.value
        update_excel_status(excel_path, row_idx, final, logger)
        return final

    # Verify
    if verify_delivery_saved(driver, po, logger):
        if pod_uploaded and pod_file:
            try:
                POD_DONE_FOLDER.mkdir(parents=True, exist_ok=True)
                dest = POD_DONE_FOLDER / pod_file.name
                if not dest.exists():
                    shutil.move(str(pod_file), str(dest))
                    logger.info(f"Moved {pod_file.name} -> _uploaded/")
            except (OSError, PermissionError) as e:
                logger.warning(f"Could not move file: {e}")

        if pod_uploaded:
            final = POStatus.VERIFIED_DELIVERY_AND_POD.value
        else:
            final = POStatus.VERIFIED_DELIVERY_ONLY.value
        logger.info(f"=== PO {po} -- {final} ===")
        update_excel_status(excel_path, row_idx, final, logger)
        return final
    else:
        final = POStatus.UNVERIFIED_STATUS_UNCHANGED.value
        logger.warning(f"=== PO {po} -- {final} ===")
        update_excel_status(excel_path, row_idx, final, logger)
        return final


def process_single_po_with_retry(driver, entry: dict, excel_path: Path, logger) -> str:
    """Retry wrapper."""
    for attempt in range(MAX_PO_RETRIES):
        result = process_single_po(driver, entry, excel_path, logger)

        # Don't retry these
        if result.startswith("VERIFIED") or result.startswith("SKIPPED") or result == "ALREADY_DONE":
            return result

        # Retry failures and unverified
        if attempt < MAX_PO_RETRIES - 1:
            logger.info(f"Retrying PO {entry['po_number']} ({attempt + 2}/{MAX_PO_RETRIES})...")
            time.sleep(3)

    return result


def record_monitor_outcome(monitor, po_number: str, result: str, duration_sec: float = 0.0):
    """
    Best-effort hook for BotMonitor integration.
    Keeps monitoring optional so legacy callers continue to work unchanged.
    """
    if not monitor:
        return

    normalized = str(result or "").strip()
    upper = normalized.upper()

    try:
        if upper.startswith("VERIFIED"):
            monitor.record(po_number, True, duration_sec, category="verified")
        elif upper.startswith("SKIPPED") or normalized == "ALREADY_DONE":
            monitor.record(po_number, False, duration_sec, error=normalized, category="skipped")
        else:
            monitor.record(po_number, False, duration_sec, error=normalized, category="failed")
    except Exception as e:
        try:
            logger = logging.getLogger("combined_bot")
            logger.warning(f"[Monitor] Could not record outcome for PO {po_number}: {e}")
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESS ALL
# ═══════════════════════════════════════════════════════════════════════════════

def process_all_entries(excel_path: Path, logger, monitor=None) -> dict:
    ensure_excel_headers(excel_path, logger)
    entries = read_pending_entries(excel_path, logger)

    if not entries:
        logger.info("No pending POs to process.")
        return {"total": 0, "verified": 0, "unverified": 0, "skipped": 0, "failed": 0}

    logger.info(f"Found {len(entries)} PO(s) to process.")

    # Check readiness of each entry
    ready_delivery_pod = 0
    ready_delivery_only = 0
    ready_pod_only = 0
    for e in entries:
        has_date = bool(e["delivery_date"])
        has_tracking = bool(e["tracking_id"])
        has_pod = has_tracking and find_pod_file(POD_FOLDER, e["tracking_id"]) is not None
        if has_date and has_pod:
            ready_delivery_pod += 1
        elif has_date and not has_tracking:
            ready_delivery_only += 1
        elif not has_date and has_pod:
            ready_pod_only += 1

    logger.info(f"  {ready_delivery_pod} PO(s) have delivery date + POD file (delivery + POD).")
    logger.info(f"  {ready_delivery_only} PO(s) have delivery date only, no tracking ID (direct delivery entry).")
    logger.info(f"  {ready_pod_only} PO(s) have POD file only, no delivery date (direct POD upload).")
    logger.info(f"  {len(entries) - ready_delivery_pod - ready_delivery_only - ready_pod_only} PO(s) not ready.")

    driver = None
    summary = {"total": len(entries), "verified": 0, "unverified": 0, "skipped": 0, "failed": 0}
    results_log = []

    try:
        driver = create_driver(logger)

        if not login(driver, logger):
            logger.error("Login failed.")
            summary["failed"] = len(entries)
            return summary

        apply_all_time_filter(driver, logger)

        for i, entry in enumerate(entries, 1):
            po = entry["po_number"]
            logger.info(f"\n{'=' * 60}")
            logger.info(f"Processing {i}/{len(entries)}: PO {po}")
            logger.info(f"{'=' * 60}")
            started = time.perf_counter()

            try:
                result = process_single_po_with_retry(driver, entry, excel_path, logger)
            except (WebDriverException, TimeoutException, NoSuchElementException, OSError) as e:
                logger.error(f"PO {po}: runtime error during processing -- {e}", exc_info=True)
                result = "FAILED - Driver/Runtime Error"
                update_excel_status(excel_path, entry["row"], result, logger)

                # Recover browser/session and continue with the remaining rows.
                try:
                    if driver:
                        driver.quit()
                except WebDriverException:
                    pass
                driver = create_driver(logger)
                if not login(driver, logger):
                    logger.error(f"PO {po}: could not re-login after recovery.")
            finally:
                duration_sec = time.perf_counter() - started
                record_monitor_outcome(monitor, po, result, duration_sec)

            if "VERIFIED" in result:
                summary["verified"] += 1
            elif "UNVERIFIED" in result:
                summary["unverified"] += 1
            elif "SKIPPED" in result or result == "ALREADY_DONE":
                summary["skipped"] += 1
            else:
                summary["failed"] += 1

            results_log.append((po, result))

            if i < len(entries):
                time.sleep(2)

    except Exception as e:
        logger.error(f"Critical error: {e}", exc_info=True)
    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed.")
            except WebDriverException:
                pass

    # Summary
    logger.info(f"\n{'=' * 60}")
    logger.info("FINAL RESULTS")
    logger.info(f"{'=' * 60}")
    for po, result in results_log:
        logger.info(f"  {po:<40} -> {result}")
    logger.info(f"{'=' * 60}")
    logger.info(f"  Verified: {summary['verified']}  |  Unverified: {summary['unverified']}  "
                f"|  Skipped: {summary['skipped']}  |  Failed: {summary['failed']}")
    logger.info(f"{'=' * 60}")

    return summary


# ═══════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD ORDER-WISE DATA (Custom Filter — Last 1 Year)
# ═══════════════════════════════════════════════════════════════════════════════

def create_download_driver(logger) -> webdriver.Chrome:
    """Create a Chrome driver with download directory set to BOT_FOLDER."""
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-features=CalculateNativeWinOcclusion")
    options.page_load_strategy = "normal"
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "password_manager.leak_detection": False,
        "download.default_directory": str(DOWNLOAD_FOLDER),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    })
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    logger.info("Launching Chrome browser (download mode)...")
    try:
        driver = webdriver.Chrome(options=options)
    except WebDriverException as e:
        logger.warning(f"Standard launch failed ({e}), trying with Service...")
        driver = webdriver.Chrome(service=Service(), options=options)

    try:
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    except WebDriverException:
        pass
    return driver


def set_date_input_value(driver, input_element, date_str: str, logger):
    """
    Set a date input field value using React-compatible JS.
    date_str should be in MM/DD/YYYY format.
    """
    restore_browser_if_needed(driver, logger, f"setting date {date_str}")
    driver.execute_script("""
        var el = arguments[0];
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(el, arguments[1]);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
    """, input_element, date_str)
    time.sleep(0.5)


def apply_custom_date_filter(driver, from_date: str, to_date: str, logger) -> bool:
    """
    Apply 'Custom Range' date filter on the B2B Orders page.
    from_date and to_date in MM/DD/YYYY format.
    """
    logger.info(f"Setting custom date filter: {from_date} to {to_date}")

    try:
        restore_browser_if_needed(driver, logger, "custom date filter")
        # Step 1: Click the date filter dropdown (currently shows "Punched Date" area)
        # Look for "Custom Range" button or the date picker area
        custom_range_btn = None
        try:
            custom_range_btn = wait_and_find(driver, By.XPATH,
                "//*[contains(text(), 'Custom Range') or contains(text(), 'custom range')]",
                timeout=5, clickable=True, logger=logger)
        except TimeoutException:
            # Try clicking the date range area to reveal options
            try:
                date_area = driver.find_element(By.XPATH,
                    "//*[contains(text(), 'Last 30 Days') or contains(text(), 'Last 30 days') "
                    "or contains(text(), 'All Time') or contains(text(), 'All time')]")
                safe_click(driver, date_area, logger)
                time.sleep(1)
                custom_range_btn = wait_and_find(driver, By.XPATH,
                    "//*[contains(text(), 'Custom Range') or contains(text(), 'custom range')]",
                    timeout=5, clickable=True, logger=logger)
            except (TimeoutException, NoSuchElementException):
                pass

        if custom_range_btn:
            safe_click(driver, custom_range_btn, logger)
            time.sleep(2)
            logger.info("Clicked 'Custom Range'.")

        # Step 2: Find the date input fields and set values
        # The screenshot shows two date inputs: "from" and "to"
        date_inputs = driver.find_elements(By.XPATH,
            "//input[@type='date' or @type='text'][contains(@placeholder, 'date') "
            "or contains(@placeholder, 'Date') or contains(@placeholder, 'MM') "
            "or contains(@placeholder, 'mm') or contains(@placeholder, 'dd')]")

        if len(date_inputs) < 2:
            # Fallback: find by position — the date inputs near "Custom Range"
            date_inputs = driver.find_elements(By.CSS_SELECTOR,
                "input[type='date']")

        if len(date_inputs) < 2:
            # Another fallback: look for inputs that look like date fields by their value format
            all_inputs = driver.find_elements(By.TAG_NAME, "input")
            date_inputs = []
            for inp in all_inputs:
                val = (inp.get_attribute("value") or "").strip()
                inp_type = (inp.get_attribute("type") or "").lower()
                if inp_type == "date" or (val and "/" in val and len(val) >= 8):
                    if inp.is_displayed():
                        date_inputs.append(inp)

        if len(date_inputs) >= 2:
            logger.info(f"Found {len(date_inputs)} date input(s).")

            # Set FROM date
            set_date_input_value(driver, date_inputs[0], from_date, logger)
            logger.info(f"Set FROM date: {from_date}")

            # Set TO date
            set_date_input_value(driver, date_inputs[1], to_date, logger)
            logger.info(f"Set TO date: {to_date}")

            time.sleep(1)

            # Close any open calendar popup by clicking elsewhere
            try:
                body = driver.find_element(By.TAG_NAME, "body")
                body.click()
                time.sleep(1)
            except (NoSuchElementException, ElementNotInteractableException):
                pass

            # Press Escape to dismiss any remaining popups/calendars
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(1)

            # Some portals require pressing Enter or clicking an Apply button
            try:
                apply_btn = driver.find_element(By.XPATH,
                    "//button[contains(text(), 'Apply') or contains(text(), 'Go') "
                    "or contains(text(), 'Filter') or contains(text(), 'Search')]")
                if apply_btn.is_displayed():
                    safe_click(driver, apply_btn, logger)
                    time.sleep(3)
                    logger.info("Clicked Apply/Filter button.")
            except NoSuchElementException:
                # Press Enter on the last date input to trigger filter
                date_inputs[1].send_keys(Keys.ENTER)
                time.sleep(3)

            # Wait for table to fully reload with data
            logger.info("Waiting for table data to load...")
            for wait_sec in range(30):
                time.sleep(1)
                try:
                    rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
                    if len(rows) > 0:
                        logger.info(f"Table loaded with {len(rows)} row(s) after {wait_sec + 1}s.")
                        break
                except (NoSuchElementException, StaleElementReferenceException):
                    pass
            else:
                logger.warning("Table may not have loaded fully, proceeding anyway...")

            # Extra wait for the page to stabilize after data loads
            time.sleep(5)
            logger.info("Custom date filter applied successfully.")
            return True
        else:
            logger.error(f"Could not find date input fields. Found {len(date_inputs)} inputs.")
            return False

    except (WebDriverException, TimeoutException) as e:
        logger.error(f"Failed to apply custom date filter: {e}")
        return False


def click_download_order_wise(driver, logger) -> bool:
    """Click ONLY the 'Download Order-wise' button (not 'Download SKU-wise')."""
    logger.info("Looking for 'Download Order-wise' button...")
    try:
        restore_browser_if_needed(driver, logger, "download button selection")

        # Wait up to 30s for buttons to exit "Downloading..." spinner state
        for _wait in range(30):
            try:
                btns = driver.find_elements(By.TAG_NAME, "button")
                btn_texts = [b.text.strip().lower() for b in btns if b.is_displayed()]
                if any("order" in t and "download" in t for t in btn_texts):
                    break  # Buttons are ready
                if any("downloading" in t for t in btn_texts):
                    if _wait == 0:
                        logger.info("Buttons still in 'Downloading...' state, waiting...")
                    time.sleep(1)
                    continue
                break  # No downloading state, proceed
            except WebDriverException:
                time.sleep(1)
        else:
            logger.warning("Buttons stuck in 'Downloading...' after 30s.")

        # -- Strategy 1: XPath targeting descendant text with Order, not SKU --
        try:
            candidates = driver.find_elements(By.XPATH,
                "//button[.//text()[contains(., 'Order')] "
                "and not(.//text()[contains(., 'SKU')])]")
            for btn in candidates:
                if btn.is_displayed():
                    btn_text = btn.text.strip()
                    logger.info(f"Strategy 1 found button: '{btn_text}'")
                    safe_click(driver, btn, logger)
                    logger.info(f"Clicked 'Download Order-wise': '{btn_text}'")
                    return True
        except (NoSuchElementException, TimeoutException):
            pass

        # -- Strategy 2: Scan ALL buttons, log each, pick by text -----------
        buttons = driver.find_elements(By.TAG_NAME, "button")
        logger.info(f"Strategy 2: scanning {len(buttons)} buttons...")
        order_btn = None
        for btn in buttons:
            if not btn.is_displayed():
                continue
            btn_text = btn.text.strip().lower()
            # Log all download-related buttons for debugging
            if "download" in btn_text or "order" in btn_text or "sku" in btn_text:
                logger.info(f"  Button: '{btn.text.strip()}'")
            if "order" in btn_text and "sku" not in btn_text:
                order_btn = btn

        if order_btn:
            logger.info(f"Strategy 2 matched: '{order_btn.text.strip()}'")
            safe_click(driver, order_btn, logger)
            logger.info(f"Clicked 'Download Order-wise': '{order_btn.text.strip()}'")
            return True

        # -- Strategy 3: JavaScript innerText match -------------------------
        logger.info("Strategy 3: using JavaScript innerText match...")
        clicked = driver.execute_script("""
            var buttons = document.querySelectorAll('button');
            for (var i = 0; i < buttons.length; i++) {
                var txt = buttons[i].innerText || '';
                if (txt.indexOf('Order') !== -1 && txt.indexOf('SKU') === -1
                    && txt.toLowerCase().indexOf('download') !== -1) {
                    buttons[i].click();
                    return txt.trim();
                }
            }
            return null;
        """)
        if clicked:
            logger.info(f"JS-clicked 'Download Order-wise': '{clicked}'")
            return True

        logger.error("Could not find 'Download Order-wise' button with any strategy.")
        return False

    except (TimeoutException, WebDriverException) as e:
        logger.error(f"Could not find 'Download Order-wise' button: {e}")
        return False


def wait_for_download(download_dir: Path, timeout: int = 120, logger=None, driver=None,
                      started_at: float | None = None) -> Path | None:
    """
    Wait for a new file to appear in download_dir.
    Returns the path to the downloaded file or None if timeout.
    """
    if logger:
        logger.info(f"Waiting for download to complete (up to {timeout}s)...")

    # Snapshot existing files
    existing = set(f.name for f in download_dir.iterdir() if f.is_file())

    for elapsed in range(timeout):
        time.sleep(1)
        # Every 10 seconds, check for portal error alert blocking download
        if driver and elapsed > 0 and elapsed % 10 == 0:
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                if logger:
                    logger.warning(f"Alert during download: '{alert_text}' -- aborting wait.")
                alert.accept()
                return None  # Signal failure so retry loop picks up
            except Exception:
                pass
        current = set(f.name for f in download_dir.iterdir() if f.is_file())
        new_files = current - existing

        # Filter out temp files (.crdownload, .tmp, .part)
        completed = [f for f in new_files
                     if not f.endswith(".crdownload")
                     and not f.endswith(".tmp")
                     and not f.endswith(".part")]

        if completed:
            downloaded = download_dir / completed[0]
            if logger:
                logger.info(f"Download complete: {downloaded.name} ({elapsed + 1}s)")
            return downloaded

        if started_at is not None:
            recent_completed = []
            for f in download_dir.iterdir():
                if not f.is_file():
                    continue
                if f.name.endswith(".crdownload") or f.name.endswith(".tmp") or f.name.endswith(".part"):
                    continue
                try:
                    if f.stat().st_mtime >= started_at:
                        recent_completed.append(f)
                except OSError:
                    continue
            if recent_completed:
                downloaded = max(recent_completed, key=lambda p: p.stat().st_mtime)
                if logger:
                    logger.info(f"Download detected via recent file timestamp: {downloaded.name} ({elapsed + 1}s)")
                return downloaded

    if logger:
        logger.error(f"Download timed out after {timeout}s.")
    return None


def rename_downloaded_file(file_path: Path, logger) -> Path:
    """Rename downloaded file to include today's date."""
    today = datetime.now().strftime("%Y-%m-%d")
    new_name = f"Filflo_OrderDump_{today}{file_path.suffix}"
    new_path = file_path.parent / new_name

    # If a file with the same name already exists, remove it
    if new_path.exists():
        new_path.unlink()

    file_path.rename(new_path)
    logger.info(f"Renamed: {file_path.name} -> {new_name}")
    return new_path


def read_csv_rows(file_path: Path, logger) -> list[list[str]]:
    """Read CSV rows using a forgiving encoding fallback chain."""
    encodings = ("utf-8-sig", "utf-8", "latin1")
    last_error = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as handle:
                rows = [[str(cell) for cell in row] for row in csv.reader(handle)]
            if logger:
                logger.info(f"Loaded {len(rows)} CSV row(s) from {file_path.name} using {encoding}.")
            return rows
        except UnicodeDecodeError as exc:
            last_error = exc
            continue

    raise UnicodeDecodeError(
        last_error.encoding if last_error else "utf-8",
        last_error.object if last_error else b"",
        last_error.start if last_error else 0,
        last_error.end if last_error else 1,
        last_error.reason if last_error else f"Could not decode CSV: {file_path}",
    )


def find_latest_downloaded_report(logger=None) -> Path | None:
    """Return the most recently modified downloaded Order-wise CSV, if present."""
    patterns = ("Order-wise*.csv", "Filflo_OrderDump_*.csv")
    candidates: list[Path] = []
    seen = set()

    for pattern in patterns:
        for path in DOWNLOAD_FOLDER.glob(pattern):
            if not path.is_file():
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            candidates.append(path)

    if not candidates:
        if logger:
            logger.warning(f"No downloaded Order-wise CSV found in {DOWNLOAD_FOLDER}")
        return None

    latest = max(candidates, key=lambda p: p.stat().st_mtime)
    if logger:
        logger.info(f"Latest downloaded report selected: {latest.name}")
    return latest


def normalize_tabular_rows(rows: list[list[str]]) -> list[list[str]]:
    """Pad/truncate rows so Google Sheet writes stay rectangular."""
    if not rows:
        return []

    width = max(len(rows[0]), 1)
    normalized = []
    for row in rows:
        cells = ["" if cell is None else str(cell) for cell in row[:width]]
        if len(cells) < width:
            cells.extend([""] * (width - len(cells)))
        normalized.append(cells)
    return normalized


def backup_google_sheet_values(sheet_title: str, worksheet_title: str, values: list[list[str]], logger) -> Path | None:
    """Persist a local CSV backup before overwriting a live Google Sheet tab."""
    if not values:
        return None

    GSHEET_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    safe_sheet = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in sheet_title).strip("_") or "sheet"
    safe_tab = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in worksheet_title).strip("_") or "tab"
    backup_path = GSHEET_BACKUP_DIR / f"{safe_sheet}_{safe_tab}_{datetime.now():%Y%m%d_%H%M%S}.csv"

    with open(backup_path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(values)

    logger.info(f"Google Sheet backup saved locally: {backup_path}")
    return backup_path


def _write_rows_to_worksheet(worksheet, rows: list[list[str]], logger, chunk_size: int = 250) -> None:
    """Write rows to a worksheet in chunks to stay under API payload limits."""
    normalized_rows = normalize_tabular_rows(rows)
    if not normalized_rows:
        worksheet.clear()
        worksheet.resize(rows=1, cols=1)
        return

    row_count = len(normalized_rows)
    col_count = len(normalized_rows[0])
    worksheet.resize(rows=max(row_count, 1), cols=max(col_count, 1))

    for start in range(0, row_count, chunk_size):
        chunk = normalized_rows[start:start + chunk_size]
        end_row = start + len(chunk)
        range_name = f"A{start + 1}:{get_column_letter(col_count)}{end_row}"
        worksheet.update(range_name=range_name, values=chunk)
        logger.info(f"Google Sheet sync progress: rows {start + 1}-{end_row} uploaded.")


def get_google_sheet_targets() -> list[dict[str, str]]:
    """Return configured Google Sheet targets, deduplicated in declaration order."""
    configured = [
        {"sheet_id": GOOGLE_SHEET_ID, "worksheet_name": GOOGLE_WORKSHEET_NAME},
        {"sheet_id": GOOGLE_SHEET_ID_2, "worksheet_name": GOOGLE_WORKSHEET_NAME_2},
    ]

    targets = []
    seen = set()
    for item in configured:
        sheet_id = (item.get("sheet_id") or "").strip()
        worksheet_name = (item.get("worksheet_name") or "Sheet1").strip() or "Sheet1"
        if not sheet_id:
            continue
        key = (sheet_id, worksheet_name)
        if key in seen:
            continue
        seen.add(key)
        targets.append({"sheet_id": sheet_id, "worksheet_name": worksheet_name})
    return targets


def align_rows_to_target_header(source_rows: list[list[str]], target_header: list[str], logger, target_label: str) -> list[list[str]]:
    """Project CSV rows onto a target worksheet header by column name."""
    normalized_rows = normalize_tabular_rows(source_rows)
    if not normalized_rows:
        return []

    source_header = normalized_rows[0]
    target_header = [str(cell) for cell in target_header if str(cell).strip()] if target_header else []

    if not target_header or target_header == source_header:
        return normalized_rows

    missing_headers = [header for header in source_header if header not in target_header]
    if missing_headers:
        raise ValueError(
            f"{target_label} is missing CSV column(s): {', '.join(missing_headers[:5])}"
        )

    source_index = {header: idx for idx, header in enumerate(source_header)}
    extra_headers = [header for header in target_header if header not in source_index]
    if extra_headers:
        preview = ", ".join(extra_headers[:5])
        suffix = "..." if len(extra_headers) > 5 else ""
        logger.info(
            f"{target_label} has {len(extra_headers)} extra column(s); "
            f"importing blanks for: {preview}{suffix}"
        )

    aligned_rows = [list(target_header)]
    for row in normalized_rows[1:]:
        aligned_rows.append([
            row[source_index[header]] if header in source_index and source_index[header] < len(row) else ""
            for header in target_header
        ])
    return aligned_rows


def sync_csv_to_google_sheet(file_path: Path, logger) -> bool:
    """Sync the downloaded CSV into all configured Google Sheet targets."""
    if not GOOGLE_SYNC_ENABLED:
        logger.info("Google Sheet sync is disabled by configuration.")
        return True

    targets = get_google_sheet_targets()
    if not targets:
        logger.info("Google Sheet sync skipped: no target sheets are configured.")
        return True

    if not GOOGLE_CLIENT_SECRET.exists() or not GOOGLE_AUTH_TOKEN.exists():
        logger.error("Google Sheet sync failed: OAuth credential files are missing.")
        return False

    try:
        import gspread
    except ImportError:
        logger.error("Google Sheet sync failed: gspread is not installed.")
        return False

    try:
        csv_rows = normalize_tabular_rows(read_csv_rows(file_path, logger))
        if not csv_rows:
            logger.error(f"Google Sheet sync aborted: CSV is empty: {file_path}")
            return False

        gc = gspread.oauth(
            credentials_filename=str(GOOGLE_CLIENT_SECRET),
            authorized_user_filename=str(GOOGLE_AUTH_TOKEN),
        )
        overall_success = True

        for index, target in enumerate(targets, start=1):
            spreadsheet = gc.open_by_key(target["sheet_id"])
            worksheet = spreadsheet.worksheet(target["worksheet_name"])
            target_label = f"Google Sheet target {index}: '{spreadsheet.title}' / '{worksheet.title}'"

            existing_values = worksheet.get_all_values()
            backup_google_sheet_values(spreadsheet.title, worksheet.title, existing_values, logger)
            rows_to_write = align_rows_to_target_header(
                csv_rows,
                existing_values[0] if existing_values else csv_rows[0],
                logger,
                target_label,
            )

            try:
                _write_rows_to_worksheet(worksheet, rows_to_write, logger)
            except Exception:
                if existing_values:
                    logger.warning(f"{target_label} failed mid-write. Attempting worksheet restore from backup data.")
                    _write_rows_to_worksheet(worksheet, existing_values, logger)
                raise

            logger.info(f"{target_label} updated with {len(rows_to_write) - 1} data row(s).")

        return overall_success

    except Exception as e:
        logger.error(f"Google Sheet sync failed: {e}", exc_info=True)
        return False


def parse_email_recipients(recipient_text: str) -> list[str]:
    """Split comma/semicolon/newline separated recipients into a clean list."""
    if not recipient_text:
        return []

    cleaned = recipient_text.replace("\xa0", " ").replace(";", ",")
    parts = [p.strip() for p in cleaned.replace("\n", ",").split(",")]
    return [p for p in parts if p]


def send_email_with_attachment(file_path: Path, logger, recipient_email: str = None) -> bool:
    """Send the downloaded file via Gmail SMTP."""
    if GMAIL_SENDER == "YOUR_GMAIL@gmail.com" or GMAIL_APP_PASSWORD == "xxxx xxxx xxxx xxxx":
        logger.warning("Email not configured. Update GMAIL_SENDER and GMAIL_APP_PASSWORD in the script.")
        logger.warning(f"File saved locally at: {file_path}")
        return False

    # Use dynamic recipient if provided, else fall back to .env default
    actual_recipient = (recipient_email or EMAIL_RECIPIENT).strip()
    recipient_list = parse_email_recipients(actual_recipient)
    if not recipient_list:
        logger.warning("No valid recipient email configured.")
        logger.info(f"File is still saved locally at: {file_path}")
        return False

    today = datetime.now().strftime("%d-%b-%Y")
    # Sanitize filename: replace ALL non-ASCII chars (like \xa0 non-breaking space) with regular space
    import unicodedata
    raw_name = file_path.name
    safe_filename = unicodedata.normalize("NFKD", raw_name)
    safe_filename = safe_filename.encode("ascii", "ignore").decode("ascii")
    if not safe_filename:
        safe_filename = f"Filflo_Download_{today}.csv"
    logger.info(f"Attachment filename (sanitized): '{safe_filename}'")

    subject = f"Filflo Order-wise Data Dump - {today}"
    body = (f"Hi,\n\n"
            f"Please find attached the Filflo B2B Order-wise data dump for the last 1 year.\n\n"
            f"File: {safe_filename}\n"
            f"Downloaded on: {today}\n\n"
            f"- Filflo Bot")

    try:
        msg = MIMEMultipart()
        msg["From"] = GMAIL_SENDER.replace("\xa0", " ").strip()
        msg["To"] = ", ".join(recipient_list)
        msg["Subject"] = subject.replace("\xa0", " ")
        msg.attach(MIMEText(body, "plain", "utf-8"))

        # Attach file
        with open(file_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=safe_filename)
        msg.attach(part)

        # Sanitize credentials: replace non-breaking spaces (\xa0) with regular spaces
        clean_sender = GMAIL_SENDER.replace("\xa0", " ").strip()
        clean_password = GMAIL_APP_PASSWORD.replace("\xa0", " ").strip()

        # Send using raw bytes to avoid any further encoding issues
        logger.info(f"Sending email to {', '.join(recipient_list)}...")
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(clean_sender, clean_password)
            server.sendmail(clean_sender, recipient_list, msg.as_bytes())

        logger.info(f"Email sent successfully to {', '.join(recipient_list)}.")
        return True

    except (smtplib.SMTPException, ConnectionError, OSError) as e:
        logger.error(f"Email failed: {e}")
        logger.info(f"File is still saved locally at: {file_path}")
        return False


def has_download_in_progress_state(driver, logger=None) -> bool:
    """
    Detect whether the portal already shows a live 'Downloading...' state.
    This can happen even when the Order-wise button text itself is no longer available.
    """
    try:
        buttons = driver.find_elements(By.TAG_NAME, "button")
        visible_texts = []
        for btn in buttons:
            if btn.is_displayed():
                txt = (btn.text or "").strip()
                if txt:
                    visible_texts.append(txt)

        downloading = [txt for txt in visible_texts if "downloading" in txt.lower()]
        if downloading:
            if logger:
                logger.info(f"Detected active download state from buttons: {downloading}")
            return True
    except WebDriverException as e:
        if logger:
            logger.warning(f"Could not inspect download button state: {e}")
    return False


def do_download_order_dump(logger, recipient_email: str = None) -> bool:
    """
    Full download workflow:
    1. Login to Filflo portal
    2. Apply Custom Range filter (last 1 year)
    3. Click 'Download Order-wise'
    4. Wait for download to complete
    5. Rename file with today's date
    6. Email the file to nandlal@anveshan.farm
    """
    logger.info("=" * 60)
    logger.info("  DOWNLOAD MODULE — Order-wise Data Dump (Last 1 Year)")
    logger.info("=" * 60)

    # Calculate date range: today and 1 year ago
    today = datetime.now()
    one_year_ago = today - timedelta(days=365)
    from_date = one_year_ago.strftime("%m/%d/%Y")
    to_date = today.strftime("%m/%d/%Y")

    logger.info(f"Date range: {from_date} to {to_date}")

    driver = None
    try:
        driver = create_download_driver(logger)

        if not login(driver, logger):
            logger.error("Login failed for download module.")
            return False

        time.sleep(3)
        dismiss_popups(driver, logger)

        # ── Retry loop: keep retrying until download succeeds ─────────────────
        MAX_ATTEMPTS       = 10    # retry up to 10 times
        RETRY_WAIT_SECONDS = 300   # 5 minutes between retries

        downloaded_file = None
        for attempt in range(1, MAX_ATTEMPTS + 1):
            logger.info(f"=== Download attempt {attempt}/{MAX_ATTEMPTS} ===")
            attempt_started_at = time.time()

            # Refresh page at the start of each attempt to clear any stuck
            # "Downloading..." spinner state on the buttons
            if attempt > 1:
                logger.info("Refreshing page to reset button states...")
                driver.refresh()
                time.sleep(5)
                dismiss_popups(driver, logger)

            # Re-apply date filter (re-navigate each attempt)
            if not apply_custom_date_filter(driver, from_date, to_date, logger):
                logger.warning(f"Date filter failed on attempt {attempt}.")
                if attempt < MAX_ATTEMPTS:
                    logger.info("Waiting 5 minutes before retry...")
                    time.sleep(RETRY_WAIT_SECONDS)
                    driver.refresh()
                    time.sleep(5)
                    dismiss_popups(driver, logger)
                continue

            time.sleep(2)

            click_started = click_download_order_wise(driver, logger)
            download_in_progress = has_download_in_progress_state(driver, logger=logger)
            if not click_started and not download_in_progress:
                logger.warning(f"Download button click failed on attempt {attempt}.")
                if attempt < MAX_ATTEMPTS:
                    logger.info("Waiting 5 minutes before retry...")
                    time.sleep(RETRY_WAIT_SECONDS)
                    driver.refresh()
                    time.sleep(5)
                    dismiss_popups(driver, logger)
                continue
            if not click_started and download_in_progress:
                logger.info("Download button already shows 'Downloading...' state; waiting for file completion.")

            # Brief pause then check for immediate error alert from portal
            time.sleep(5)
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                logger.warning(f"Portal error alert: '{alert_text}'")
                alert.accept()
                logger.warning(f"Download failed on attempt {attempt} due to portal error.")
                if attempt < MAX_ATTEMPTS:
                    logger.info("Waiting 5 minutes before retry...")
                    time.sleep(RETRY_WAIT_SECONDS)
                    driver.refresh()
                    time.sleep(5)
                    dismiss_popups(driver, logger)
                continue
            except (NoAlertPresentException, WebDriverException):
                pass  # No alert = download started normally

            # Wait up to 10 minutes for the file to appear
            downloaded_file = wait_for_download(
                DOWNLOAD_FOLDER,
                timeout=600,
                logger=logger,
                driver=driver,
                started_at=attempt_started_at,
            )

            if downloaded_file:
                # Safety check: reject if we accidentally got SKU-wise
                if "sku" in downloaded_file.name.lower():
                    logger.warning(f"Wrong file downloaded (SKU-wise): {downloaded_file.name} -- deleting and retrying.")
                    try:
                        downloaded_file.unlink()
                    except OSError:
                        pass
                    downloaded_file = None
                    if attempt < MAX_ATTEMPTS:
                        driver.refresh()
                        time.sleep(5)
                        dismiss_popups(driver, logger)
                    continue
                logger.info(f"Download succeeded on attempt {attempt}: {downloaded_file.name}")
                break

            logger.warning(f"Download timed out on attempt {attempt}.")
            if attempt < MAX_ATTEMPTS:
                logger.info("Waiting 5 minutes before retry...")
                time.sleep(RETRY_WAIT_SECONDS)
                driver.refresh()
                time.sleep(5)
                dismiss_popups(driver, logger)
            else:
                logger.error(f"All {MAX_ATTEMPTS} download attempts exhausted.")

        if not downloaded_file:
            logger.error("Download did not complete after all retries.")
            return False

        # Send email with the downloaded file as-is (no rename)
        email_sent = send_email_with_attachment(downloaded_file, logger, recipient_email=recipient_email)
        sheet_synced = sync_csv_to_google_sheet(downloaded_file, logger)

        if email_sent:
            logger.info(f"Download + Email complete: {downloaded_file.name}")
        else:
            logger.error(f"Download complete but email failed: {downloaded_file.name}")

        if sheet_synced:
            logger.info(f"Google Sheet sync complete: {downloaded_file.name}")
        else:
            logger.error(f"Download complete but Google Sheet sync failed: {downloaded_file.name}")

        return email_sent and sheet_synced

    except Exception as e:
        logger.error(f"Download module failed: {e}", exc_info=True)
        return False
    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed (download mode).")
            except WebDriverException:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Filflo B2B Combined Bot v3")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_PATH))
    parser.add_argument("--once", action="store_true")
    parser.add_argument("--interval", type=int, default=POLL_INTERVAL_SECONDS)
    parser.add_argument("--reset", action="store_true",
                        help="Clear all Status values before running")
    parser.add_argument("--download", action="store_true",
                        help="Download Order-wise data dump (last 1 year) and email it")
    args = parser.parse_args()

    BOT_FOLDER.mkdir(parents=True, exist_ok=True)
    POD_FOLDER.mkdir(parents=True, exist_ok=True)
    POD_DONE_FOLDER.mkdir(parents=True, exist_ok=True)

    logger = setup_logging(LOG_DIR)

    # ── Download mode ─────────────────────────────────────────────────────
    if args.download:
        logger.info("=" * 60)
        logger.info("  Filflo B2B Bot — DOWNLOAD MODE")
        logger.info("=" * 60)
        success = do_download_order_dump(logger)
        logger.info(f"Download {'succeeded' if success else 'FAILED'}.")
        sys.exit(0 if success else 1)

    # ── Normal mode (delivery + POD processing) ──────────────────────────
    logger.info("=" * 60)
    logger.info("  Filflo B2B Combined Bot v3")
    logger.info("  Delivery Date + POD Upload in ONE form")
    logger.info(f"  Excel: {args.excel}")
    logger.info(f"  POD folder: {POD_FOLDER}")
    logger.info(f"  Mode: {'Single run' if args.once else 'Continuous'}")
    logger.info("=" * 60)

    excel_path = Path(args.excel)

    if not excel_path.is_file():
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    if args.reset:
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            cleared = 0
            for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
                if ws.cell(row=row_idx, column=COL_STATUS).value:
                    ws.cell(row=row_idx, column=COL_STATUS, value=None)
                    cleared += 1
            wb.save(excel_path)
            wb.close()
            logger.info(f"Reset: cleared {cleared} row(s).")
        except PermissionError:
            logger.error("Excel file is open. Close it first.")
            sys.exit(1)

    if args.once:
        summary = process_all_entries(excel_path, logger)
        logger.info(f"\nDone: {summary['verified']} verified, {summary['unverified']} unverified, "
                     f"{summary['skipped']} skipped, {summary['failed']} failed "
                     f"out of {summary['total']}.")
    else:
        logger.info(f"Monitoring every {args.interval}s. Press Ctrl+C to stop.")
        try:
            while True:
                process_all_entries(excel_path, logger)
                logger.info(f"Sleeping {args.interval}s...")
                time.sleep(args.interval)
        except KeyboardInterrupt:
            logger.info("\nStopped by user.")

    logger.info("Bot v3 finished.")


if __name__ == "__main__":
    main()

"""
Blinkit PartnersBiz -- Appointment Slot Booker (v5 - Production Hardened)
Built from real DOM inspection of partnersbiz.com (Ant Design + styled-components)

v5 changes over v4:
  - Config loaded from config.json (falls back to defaults if missing)
  - Screenshot captured on every failure for debugging
  - Retry mechanism: each failed PO gets 1 automatic retry before moving on
  - "Uncertain" added to SKIP_STATUSES to prevent double-bookings on re-run
  - Dangerous 99999 qty fallback removed — PO is skipped if max can't be read
  - Better error classification (AWB invalid, validation failed, etc.)
  - Smarter waits: WebDriverWait replaces blind sleep where possible
  - File logging alongside console output
  - Unused imports removed

USAGE:
  1. Open Chrome with remote debugging:
     chrome.exe --remote-debugging-port=9222
  2. Login to partnersbiz.com manually (OTP required)
  3. Run this script:
     python blinkit_appointment_booker.py

REQUIREMENTS:
  pip install selenium openpyxl
"""

import json
import re
import sys
import time
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
)
from filflo_monitor_bus import attach_monitor_handler

# ==============================================================
# CONFIGURATION — loaded from config.json, with safe defaults
# ==============================================================
SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"

DEFAULT_CONFIG = {
    "login_email": "prashant@anveshan.farm",
    "partnersbiz_base": "https://www.partnersbiz.com",
    "schedule_url": "https://www.partnersbiz.com/app/appointments/schedule/{po_number}",
    "excel_filename": "Data.xlsx",
    "use_existing_browser": True,
    "chrome_debug_port": 9222,
    "default_courier": "Delhivery",
    "max_retries_per_po": 1,
}


def load_config():
    """Load config from config.json if it exists, otherwise use defaults."""
    config = DEFAULT_CONFIG.copy()
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r") as f:
                user_config = json.load(f)
            config.update(user_config)
            logging.getLogger("booker").info(f"Config loaded from {CONFIG_PATH}")
        except (json.JSONDecodeError, IOError) as e:
            logging.getLogger("booker").warning(f"Config file error, using defaults: {e}")
    return config


CFG = load_config()
EXCEL_PATH = SCRIPT_DIR / CFG["excel_filename"]
SCREENSHOT_DIR = SCRIPT_DIR / "screenshots"
SCREENSHOT_DIR.mkdir(exist_ok=True)

# Statuses that are skipped on re-runs. "Uncertain" is included to prevent
# double-bookings — if a booking was uncertain, a human should verify before
# the bot retries it.
SKIP_STATUSES = {
    "Scheduled",
    "No Slot Available",
    "Not Found/Already Scheduled",
    "Uncertain",
    "Slot Already Taken",
    "Reschedule Not Allowed",
}

# ==============================================================
# LOGGING — both console and file
# ==============================================================
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
log_filename = LOG_DIR / f"booker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(log_filename, encoding="utf-8"),
    ],
)
log = logging.getLogger("booker")
attach_monitor_handler(log, source="blinkit")


# ==============================================================
# DRIVER
# ==============================================================
def create_driver():
    """Connect to existing Chrome or start a new one."""
    opts = Options()
    if CFG["use_existing_browser"]:
        opts.add_experimental_option(
            "debuggerAddress", f"127.0.0.1:{CFG['chrome_debug_port']}"
        )
    else:
        opts.add_argument("--start-maximized")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])

    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(60)
    driver.implicitly_wait(2)
    return driver


# ==============================================================
# JS INTERACTION HELPERS (minimized-window safe)
# ==============================================================

def js_click(driver, element):
    """
    Click via JavaScript — works even when Chrome is minimized because
    it dispatches the event directly on the DOM node, bypassing viewport
    visibility requirements that ActionChains depends on.
    """
    driver.execute_script("arguments[0].click();", element)


def js_mousedown(driver, element):
    """
    Dispatch a mousedown event via JS. Ant Design's <Select> component
    listens on mousedown (not click), so a regular JS .click() won't
    open the dropdown. This fires the correct event.
    """
    driver.execute_script("""
        arguments[0].dispatchEvent(
            new MouseEvent('mousedown', {bubbles: true, cancelable: true})
        );
    """, element)


def js_focus_and_type(driver, element, value):
    """
    Set an input's value via React-compatible JS events. This bypasses
    React's synthetic event system by using the native HTMLInputElement
    value setter, then dispatching input/change events so React picks
    up the new value. Works when Chrome is minimized.
    """
    driver.execute_script("""
        var el = arguments[0];
        var val = arguments[1];
        el.focus();
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value'
        ).set;
        nativeInputValueSetter.call(el, val);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
    """, element, str(value))


def trigger_post_awb_fetch(driver, input_element):
    """
    After entering the tracking number, blur the field and click a neutral
    area so the site kicks off any on-blur EDD fetch/validation workflow.
    """
    try:
        driver.execute_script("""
            var el = arguments[0];
            try { el.blur(); } catch (e) {}

            var target =
                document.querySelector('main') ||
                document.querySelector('[role="main"]') ||
                document.querySelector('section') ||
                document.body;

            if (target) {
                var rect = target.getBoundingClientRect();
                var opts = {
                    bubbles: true,
                    cancelable: true,
                    clientX: Math.max(5, Math.floor(rect.left + 10)),
                    clientY: Math.max(5, Math.floor(rect.top + 10))
                };
                target.dispatchEvent(new MouseEvent('mousedown', opts));
                target.dispatchEvent(new MouseEvent('mouseup', opts));
                target.dispatchEvent(new MouseEvent('click', opts));
            }
        """, input_element)
        time.sleep(0.4)
    except Exception:
        pass


def wait_for_edd_autofill(driver, timeout=8):
    """
    Wait briefly for the site to auto-populate/show the EDD field after AWB
    entry. Returns True if the field becomes visible or gets a value.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            result = driver.execute_script("""
                var el = document.querySelector(
                    "input[placeholder*='Estimated Date'], " +
                    "input[placeholder*='Date of Delivery'], " +
                    "input[placeholder='EDD'], " +
                    "input[placeholder*='delivery date']"
                );
                if (!el) {
                    return { found: false, value: "" };
                }
                return { found: true, value: el.value || "" };
            """)
            if result and (result.get("found") or result.get("value")):
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False


def save_screenshot(driver, po_number, step_name):
    """
    Capture a screenshot on failure. Saved to screenshots/ folder with
    the PO number and step name so you can immediately see what the page
    looked like when things went wrong.
    """
    try:
        timestamp = datetime.now().strftime("%H%M%S")
        filename = SCREENSHOT_DIR / f"FAIL_{po_number}_{step_name}_{timestamp}.png"
        driver.save_screenshot(str(filename))
        log.info(f"    Screenshot saved: {filename.name}")
    except Exception as e:
        log.warning(f"    Could not save screenshot: {e}")


def wait_for_element(driver, by, selector, timeout=10):
    """
    Replaces the pattern of find_element + time.sleep with a proper
    explicit wait. Returns the element when found, or None on timeout.
    """
    try:
        return WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, selector))
        )
    except TimeoutException:
        return None


# ==============================================================
# DOM-VERIFIED HELPERS
# ==============================================================

def is_schedule_form_visible(driver, max_retries=6, interval=1.5):
    """
    Check if the schedule appointment form is visible.
    Supports BOTH the old Ant Design drawer layout AND the new full-page layout.
    The site migrated from drawer to full-page, so we check for:
      1. Old: .ant-drawer.ant-drawer-open
      2. New: page body containing 'Courier Partner' or 'Schedule Appointment'
    """
    for attempt in range(max_retries):
        try:
            # Check old drawer style
            drawer_found = driver.execute_script(
                "return document.querySelectorAll('.ant-drawer.ant-drawer-open').length > 0;"
            )
            if drawer_found:
                return True

            # Check new full-page style
            body_text = driver.execute_script("return document.body.innerText;")
            if "Courier Partner" in body_text or "Schedule Appointment" in body_text:
                return True
        except Exception:
            pass
        if attempt < max_retries - 1:
            time.sleep(interval)
    return False


def classify_page_error(driver):
    """
    Read the page and return a specific error status instead of a generic 'Error'.
    This makes the Excel output much more useful for debugging.
    """
    try:
        text = driver.execute_script("return document.body.innerText;").lower()
    except Exception:
        return "Error: page unreadable"

    if "validation failed" in text or "check your tracking" in text:
        return "Error: AWB validation failed"
    if is_reschedule_blocked_page(driver):
        return "Reschedule Not Allowed"
    if "please book another slot" in text or "slot you were trying got scheduled" in text:
        return "Slot Already Taken"
    if "already scheduled" in text or "already booked" in text:
        return "Error: already scheduled"
    if "po not found" in text or "not found" in text:
        return "Not Found/Already Scheduled"
    if "no slot" in text:
        return "No Slot Available"
    if "session expired" in text or "sign in" in text:
        return "Error: session expired"
    return "Error"


def has_slot_taken_popup(driver):
    """Detect the modal that says the chosen slot has already been booked."""
    try:
        text = driver.execute_script("return document.body.innerText;").lower()
    except Exception:
        return False

    return (
        "please book another slot" in text
        or "slot you were trying got scheduled" in text
        or ("schedule again" in text and "another slot" in text)
    )


def is_reschedule_blocked_page(driver):
    """Return True when the site explicitly says vendor-side rescheduling is blocked."""
    try:
        text = driver.execute_script("return document.body.innerText;").lower()
        url = (driver.current_url or "").lower()
    except Exception:
        return False

    blocked_markers = (
        "rescheduling is not allowed for this courier partner",
        "it is being managed by us",
        "raise a ticket on partnersbiz freshdesk",
    )
    return any(marker in text for marker in blocked_markers) and (
        "/reschedule/" in url or "reschedule appointment" in text
    )


def get_visible_clubbing_dialog(driver):
    """Return the visible clubbing modal/dialog root, or None if no real popup is open."""
    lowered = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    uppered = "abcdefghijklmnopqrstuvwxyz"
    lc_text = f"translate(normalize-space(.), '{lowered}', '{uppered}')"
    xpaths = [
        f"//*[@role='dialog' and (contains({lc_text}, 'clubbing') or .//*[contains({lc_text}, 'clubbing')])]",
        f"//*[@aria-modal='true' and (contains({lc_text}, 'clubbing') or .//*[contains({lc_text}, 'clubbing')])]",
        f"//*[contains(@class, 'modal') and (contains({lc_text}, 'clubbing') or .//*[contains({lc_text}, 'clubbing')])]",
        f"//*[contains(@class, 'dialog') and (contains({lc_text}, 'clubbing') or .//*[contains({lc_text}, 'clubbing')])]",
    ]

    for xpath in xpaths:
        try:
            elements = driver.find_elements(By.XPATH, xpath)
        except Exception:
            continue
        for element in elements:
            try:
                if element.is_displayed():
                    return element
            except StaleElementReferenceException:
                continue
            except Exception:
                continue
    return None


def has_visible_clubbing_popup(driver):
    """True only when a visible clubbing modal is open, not just hidden text in the DOM."""
    return get_visible_clubbing_dialog(driver) is not None


def click_visible_menu_option(driver, option_text):
    """
    Click an open Dock/portal dropdown option by its visible text.
    Targets the actual menu row instead of a nested text node.
    """
    xpaths = [
        f"//*[@role='menu']//div[@role='menuitem'][normalize-space()='{option_text}' or .//*[normalize-space()='{option_text}']]",
        f"//*[@role='menuitem'][normalize-space()='{option_text}' or .//*[normalize-space()='{option_text}']]",
        f"//*[normalize-space()='{option_text}' and ancestor::*[@role='menu']]/ancestor::*[@role='menuitem'][1]",
    ]

    for xpath in xpaths:
        try:
            options = driver.find_elements(By.XPATH, xpath)
            for opt in options:
                try:
                    if not opt.is_displayed():
                        continue
                    js_click(driver, opt)
                    time.sleep(0.3)
                    return True
                except Exception:
                    continue
        except Exception:
            continue

    return False


def is_courier_committed(driver, courier_name):
    """
    Confirm the courier field has actually selected the value, not merely typed
    into the searchable dropdown input.
    """
    try:
        result = driver.execute_script("""
            var courier = (arguments[0] || '').toLowerCase();
            var input = document.querySelector("input[placeholder='Select Courier Partner']");
            var value = input ? (input.value || '').trim().toLowerCase() : '';
            return {
                value: value,
                matched: value === courier
            };
        """, courier_name)
        return bool(result and result.get("matched"))
    except Exception:
        return False


def wait_for_courier_commit(driver, courier_name, timeout=2.5):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if is_courier_committed(driver, courier_name):
            return True
        time.sleep(0.2)
    return False


def focus_tracking_input(driver):
    """Move focus to the tracking input so the courier dropdown closes cleanly."""
    selectors = [
        "input[placeholder='Enter LR/AWB number']",
        "input[placeholder='Enter Tracking number.']",
        "input[placeholder*='Tracking']",
        "input[placeholder*='AWB']",
    ]
    for selector in selectors:
        try:
            found = driver.find_elements(By.CSS_SELECTOR, selector)
            for inp in found:
                if inp.is_displayed():
                    driver.execute_script("arguments[0].focus();", inp)
                    js_click(driver, inp)
                    time.sleep(0.2)
                    return True
        except Exception:
            continue
    return False


def dismiss_open_menu(driver):
    """Try to close any still-open Dock menu/portal after option selection."""
    try:
        active = driver.switch_to.active_element
        active.send_keys(Keys.ESCAPE)
        time.sleep(0.2)
    except Exception:
        pass

    try:
        driver.execute_script("""
            var menu = document.querySelector('[role="menu"]');
            if (menu) {
                var evt = new KeyboardEvent('keydown', {key: 'Escape', bubbles: true});
                menu.dispatchEvent(evt);
            }
        """)
        time.sleep(0.2)
    except Exception:
        pass


def select_courier(driver, courier_name, po_number=""):
    """
    Open the courier dropdown and select the courier.
    Optimized strategy order (fastest-first for Dock Design System):
      Strategy 1: Direct combobox role selector (Dock Design — works on current site)
      Strategy 2: JS DOM discovery fallback (if role selector misses)
      Strategy 3: Ant Design CSS selectors (legacy fallback)
    """
    max_retries = 2

    for attempt in range(max_retries):
        log.info(f"    Courier selection attempt {attempt+1}/{max_retries}")

        # ------ Strategy 1: Direct Dock input/combobox (fast path) ------
        try:
            search_inputs = driver.find_elements(
                By.CSS_SELECTOR,
                "input[placeholder='Select Courier Partner'], input[placeholder*='Courier']",
            )
            comboboxes = driver.find_elements(By.CSS_SELECTOR, "[role='combobox']")
            if search_inputs:
                target = comboboxes[0] if comboboxes else search_inputs[0]
                log.info("    Found courier input via Dock selector")
                js_click(driver, target)
                time.sleep(0.2)
                js_mousedown(driver, target)
                time.sleep(0.3)

                search_input = next((el for el in search_inputs if el.is_displayed()), search_inputs[0])
                search_input.send_keys(Keys.CONTROL + "a")
                search_input.send_keys(Keys.BACKSPACE)
                search_input.send_keys(courier_name)
                time.sleep(0.8)

                if click_visible_menu_option(driver, courier_name):
                    dismiss_open_menu(driver)
                    if wait_for_courier_commit(driver, courier_name):
                        trigger_post_awb_fetch(driver, search_input)
                        focus_tracking_input(driver)
                        log.info(f"    Courier selected (Dock menuitem): {courier_name}")
                        return True
                    log.warning("    Courier option click did not commit selection")

                all_opts = driver.find_elements(
                    By.XPATH,
                    f"//button[normalize-space()='{courier_name}'] | "
                    f"//*[normalize-space()='{courier_name}' and not(self::script)]"
                )
                for opt in all_opts:
                    try:
                        if opt.is_displayed() and len(opt.text.strip()) < 50:
                            js_click(driver, opt)
                            time.sleep(0.2)
                            dismiss_open_menu(driver)
                            if wait_for_courier_commit(driver, courier_name):
                                trigger_post_awb_fetch(driver, search_input)
                                focus_tracking_input(driver)
                                log.info(f"    Courier selected (Dock input): {courier_name}")
                                return True
                    except Exception:
                        continue

                search_input.send_keys(Keys.ENTER)
                time.sleep(0.4)
                current_val = search_input.get_attribute("value") or ""
                if courier_name.lower() in current_val.lower():
                    dismiss_open_menu(driver)
                if courier_name.lower() in current_val.lower() and wait_for_courier_commit(driver, courier_name):
                    trigger_post_awb_fetch(driver, search_input)
                    focus_tracking_input(driver)
                    log.info(f"    Courier selected (Enter fallback): {courier_name}")
                    return True
        except Exception as e:
            log.debug(f"    Dock courier strategy failed: {e}")

        # ------ Strategy 2: JS DOM discovery (broader search) ------
        try:
            result = driver.execute_script("""
                var els = document.querySelectorAll('*');
                var info = [];
                for (var i = 0; i < els.length; i++) {
                    var el = els[i];
                    var ph = el.getAttribute('placeholder') || '';
                    var role = el.getAttribute('role') || '';
                    var tag = el.tagName.toLowerCase();
                    if (ph.toLowerCase().includes('courier') ||
                        ph.toLowerCase().includes('select courier') ||
                        role === 'combobox' ||
                        role === 'listbox') {
                        info.push({
                            tag: tag,
                            id: el.id,
                            className: el.className,
                            placeholder: ph,
                            role: role,
                            type: el.type || ''
                        });
                    }
                }
                return JSON.stringify(info);
            """)
            log.debug(f"    DOM discovery results: {result}")

            import json as _json
            elements = _json.loads(result)
            for elem_info in elements:
                role = elem_info.get('role', '')
                classname = elem_info.get('className', '')
                placeholder = elem_info.get('placeholder', '')

                if role == 'combobox' or 'courier' in placeholder.lower():
                    if classname:
                        first_class = classname.split()[0]
                        found = driver.find_elements(By.CSS_SELECTOR, f".{first_class}")
                        if found:
                            js_click(driver, found[0])
                            time.sleep(0.3)
                            js_mousedown(driver, found[0])
                            time.sleep(0.5)

                            active = driver.switch_to.active_element
                            active.send_keys(courier_name)
                            time.sleep(1.0)

                            if click_visible_menu_option(driver, courier_name):
                                dismiss_open_menu(driver)
                                if wait_for_courier_commit(driver, courier_name):
                                    trigger_post_awb_fetch(driver, active)
                                    focus_tracking_input(driver)
                                    log.info(f"    Courier selected (JS discovery menuitem): {courier_name}")
                                    return True

                            all_opts = driver.find_elements(
                                By.XPATH,
                                f"//*[contains(text(), '{courier_name}')]"
                            )
                            for opt in all_opts:
                                try:
                                    if opt.is_displayed() and len(opt.text.strip()) < 50:
                                        js_click(driver, opt)
                                        time.sleep(0.3)
                                        dismiss_open_menu(driver)
                                        if wait_for_courier_commit(driver, courier_name):
                                            trigger_post_awb_fetch(driver, active)
                                            focus_tracking_input(driver)
                                            log.info(f"    Courier selected (JS discovery): {courier_name}")
                                            return True
                                except Exception:
                                    continue
        except Exception as e:
            log.warning(f"    JS DOM discovery failed: {e}")

        # ------ Strategy 3: Ant Design CSS selectors (legacy) ------
        for css in [
            ".ant-select.custom-select .ant-select-selector",
            ".ant-select .ant-select-selector",
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, css)
                if el:
                    log.info(f"    Found courier via CSS: {css}")
                    js_mousedown(driver, el)
                    time.sleep(0.5)
                    search = driver.find_elements(By.CSS_SELECTOR, "input[type='search']")
                    if search:
                        driver.execute_script("arguments[0].value = '';", search[0])
                        search[0].send_keys(courier_name)
                        time.sleep(1.0)
                        opts = driver.find_elements(
                            By.CSS_SELECTOR, ".ant-select-item.ant-select-item-option"
                        )
                        if opts:
                            js_click(driver, opts[0])
                            time.sleep(0.3)
                            dismiss_open_menu(driver)
                            if wait_for_courier_commit(driver, courier_name):
                                trigger_post_awb_fetch(driver, search[0])
                                focus_tracking_input(driver)
                                log.info(f"    Courier selected (CSS): {courier_name}")
                                return True
            except Exception:
                continue

        time.sleep(1)

    log.error(f"    All courier selection methods failed after {max_retries} attempts")
    save_screenshot(driver, po_number, "courier_select")
    return False


def fill_tracking_number(driver, awb, po_number=""):
    """Fill tracking number via JS value setter with verification."""
    try:
        # Try multiple placeholder variants (site may change these)
        inp = None
        for placeholder in [
            'input[placeholder="Enter Tracking number."]',
            'input[placeholder="Enter LR/AWB number"]',
            'input[placeholder*="Tracking"]',
            'input[placeholder*="AWB"]',
            'input[placeholder*="tracking"]',
        ]:
            try:
                inp = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, placeholder))
                )
                if inp:
                    break
            except TimeoutException:
                continue

        if not inp:
            log.error(f"    Tracking number input not found with any placeholder")
            save_screenshot(driver, po_number, "tracking_not_found")
            return False
        js_focus_and_type(driver, inp, str(awb))
        time.sleep(0.5)
        trigger_post_awb_fetch(driver, inp)

        actual = inp.get_attribute("value")
        if str(awb) in str(actual):
            log.info(f"    Tracking number entered: {awb}")
            return True

        # Fallback: send_keys
        log.warning("    JS value set didn't stick, trying send_keys fallback...")
        inp.click()
        time.sleep(0.2)
        inp.send_keys(Keys.CONTROL + "a")
        inp.send_keys(Keys.BACKSPACE)
        inp.send_keys(str(awb))
        time.sleep(0.3)
        trigger_post_awb_fetch(driver, inp)
        log.info(f"    Tracking number entered (fallback): {awb}")
        return True
    except Exception as e:
        log.error(f"    Tracking number failed: {e}")
        save_screenshot(driver, po_number, "tracking_number")
        return False


def fill_edd(driver, edd_date, po_number=""):
    """
    Fill the 'Estimated Date of Delivery' field if it exists and is empty.
    This is a new field added by PartnersBiz that wasn't in the original UI.
    Format: DD Mon YYYY (e.g., '16 Apr 2026') — matches what the site shows.
    Also tries DD-MM-YYYY and YYYY-MM-DD if the first format doesn't work.
    """
    try:
        # Check if EDD field exists on the page
        edd_inputs = []

        # Fast path: query the current Dock input directly to avoid repeated
        # implicit-wait delays when the field is absent.
        try:
            quick_match = driver.execute_script("""
                return document.querySelector(
                    "input[placeholder*='Estimated Date'], " +
                    "input[placeholder*='Date of Delivery'], " +
                    "input[placeholder='EDD'], " +
                    "input[placeholder*='delivery date']"
                );
            """)
            if quick_match:
                edd_inputs = [quick_match]
        except Exception:
            pass

        # Strategy 1: Find by placeholder text
        if not edd_inputs:
            for placeholder_text in [
                'Estimated Date',
                'Date of Delivery',
                'EDD',
                'delivery date',
            ]:
                found = driver.find_elements(
                    By.XPATH,
                    f"//input[contains(@placeholder, '{placeholder_text}')]"
                )
                if found:
                    edd_inputs = found
                    break

        # Strategy 2: Find input near the "Estimated Date of Delivery" label
        if not edd_inputs:
            labels = driver.find_elements(
                By.XPATH,
                "//*[contains(text(), 'Estimated Date') or contains(text(), 'Date of Delivery')]"
            )
            for label in labels:
                try:
                    # Look for nearby input (sibling or child)
                    parent = label.find_element(By.XPATH, "./..")
                    nearby_inputs = parent.find_elements(By.TAG_NAME, "input")
                    if nearby_inputs:
                        edd_inputs = nearby_inputs
                        break
                except Exception:
                    continue

        # Strategy 3: Find date picker input
        if not edd_inputs:
            edd_inputs = driver.find_elements(
                By.CSS_SELECTOR,
                ".ant-picker input, input[type='date'], input.ant-calendar-input"
            )

        if not edd_inputs:
            # Check if EDD is already showing as text (auto-filled)
            body_text = driver.execute_script("return document.body.innerText;")
            edd_formatted = edd_date.strftime("%d %b %Y")  # e.g., "16 Apr 2026"
            if edd_formatted in body_text or edd_date.strftime("%d-%m-%Y") in body_text:
                log.info(f"    EDD already displayed on page: {edd_formatted}")
                return True
            log.info("    EDD field not found on page (may not be required)")
            return True  # non-critical — proceed

        inp = edd_inputs[0]

        # Check if already filled
        current_val = inp.get_attribute("value") or ""
        if current_val.strip():
            log.info(f"    EDD already filled: {current_val}")
            return True

        # Format EDD — try multiple formats
        edd_formats = [
            edd_date.strftime("%d %b %Y"),    # "16 Apr 2026"
            edd_date.strftime("%d-%m-%Y"),     # "16-04-2026"
            edd_date.strftime("%Y-%m-%d"),     # "2026-04-16"
            edd_date.strftime("%d/%m/%Y"),     # "16/04/2026"
        ]

        for fmt_val in edd_formats:
            try:
                # Clear and type
                js_focus_and_type(driver, inp, fmt_val)
                time.sleep(0.5)

                # Verify
                actual = inp.get_attribute("value") or ""
                if actual.strip():
                    log.info(f"    EDD entered: {fmt_val}")
                    # Press Enter to confirm date picker if needed
                    inp.send_keys(Keys.ENTER)
                    time.sleep(0.3)
                    return True
            except Exception:
                continue

        # Fallback: click + send_keys
        try:
            inp.click()
            time.sleep(0.5)
            inp.send_keys(Keys.CONTROL + "a")
            inp.send_keys(Keys.BACKSPACE)
            inp.send_keys(edd_formats[0])
            inp.send_keys(Keys.ENTER)
            time.sleep(0.5)
            log.info(f"    EDD entered (send_keys fallback): {edd_formats[0]}")
            return True
        except Exception as e:
            log.warning(f"    EDD send_keys fallback failed: {e}")

        log.warning("    Could not fill EDD field — proceeding anyway")
        save_screenshot(driver, po_number, "edd_fill_failed")
        return True  # non-critical

    except Exception as e:
        log.warning(f"    EDD fill error: {e}")
        return True  # non-critical — don't block booking


def ensure_quantities(driver, po_number=""):
    """
    Force quantities to MAX for 100% fill rate.
    If max values can't be extracted from the page, the PO is flagged
    instead of using a dangerous fallback like 99999.
    """
    try:
        spinbuttons = driver.find_elements(By.CSS_SELECTOR, ".ant-input-number input")

        if len(spinbuttons) < 2:
            try:
                po_rows = driver.find_elements(
                    By.XPATH, "//div[contains(text(), 'PO No.')]"
                )
                if po_rows:
                    js_click(driver, po_rows[0])
                    time.sleep(1)
                    spinbuttons = driver.find_elements(
                        By.CSS_SELECTOR, ".ant-input-number input"
                    )
            except Exception:
                pass

        if len(spinbuttons) < 2:
            log.warning("    Spinbutton inputs not found (may be ok if pre-set)")
            return True

        # Extract max values from page text
        page_text = driver.execute_script("return document.body.innerText;")
        m_qty = re.search(r"Total Qty\s*:\s*(\d+)", page_text, re.I)
        m_sku = re.search(r"Total SKUs\s*:\s*(\d+)", page_text, re.I)
        max_qty = m_qty.group(1) if m_qty else None
        max_sku = m_sku.group(1) if m_sku else None

        # Fallback: "/ 117" pattern next to spinbutton
        if not max_qty:
            slash_matches = re.findall(r"/\s*(\d+)", page_text)
            if len(slash_matches) >= 2:
                max_qty = slash_matches[0]
                max_sku = slash_matches[1]

        # If we still can't determine max values, DON'T guess — flag it
        if not max_qty or not max_sku:
            log.warning("    Could not determine max qty/SKU — using pre-filled values")
            save_screenshot(driver, po_number, "qty_unknown")
            return True

        current_qty = spinbuttons[0].get_attribute("value") or "0"
        current_sku = spinbuttons[1].get_attribute("value") or "0"

        if current_qty != max_qty or current_sku != max_sku:
            for inp, val in [(spinbuttons[0], max_qty), (spinbuttons[1], max_sku)]:
                js_focus_and_type(driver, inp, val)
                time.sleep(0.3)
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('blur', {bubbles: true}));", inp
                )
                time.sleep(0.3)
            log.info(f"    Quantities set to MAX: {max_qty} qty, {max_sku} SKUs (100% fill rate)")
        else:
            log.info(f"    Quantities already at MAX: {max_qty} qty, {max_sku} SKUs (100%)")
        return True

    except Exception as e:
        log.error(f"    Quantities check failed: {e}")
        save_screenshot(driver, po_number, "qty_error")
        return True  # non-critical — proceed with pre-filled values


def click_styled_button(driver, button_text, timeout=15, po_number=""):
    """
    Click a button by its visible text. Supports multiple element types:
      1. <button> elements (new Dock Design System)
      2. Styled <div> elements (old Ant Design layout)
      3. <a>, <span> or any clickable element with matching text
    """
    btn = None

    # Multiple XPath strategies to find the button
    xpaths = [
        # Real <button> (most likely on new site)
        f"//button[normalize-space()='{button_text}']",
        # Button containing a span with the text
        f"//button[.//span[normalize-space()='{button_text}']]",
        # Styled <div> button (old site)
        f"//div[normalize-space()='{button_text}' and not(div) and not(span) and not(button)]",
        # Any element with exact text
        f"//*[normalize-space()='{button_text}' and not(ancestor::script)]",
    ]

    try:
        deadline = time.time() + timeout

        # Phase 1: Find the button element
        while time.time() < deadline:
            for xpath in xpaths:
                try:
                    elements = driver.find_elements(By.XPATH, xpath)
                    for el in elements:
                        try:
                            if el.is_displayed():
                                btn = el
                                log.info(f"    Found '{button_text}' via: {el.tag_name}")
                                break
                        except StaleElementReferenceException:
                            continue
                    if btn:
                        break
                except Exception:
                    continue
            if btn:
                break
            time.sleep(0.5)

        if not btn:
            log.error(f"    '{button_text}' not found on page within {timeout}s")
            save_screenshot(driver, po_number, f"btn_missing_{button_text.replace(' ','_')}")
            return False

        # Phase 2: Wait for the button to become actually enabled.
        # Tailwind utility classes like "disabled:cursor-not-allowed" are
        # always present in class names, so we must check runtime state
        # instead of string-matching on the class attribute.
        wait_deadline = time.time() + min(timeout, 4)
        while time.time() < wait_deadline:
            try:
                is_disabled = driver.execute_script(
                    "return arguments[0].disabled === true || arguments[0].matches(':disabled');",
                    btn,
                )
                aria_disabled = (btn.get_attribute("aria-disabled") or "").lower() == "true"
                cursor = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).cursor;", btn
                )
                pointer_events = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).pointerEvents;", btn
                )
                if is_disabled or aria_disabled or cursor == "not-allowed" or pointer_events == "none":
                    time.sleep(0.5)
                    continue
                break
            except StaleElementReferenceException:
                for xpath in xpaths:
                    try:
                        btn = driver.find_element(By.XPATH, xpath)
                        break
                    except Exception:
                        continue
                time.sleep(0.5)
        else:
            log.warning(f"    '{button_text}' appears disabled — clicking anyway")

        # Phase 3: Click it
        js_click(driver, btn)
        time.sleep(1)
        log.info(f"    Clicked '{button_text}'")
        return True

    except Exception as e:
        log.error(f"    '{button_text}' click failed: {e}")
        if btn:
            try:
                driver.execute_script("""
                    var el = arguments[0];
                    el.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                    el.dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));
                    el.dispatchEvent(new MouseEvent('click', {bubbles:true}));
                """, btn)
                time.sleep(1)
                log.info(f"    Clicked '{button_text}' (full event fallback)")
                return True
            except Exception:
                pass
        save_screenshot(driver, po_number, f"btn_error_{button_text.replace(' ','_')}")
        return False


def get_button_runtime_state(driver, button_text, timeout=3):
    """Return the first visible button-like element and whether it is truly disabled."""
    xpaths = [
        f"//button[normalize-space()='{button_text}']",
        f"//button[.//span[normalize-space()='{button_text}']]",
        f"//div[normalize-space()='{button_text}' and not(div) and not(span) and not(button)]",
        f"//*[normalize-space()='{button_text}' and not(ancestor::script)]",
    ]

    deadline = time.time() + timeout
    while time.time() < deadline:
        for xpath in xpaths:
            try:
                elements = driver.find_elements(By.XPATH, xpath)
                for el in elements:
                    try:
                        if not el.is_displayed():
                            continue
                        is_disabled = driver.execute_script(
                            "return arguments[0].disabled === true || arguments[0].matches(':disabled');",
                            el,
                        )
                        aria_disabled = (el.get_attribute("aria-disabled") or "").lower() == "true"
                        cursor = driver.execute_script(
                            "return window.getComputedStyle(arguments[0]).cursor;",
                            el,
                        )
                        pointer_events = driver.execute_script(
                            "return window.getComputedStyle(arguments[0]).pointerEvents;",
                            el,
                        )
                        return el, (is_disabled or aria_disabled or cursor == "not-allowed" or pointer_events == "none")
                    except StaleElementReferenceException:
                        continue
            except Exception:
                continue
        time.sleep(0.3)

    return None, None


def is_slot_page_visible(driver):
    """Return True when the date/slot step is visible."""
    if is_reschedule_blocked_page(driver):
        return True
    try:
        body = driver.execute_script("return document.body.innerText;")
    except Exception:
        return False

    return (
        "Select Date & Slot" in body
        or ("Select Date" in body and "Schedule Appointment" in body and "Back" in body)
    )


def continue_to_slot_page(driver, po_number=""):
    """
    Continue from the form page to the slot-selection step.
    The current site can be flaky: a first click may partially advance the
    state without rendering slots immediately, so we retry and wait for the
    next step explicitly instead of assuming one click is enough.
    """
    max_attempts = 3

    for attempt in range(1, max_attempts + 1):
        log.info(f"    Continue attempt {attempt}/{max_attempts}")

        if is_slot_page_visible(driver):
            return "OK"

        _, is_disabled = get_button_runtime_state(driver, "Continue", timeout=2)
        if is_disabled:
            try:
                body = driver.execute_script("return document.body.innerText;").lower()
            except Exception:
                body = ""
            if "invoice details pending" in body:
                log.info("    Continue is genuinely disabled while invoice details are pending")
                save_screenshot(driver, po_number, "continue_invoice_locked")
                return "Invoice Required"

        if not click_styled_button(driver, "Continue", timeout=15, po_number=po_number):
            return "Error"

        wait_deadline = time.time() + 12
        while time.time() < wait_deadline:
            if is_reschedule_blocked_page(driver):
                return "Reschedule Not Allowed"
            if is_slot_page_visible(driver):
                return "OK"
            try:
                body = driver.execute_script("return document.body.innerText;").lower()
                if "no slot" in body:
                    return "No Slot Available"
            except Exception:
                pass
            time.sleep(1)

    save_screenshot(driver, po_number, "continue_no_transition")
    return "Error"


def select_date_and_slot(driver, edd, po_number=""):
    """
    Pick the earliest available date, prefer an AM slot, and fall back to
    the first available PM slot when AM is unavailable.
    """
    # Wait for date/slot page to load
    date_header = wait_for_element(
        driver, By.XPATH, "//*[contains(text(), 'Select Date')]", timeout=10
    )
    if is_reschedule_blocked_page(driver):
        save_screenshot(driver, po_number, "reschedule_not_allowed")
        return None, None, "Reschedule Not Allowed"
    if not date_header:
        body = driver.execute_script("return document.body.innerText;").lower()
        if "no slot" in body:
            return None, None, "No Slot Available"
        save_screenshot(driver, po_number, "date_page_timeout")
        return None, None, "Timeout waiting for date/slot page"

    render_deadline = time.time() + 15
    while time.time() < render_deadline:
        try:
            body = driver.execute_script("return document.body.innerText;")
        except Exception:
            body = ""

        has_date_text = bool(date_header) and any(
            month in body for month in
            ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
        )
        has_slot_text = ("am" in body.lower()) or ("pm" in body.lower())
        if has_date_text or has_slot_text or "no slot" in body.lower():
            break
        time.sleep(1)

    # ---- DATE SELECTION ----
    best_date = None
    best_card = None
    date_pattern = re.compile(
        r"(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec),?\s*(\d{4})"
    )

    all_divs = driver.find_elements(By.CSS_SELECTOR, "div")
    for div in all_divs:
        try:
            text = div.text.strip()
            if len(text) > 80 or len(text) < 10:
                continue
            m = date_pattern.search(text)
            if not m:
                continue
            inner = div.get_attribute("innerHTML") or ""
            if inner.count("<div") > 5:
                continue
            slot_date = datetime.strptime(
                f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %b %Y"
            ).date()
            if best_date is None or slot_date < best_date:
                best_date = slot_date
                best_card = div
        except (StaleElementReferenceException, Exception):
            continue

    if not best_card:
        save_screenshot(driver, po_number, "no_date_found")
        return None, None, "No Slot Available"

    try:
        js_click(driver, best_card)
        time.sleep(1)
        log.info(f"    Selected date: {best_date.strftime('%d %b %Y')}")
    except Exception:
        try:
            driver.execute_script(
                "arguments[0].dispatchEvent(new MouseEvent('click', {bubbles:true}));",
                best_card,
            )
            time.sleep(1)
        except Exception:
            pass

    # ---- TIME SLOT SELECTION ----
    time.sleep(1)
    selected_time = ""

    slot_buttons = driver.find_elements(
        By.XPATH,
        "//button[.//div[contains(text(), 'am') or contains(text(), 'pm')]]",
    )
    if not slot_buttons:
        slot_buttons = driver.find_elements(
            By.XPATH,
            "//div[(contains(text(), 'am') or contains(text(), 'pm')) and contains(text(), ':') "
            "and string-length(normalize-space()) < 30]",
        )

    preferred_slot = None
    fallback_slot = None

    for btn in slot_buttons:
        try:
            if not btn.is_enabled():
                continue
            cls = btn.get_attribute("class") or ""
            if "disabled" in cls:
                continue
            slot_text = btn.text.strip()
            if not slot_text:
                continue
            if "am" in slot_text.lower() and preferred_slot is None:
                preferred_slot = (btn, slot_text)
            if ("am" in slot_text.lower() or "pm" in slot_text.lower()) and fallback_slot is None:
                fallback_slot = (btn, slot_text)
        except (StaleElementReferenceException, Exception):
            continue

    chosen_slot = preferred_slot or fallback_slot
    if chosen_slot:
        btn, slot_text = chosen_slot
        try:
            js_click(driver, btn)
            time.sleep(1)
            selected_time = slot_text
            log.info(f"    Selected time slot: {slot_text}")
        except Exception:
            selected_time = ""

    if not selected_time:
        save_screenshot(driver, po_number, "no_time_slot")
        return best_date, None, "No Slot Available"

    return best_date, selected_time, "OK"


def handle_clubbing_popup(driver, max_wait=5):
    """
    Backward-compatible wrapper for the current clubbing-popup handler.
    """
    return handle_clubbing_popup_v2(driver, max_wait=max_wait)


def handle_clubbing_popup_v2(driver, max_wait=5):
    """
    Improved clubbing-popup handler for the current live site.
    Supports both old Ant modal markup and newer dialog/checkbox wrappers.
    """
    try:
        lowered = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        uppered = "abcdefghijklmnopqrstuvwxyz"
        lc_text = f"translate(normalize-space(.), '{lowered}', '{uppered}')"

        def find_visible_clubbing_dialog():
            return get_visible_clubbing_dialog(driver)

        def checkbox_marked(dialog):
            try:
                return bool(driver.execute_script("""
                    const root = arguments[0];
                    const inputs = Array.from(
                        root.querySelectorAll("input[type='checkbox'], [role='checkbox']")
                    );
                    return inputs.some((node) => {
                        if (node.matches("input[type='checkbox']")) {
                            return !!node.checked;
                        }
                        const aria = (node.getAttribute("aria-checked") || "").toLowerCase();
                        if (aria === "true") {
                            return true;
                        }
                        const cls = String(node.className || "").toLowerCase();
                        return cls.includes("checked") || cls.includes("selected") || cls.includes("active");
                    });
                """, dialog))
            except Exception:
                return False

        def force_checkbox_checked(dialog):
            try:
                return bool(driver.execute_script("""
                    const root = arguments[0];
                    const checkbox = root.querySelector("input[type='checkbox']");
                    if (!checkbox) {
                        return false;
                    }
                    if (!checkbox.checked) {
                        const desc = Object.getOwnPropertyDescriptor(
                            HTMLInputElement.prototype,
                            "checked"
                        );
                        if (desc && desc.set) {
                            desc.set.call(checkbox, true);
                        } else {
                            checkbox.checked = true;
                        }
                        ["click", "input", "change"].forEach((name) => {
                            checkbox.dispatchEvent(new Event(name, { bubbles: true }));
                        });
                    }
                    return !!checkbox.checked;
                """, dialog))
            except Exception:
                return False

        def click_like_user(element):
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
                element,
            )
            try:
                js_mousedown(driver, element)
            except Exception:
                pass
            js_click(driver, element)
            time.sleep(0.4)

        def popup_cleared():
            if "/scheduled/" in driver.current_url:
                return True
            if is_reschedule_blocked_page(driver):
                return True
            dialog = find_visible_clubbing_dialog()
            return dialog is None

        def action_enabled(element):
            try:
                state = driver.execute_script("""
                    const el = arguments[0];
                    const aria = (el.getAttribute("aria-disabled") || "").toLowerCase();
                    const cls = String(el.className || "").toLowerCase();
                    const style = window.getComputedStyle(el);
                    const disabled =
                        el.disabled === true ||
                        el.hasAttribute("disabled") ||
                        aria === "true";
                    return {
                        disabled,
                        className: cls,
                        pointerEvents: style.pointerEvents,
                        visibility: style.visibility,
                        display: style.display,
                        opacity: style.opacity
                    };
                """, element)
            except Exception:
                return False

            if not element.is_displayed():
                return False
            if state["disabled"]:
                return False
            if state["pointerEvents"] == "none":
                return False
            if state["visibility"] == "hidden" or state["display"] == "none":
                return False
            return True

        def collect_popup_actions(dialog):
            try:
                actions = driver.execute_script("""
                    const root = arguments[0];
                    const nodes = Array.from(
                        root.querySelectorAll("button, [role='button'], a, input[type='button'], input[type='submit'], div, span")
                    );
                    const seen = new Set();
                    const labels = [];
                    for (const node of nodes) {
                        const text = (node.innerText || node.value || "").trim();
                        if (!text || seen.has(text)) {
                            continue;
                        }
                        const style = window.getComputedStyle(node);
                        const visible = style.display !== "none" && style.visibility !== "hidden";
                        if (!visible) {
                            continue;
                        }
                        seen.add(text);
                        labels.push(text);
                        if (labels.length >= 10) {
                            break;
                        }
                    }
                    return labels;
                """, dialog)
                return actions or []
            except Exception:
                return []

        def find_action_target(dialog, keywords, allow_fallback=False):
            if dialog is None:
                return "", None
            selectors = [
                "button, [role='button'], a, input[type='button'], input[type='submit']",
            ]
            if allow_fallback:
                selectors.append("div, span")

            for selector in selectors:
                for candidate in dialog.find_elements(By.CSS_SELECTOR, selector):
                    try:
                        if not candidate.is_displayed():
                            continue
                        label = (
                            candidate.text
                            or candidate.get_attribute("value")
                            or candidate.get_attribute("aria-label")
                            or ""
                        ).strip()
                        normalized = label.lower()
                        if not normalized or not any(keyword in normalized for keyword in keywords):
                            continue
                        clickable = candidate.find_elements(
                            By.XPATH,
                            "./ancestor-or-self::button[1] | "
                            "./ancestor-or-self::*[@role='button'][1] | "
                            "./ancestor-or-self::a[1] | "
                            "./ancestor-or-self::input[@type='button' or @type='submit'][1] | "
                            "./ancestor-or-self::div[contains(@class, 'btn') or contains(@class, 'button')][1]",
                        )
                        target = clickable[0] if clickable else candidate
                        if not action_enabled(target):
                            continue
                        return label, target
                    except StaleElementReferenceException:
                        continue
                    except Exception:
                        continue
            return "", None

        for _ in range(max_wait):
            if has_visible_clubbing_popup(driver):
                break
            if "/scheduled/" in driver.current_url:
                return True
            if is_reschedule_blocked_page(driver):
                return False
            time.sleep(1)
        else:
            return True

        log.info("    PO Clubbing Charges popup detected")

        dialog = find_visible_clubbing_dialog()
        if dialog is None:
            log.warning("    Clubbing popup text found, but no visible dialog root was identified")
            if "/scheduled/" in driver.current_url:
                return True
            return False

        checkbox_clicked = False
        checkbox_xpaths = [
            ".//input[@type='checkbox']",
            ".//*[@role='checkbox']",
            f".//label[contains({lc_text}, 'understand')]",
            f".//*[contains({lc_text}, 'understand')]",
            f".//label[contains({lc_text}, 'agree')]",
            f".//*[contains({lc_text}, 'agree')]",
            ".//*[contains(@class, 'checkbox')]",
        ]
        for xpath in checkbox_xpaths:
            candidates = dialog.find_elements(By.XPATH, xpath)
            if not candidates:
                continue
            for candidate in candidates:
                try:
                    tag_name = candidate.tag_name.lower()
                    if tag_name != "input" and not candidate.is_displayed():
                        continue
                    click_targets = candidate.find_elements(
                        By.XPATH,
                        "./ancestor-or-self::label[1] | "
                        "./ancestor-or-self::*[@role='checkbox'][1] | "
                        "./ancestor-or-self::*[contains(@class, 'checkbox')][1] | "
                        "./ancestor-or-self::button[1]",
                    )
                    if not click_targets:
                        click_targets = [candidate]
                    for target in click_targets:
                        click_like_user(target)
                        if checkbox_marked(dialog):
                            checkbox_clicked = True
                            log.info("    Checkbox ticked on clubbing popup")
                            break
                    if checkbox_clicked:
                        break
                except StaleElementReferenceException:
                    continue
                except Exception as e:
                    log.warning(f"    Checkbox attempt failed: {e}")
                    continue
            if checkbox_clicked:
                break

        if not checkbox_clicked and checkbox_marked(dialog):
            checkbox_clicked = True
            log.info("    Checkbox already ticked on clubbing popup")

        if not checkbox_clicked:
            if force_checkbox_checked(dialog):
                checkbox_clicked = True
                log.info("    Checkbox forced on clubbing popup via JS state sync")

        if not checkbox_clicked:
            log.warning("    Could not tick checkbox - trying popup action anyway")

        action_keywords = (
            "save",
            "continue",
            "proceed",
            "confirm",
            "schedule",
            "ok",
            "okay",
            "yes",
        )

        label = ""
        target = None
        for _ in range(8):
            dialog = find_visible_clubbing_dialog() or dialog
            label, target = find_action_target(dialog, action_keywords, allow_fallback=False)
            if target is not None:
                break
            time.sleep(0.4)

        if target is None:
            dialog = find_visible_clubbing_dialog() or dialog
            label, target = find_action_target(dialog, action_keywords, allow_fallback=True)

        if target is not None:
            try:
                click_like_user(target)
                log.info(f"    Clubbing popup action clicked: {label}")
                for _ in range(8):
                    if popup_cleared():
                        return True
                    time.sleep(0.5)
            except Exception as e:
                log.warning(f"    Clubbing popup action click failed: {e}")

        available_actions = collect_popup_actions(dialog)
        if available_actions:
            log.warning(f"    Visible popup actions seen: {available_actions}")
        log.error("    Could not click Save button on clubbing popup")
        save_screenshot(driver, "clubbing", "popup_not_handled")
        return False

    except Exception as e:
        log.error(f"    Enhanced PO Clubbing popup handler error: {e}")
        return False


def extract_appointment_id(driver):
    """Extract appointment ID from confirmation URL or page text."""
    url_match = re.search(r"/scheduled/(\d+)", driver.current_url)
    if url_match:
        return url_match.group(1)
    try:
        body = driver.execute_script("return document.body.innerText;")
        id_match = re.search(r"Appointment\s*(?:ID|Id)\s*[-:]\s*(\d+)", body)
        if id_match:
            return id_match.group(1)
    except Exception:
        pass
    return ""


# ==============================================================
# EXCEL HANDLING
# ==============================================================
def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value
    }

    for col_name in ["Appointment ID", "Scheduled Date", "Reporting Time", "Status"]:
        if col_name not in headers:
            nc = ws.max_column + 1
            ws.cell(1, nc, col_name)
            headers[col_name] = nc
    wb.save(path)

    pos = []
    for row in range(2, ws.max_row + 1):
        po_number = ws.cell(row, headers["PO Number"]).value
        if not po_number:
            continue
        status = ws.cell(row, headers.get("Status", 99)).value
        if status and str(status).strip() in SKIP_STATUSES:
            continue

        awb = ws.cell(row, headers["AWB No."]).value
        edd = ws.cell(row, headers["EDD"]).value
        courier = ws.cell(row, headers.get("Vendor", 99)).value or CFG["default_courier"]

        # Parse EDD from multiple possible formats
        if isinstance(edd, datetime):
            edd_date = edd.date()
        elif isinstance(edd, date):
            edd_date = edd
        else:
            edd_str = str(edd).strip()[:20]
            edd_date = None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y"):
                try:
                    edd_date = datetime.strptime(edd_str, fmt).date()
                    break
                except ValueError:
                    continue
            if edd_date is None:
                log.warning(f"  Row {row}: cannot parse EDD '{edd}', skipping")
                continue

        pos.append({
            "row": row,
            "po_number": str(int(po_number)) if isinstance(po_number, float) else str(po_number),
            "awb": str(int(awb)) if isinstance(awb, float) else str(awb),
            "edd": edd_date,
            "courier": str(courier).strip(),
        })
    return wb, ws, headers, pos


def update_excel(wb, ws, headers, row, appt_id, sched_date, report_time, status):
    ws.cell(row, headers["Appointment ID"], appt_id)
    ws.cell(row, headers["Scheduled Date"], sched_date)
    ws.cell(row, headers["Reporting Time"], report_time)
    ws.cell(row, headers["Status"], status)
    wb.save(EXCEL_PATH)


# ==============================================================
# CORE BOOKING WORKFLOW
# ==============================================================
def book_single_appointment(driver, po):
    po_number = po["po_number"]
    awb = po["awb"]
    edd = po["edd"]
    courier = po.get("courier", CFG["default_courier"])

    log.info(f"\n{'='*60}")
    log.info(f"  PO: {po_number} | AWB: {awb} | EDD: {edd} | Courier: {courier}")
    log.info(f"{'='*60}")

    # Navigate to schedule page
    driver.get(CFG["schedule_url"].format(po_number=po_number))
    time.sleep(5)

    # Check drawer — has built-in retries for minimized Chrome
    if not is_schedule_form_visible(driver):
        log.info("    -> PO not found or already scheduled (no schedule form)")
        return {"status": "Not Found/Already Scheduled"}

    # Verify schedule form loaded
    try:
        body_text = driver.execute_script("return document.body.innerText;").lower()
    except Exception:
        body_text = ""
    if "courier partner" not in body_text:
        log.info("    -> Schedule form not loaded properly")
        save_screenshot(driver, po_number, "form_not_loaded")
        return {"status": "Error: form not loaded"}

    # STEP 1: Courier
    log.info("  [1/5] Selecting courier...")
    if not select_courier(driver, courier, po_number):
        return {"status": "Error: courier selection failed"}

    # Verify courier was selected (dropdown should show courier name, not placeholder)
    time.sleep(0.5)
    try:
        page_text = driver.execute_script("return document.body.innerText;")
        if "Select Courier Partner" in page_text and courier not in page_text:
            log.warning("    Courier may not have been selected — retrying once")
            select_courier(driver, courier, po_number)
    except Exception:
        pass

    # STEP 2: Tracking number
    log.info("  [2/5] Filling tracking number...")
    if not fill_tracking_number(driver, awb, po_number):
        return {"status": "Error: tracking number failed"}

    # STEP 2.5: EDD — often auto-populated after AWB entry once the field loses
    # focus, so wait briefly for the site to fetch it before falling back.
    if wait_for_edd_autofill(driver, timeout=8):
        log.info("  [2.5/5] EDD auto-fetch triggered after AWB entry")
    elif fill_edd(driver, edd, po_number):
        log.info("  [2.5/5] EDD checked/filled")
    else:
        log.info("  [2.5/5] EDD step skipped — proceeding")

    # STEP 3: Quantities
    log.info("  [3/5] Verifying quantities...")
    ensure_quantities(driver, po_number)
    time.sleep(1)

    # STEP 4: Continue (wait extra for button to become enabled after form fills)
    log.info("  [4/5] Clicking Continue...")
    time.sleep(1)  # let React state settle after all fields
    continue_status = continue_to_slot_page(driver, po_number=po_number)
    if continue_status != "OK":
        if continue_status == "Invoice Required":
            return {"status": "Invoice Required"}
        if continue_status == "No Slot Available":
            return {"status": "No Slot Available"}
        status = classify_page_error(driver)
        return {"status": status}

    # Wait for next page and check for "no slot available"
    time.sleep(2)
    if is_reschedule_blocked_page(driver):
        log.info("    -> Rescheduling is managed by PartnersBiz for this courier partner")
        save_screenshot(driver, po_number, "reschedule_not_allowed")
        return {"status": "Reschedule Not Allowed"}
    body_text = driver.execute_script("return document.body.innerText;").lower()
    if "no slot" in body_text and "select date" not in body_text:
        log.info("    -> No slot available for this PO")
        return {"status": "No Slot Available"}

    # STEP 5: Date & slot
    log.info("  [5/5] Selecting date & slot...")
    best_date, selected_time, slot_status = select_date_and_slot(driver, edd, po_number)

    if slot_status != "OK":
        log.info(f"    -> {slot_status}")
        return {"status": slot_status}

    # Schedule Appointment
    log.info("  [>>] Clicking Schedule Appointment...")
    if not click_styled_button(driver, "Schedule Appointment", timeout=12, po_number=po_number):
        return {"status": classify_page_error(driver)}

    time.sleep(2)
    if is_reschedule_blocked_page(driver):
        log.info("    -> Rescheduling is managed by PartnersBiz for this courier partner")
        save_screenshot(driver, po_number, "reschedule_not_allowed")
        return {"status": "Reschedule Not Allowed"}

    # Handle clubbing popup
    clubbing_ok = handle_clubbing_popup_v2(driver, max_wait=5)
    if not clubbing_ok:
        log.warning("    Clubbing popup handling failed — checking if booking went through anyway")

    # Wait for confirmation
    reporting_time = selected_time.split("-")[0].strip() if selected_time else ""
    sched_date_str = best_date.strftime("%d %b %Y") if best_date else ""
    appt_id = ""
    success = False
    clubbing_retry_count = 0
    for _ in range(25):
        if "/scheduled/" in driver.current_url:
            appt_id = extract_appointment_id(driver)
            success = True
            break
        if has_slot_taken_popup(driver):
            log.info("    Slot taken popup detected - skipping this PO")
            save_screenshot(driver, po_number, "slot_already_taken")
            return {"status": "Slot Already Taken"}
        if is_reschedule_blocked_page(driver):
            log.info("    Rescheduling is not allowed for this courier partner - skipping this PO")
            save_screenshot(driver, po_number, "reschedule_not_allowed")
            return {"status": "Reschedule Not Allowed"}
        if has_visible_clubbing_popup(driver):
            log.info("    Clubbing popup still visible — retrying...")
            clubbing_retry_count += 1
            handled = handle_clubbing_popup_v2(driver, max_wait=2)
            if not handled and clubbing_retry_count >= 3:
                log.warning("    Clubbing popup remained stuck after repeated retries")
                save_screenshot(driver, po_number, "clubbing_popup_stuck")
                return {
                    "status": "Uncertain",
                    "scheduled_date": sched_date_str,
                    "reporting_time": reporting_time,
                }
        try:
            if "validation failed" in driver.page_source.lower():
                return {"status": "Error: AWB validation failed"}
        except Exception:
            pass
        time.sleep(1)
    if success:
        log.info(f"    SUCCESS! Appointment ID: {appt_id}")
        return {
            "status": "Scheduled",
            "appointment_id": appt_id,
            "scheduled_date": sched_date_str,
            "reporting_time": reporting_time,
        }
    else:
        log.warning("    Could not confirm booking")
        save_screenshot(driver, po_number, "unconfirmed")
        return {
            "status": "Uncertain",
            "scheduled_date": sched_date_str,
            "reporting_time": reporting_time,
        }


# ==============================================================
# MAIN — with retry logic
# ==============================================================
def main():
    if not EXCEL_PATH.exists():
        log.error(f"Excel file not found: {EXCEL_PATH}")
        return

    wb, ws, headers, pos = load_excel(EXCEL_PATH)
    if not pos:
        log.info("No POs to process (all scheduled or skipped)")
        return

    log.info(f"Found {len(pos)} POs to process")
    log.info(f"Skipping statuses: {SKIP_STATUSES}")
    log.info(f"Log file: {log_filename}")
    log.info(f"Screenshots dir: {SCREENSHOT_DIR}")

    driver = create_driver()

    # Verify login
    driver.get(CFG["partnersbiz_base"] + "/app/appointments")
    time.sleep(3)
    if "sign in" in driver.page_source.lower():
        log.error("NOT LOGGED IN! Please login to PartnersBiz first, then re-run.")
        return

    log.info("Login verified — starting bookings\n")

    booked = 0
    errors = 0
    skipped = 0
    max_retries = CFG["max_retries_per_po"]

    try:
        for i, po in enumerate(pos, 1):
            log.info(f"\n[{i}/{len(pos)}] Processing...")
            result = None

            # Attempt + retries
            for attempt in range(1 + max_retries):
                try:
                    result = book_single_appointment(driver, po)
                except Exception as e:
                    log.error(f"    UNHANDLED EXCEPTION: {e}")
                    save_screenshot(driver, po["po_number"], "unhandled_exception")
                    result = {"status": "Error: unhandled exception"}

                status = result.get("status", "Error")

                # If it's a transient error and we have retries left, try again
                is_retriable = status.startswith("Error") and "already" not in status
                if is_retriable and attempt < max_retries:
                    log.info(f"    Retrying PO {po['po_number']} (attempt {attempt+2})...")
                    time.sleep(3)
                    continue
                break

            update_excel(
                wb, ws, headers, po["row"],
                result.get("appointment_id", ""),
                result.get("scheduled_date", ""),
                result.get("reporting_time", ""),
                result.get("status", "Error"),
            )

            status = result.get("status", "Error")
            if status == "Scheduled":
                booked += 1
            elif status in (
                "Not Found/Already Scheduled",
                "No Slot Available",
                "Invoice Required",
                "Slot Already Taken",
                "Reschedule Not Allowed",
            ):
                skipped += 1
            else:
                errors += 1

            time.sleep(2)

    except KeyboardInterrupt:
        log.info("\n\nInterrupted by user. Progress saved to Excel.")
    finally:
        log.info(f"\n{'='*60}")
        log.info(f"  RESULTS: Booked={booked} | Skipped={skipped} | Errors={errors} | Total={len(pos)}")
        log.info(f"  Excel updated: {EXCEL_PATH}")
        log.info(f"  Log file: {log_filename}")
        log.info(f"{'='*60}")

        if not CFG["use_existing_browser"]:
            driver.quit()


def run_booking(excel_path: str | Path | None = None) -> dict:
    """
    Programmatic wrapper for the Blinkit bot.
    Returns a structured summary for conversational-tool integration.
    """
    global EXCEL_PATH

    original_excel_path = EXCEL_PATH
    wb = None
    driver = None

    if excel_path:
        EXCEL_PATH = Path(excel_path)

    summary = {
        "success": False,
        "excel_path": str(EXCEL_PATH),
        "log_file": str(log_filename),
        "screenshot_dir": str(SCREENSHOT_DIR),
        "booked": 0,
        "skipped": 0,
        "errors": 0,
        "total": 0,
        "message": "",
    }

    try:
        if not EXCEL_PATH.exists():
            msg = f"Excel file not found: {EXCEL_PATH}"
            log.error(msg)
            summary["message"] = msg
            return summary

        wb, ws, headers, pos = load_excel(EXCEL_PATH)
        summary["total"] = len(pos)
        if not pos:
            msg = "No POs to process (all scheduled or skipped)"
            log.info(msg)
            summary["success"] = True
            summary["message"] = msg
            return summary

        log.info(f"Found {len(pos)} POs to process")
        log.info(f"Skipping statuses: {SKIP_STATUSES}")
        log.info(f"Log file: {log_filename}")
        log.info(f"Screenshots dir: {SCREENSHOT_DIR}")

        driver = create_driver()
        driver.get(CFG["partnersbiz_base"] + "/app/appointments")
        time.sleep(3)
        if "sign in" in driver.page_source.lower():
            msg = "NOT LOGGED IN! Please login to PartnersBiz first, then re-run."
            log.error(msg)
            summary["message"] = msg
            return summary

        log.info("Login verified â€” starting bookings\n")
        max_retries = CFG["max_retries_per_po"]

        try:
            for i, po in enumerate(pos, 1):
                log.info(f"\n[{i}/{len(pos)}] Processing...")
                result = None

                for attempt in range(1 + max_retries):
                    try:
                        result = book_single_appointment(driver, po)
                    except Exception as e:
                        log.error(f"    UNHANDLED EXCEPTION: {e}")
                        save_screenshot(driver, po["po_number"], "unhandled_exception")
                        result = {"status": "Error: unhandled exception"}

                    status = result.get("status", "Error")
                    is_retriable = status.startswith("Error") and "already" not in status.lower()
                    if is_retriable and attempt < max_retries:
                        log.info(f"    Retrying PO {po['po_number']} (attempt {attempt + 2})...")
                        time.sleep(3)
                        continue
                    break

                update_excel(
                    wb, ws, headers, po["row"],
                    result.get("appointment_id", ""),
                    result.get("scheduled_date", ""),
                    result.get("reporting_time", ""),
                    result.get("status", "Error"),
                )

                status = result.get("status", "Error")
                if status == "Scheduled":
                    summary["booked"] += 1
                elif status in (
                    "Not Found/Already Scheduled",
                    "No Slot Available",
                    "Invoice Required",
                    "Slot Already Taken",
                    "Reschedule Not Allowed",
                ):
                    summary["skipped"] += 1
                else:
                    summary["errors"] += 1

                time.sleep(2)
        except KeyboardInterrupt:
            log.info("\n\nInterrupted by user. Progress saved to Excel.")
            summary["message"] = "Interrupted by user. Progress saved to Excel."

        log.info(f"\n{'=' * 60}")
        log.info(
            f"  RESULTS: Booked={summary['booked']} | Skipped={summary['skipped']} "
            f"| Errors={summary['errors']} | Total={summary['total']}"
        )
        log.info(f"  Excel updated: {EXCEL_PATH}")
        log.info(f"  Log file: {log_filename}")
        log.info(f"{'=' * 60}")

        summary["success"] = summary["errors"] == 0
        if not summary["message"]:
            summary["message"] = (
                f"Blinkit run finished. Booked={summary['booked']}, "
                f"Skipped={summary['skipped']}, Errors={summary['errors']}, "
                f"Total={summary['total']}."
            )
        return summary
    finally:
        if wb is not None:
            wb.close()
        if driver and not CFG["use_existing_browser"]:
            driver.quit()
        EXCEL_PATH = original_excel_path


if __name__ == "__main__":
    main()

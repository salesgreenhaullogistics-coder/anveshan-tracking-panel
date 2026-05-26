"""
Shared Instamart sync helpers for Python-side fetch and push workflows.
"""

from __future__ import annotations

import logging
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import gspread
import openpyxl
from openpyxl.utils.datetime import from_excel as from_excel_serial
from filflo_monitor_bus import attach_monitor_handler


BOT_FOLDER = Path(__file__).resolve().parent
LOG_DIR = BOT_FOLDER / "logs"
DATA_XLSX_PATH = BOT_FOLDER / "Data.xlsx"

GOOGLE_CLIENT_SECRET = BOT_FOLDER / "client_secret.json"
GOOGLE_AUTH_TOKEN = BOT_FOLDER / "authorized_user.json"

OMKARA_SHEET_ID = "1VcGBpoD_ev1p_1XpZ4yMMVv_HdBP19Gqo-2sPFMrCho"
GRACIOUS_SHEET_ID = "1ZlB2B1UhpSFtVkxRRoQG7uU56iBZTwZ6GGd7Mudl0z8"

DATA_WORKBOOK_HEADERS = [
    "PO Number",
    "AWB No.",
    "EDD",
    "Vendor",
    "Appointment ID",
    "Scheduled Date",
    "Reporting Time",
    "Status",
]

ALLOWED_PLATFORM_VALUES = {"scootsy", "instamart"}
PENDING_APPOINTMENT_VALUES = {"", "na", "n/a", "none", "no slot available"}

TRACKER_DATE_FORMAT = "%d-%b-%y"


@dataclass(frozen=True, slots=True)
class TrackerDefinition:
    key: str
    display_name: str
    sheet_id: str
    platform_column_index: int
    status_column_index: int
    appointment_column_index: int
    edd_column_index: int
    po_column_index: int
    po_lookup_column: str
    scheduled_date_column: str
    appointment_id_column: str
    reporting_time_column: str


TRACKER_DEFINITIONS: dict[str, TrackerDefinition] = {
    "omkara": TrackerDefinition(
        key="omkara",
        display_name="Omkara",
        sheet_id=OMKARA_SHEET_ID,
        platform_column_index=5,
        status_column_index=10,
        appointment_column_index=12,
        edd_column_index=11,
        po_column_index=22,
        po_lookup_column="W",
        scheduled_date_column="M",
        appointment_id_column="P",
        reporting_time_column="Q",
    ),
    "gracious": TrackerDefinition(
        key="gracious",
        display_name="Gracious",
        sheet_id=GRACIOUS_SHEET_ID,
        platform_column_index=6,
        status_column_index=13,
        appointment_column_index=15,
        edd_column_index=22,
        po_column_index=30,
        po_lookup_column="AE",
        scheduled_date_column="P",
        appointment_id_column="AB",
        reporting_time_column="AD",
    ),
}


def setup_instamart_logging(logger_name: str, log_prefix: str) -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"{log_prefix}_{datetime.now():%Y%m%d}.log"

    logger = logging.getLogger(logger_name)
    logger.setLevel(logging.DEBUG)
    if logger.handlers:
        attach_monitor_handler(logger, source=log_prefix)
        return logger

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s | %(levelname)-8s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )

    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(errors="replace")
    except Exception:
        pass

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(
        logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%H:%M:%S")
    )

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    attach_monitor_handler(logger, source=log_prefix)
    return logger


def create_google_client() -> gspread.Client:
    return gspread.oauth(
        credentials_filename=str(GOOGLE_CLIENT_SECRET),
        authorized_user_filename=str(GOOGLE_AUTH_TOKEN),
    )


def current_tracker_sheet_names(reference_date: datetime | None = None) -> dict[str, str]:
    now = reference_date or datetime.now()
    return {
        "omkara": now.strftime("%B"),
        "gracious": now.strftime("%b. %y"),
    }


def normalize_sheet_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_po_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return normalize_sheet_text(value)


def normalize_status(value: Any) -> str:
    return normalize_sheet_text(value).lower().replace(" ", "")


def normalize_platform(value: Any) -> str:
    return normalize_sheet_text(value).lower()


def is_pending_appointment_value(value: Any) -> bool:
    return normalize_sheet_text(value).lower() in PENDING_APPOINTMENT_VALUES


def safe_row_value(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else ""


def extract_instamart_rows_from_values(
    values: list[list[Any]], tracker_definition: TrackerDefinition
) -> tuple[list[dict[str, str]], int]:
    extracted_rows: list[dict[str, str]] = []
    skipped_missing_fields = 0

    for row_number, row in enumerate(values[1:], start=2):
        platform_value = normalize_platform(
            safe_row_value(row, tracker_definition.platform_column_index)
        )
        status_value = normalize_status(safe_row_value(row, tracker_definition.status_column_index))
        appointment_value = safe_row_value(row, tracker_definition.appointment_column_index)
        po_number = normalize_po_number(safe_row_value(row, tracker_definition.po_column_index))
        edd = normalize_sheet_text(safe_row_value(row, tracker_definition.edd_column_index))

        if platform_value not in ALLOWED_PLATFORM_VALUES:
            continue

        if status_value != "intransit":
            continue

        if not is_pending_appointment_value(appointment_value):
            continue

        if not po_number or not edd:
            skipped_missing_fields += 1
            continue

        extracted_rows.append(
            {
                "PO Number": po_number,
                "EDD": edd,
                "_tracker": tracker_definition.display_name,
                "_row_number": str(row_number),
            }
        )

    return extracted_rows, skipped_missing_fields


def dedupe_rows_by_po_number(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], int]:
    seen_po_numbers: set[str] = set()
    unique_rows: list[dict[str, str]] = []
    duplicate_count = 0

    for row in rows:
        po_number = normalize_po_number(row.get("PO Number"))
        edd = normalize_sheet_text(row.get("EDD"))
        if not po_number or not edd:
            continue
        if po_number in seen_po_numbers:
            duplicate_count += 1
            continue
        seen_po_numbers.add(po_number)
        unique_rows.append({"PO Number": po_number, "EDD": edd})

    return unique_rows, duplicate_count


def _header_map(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column_number in range(1, worksheet.max_column + 1):
        header_value = normalize_sheet_text(worksheet.cell(row=1, column=column_number).value)
        if header_value:
            headers[header_value] = column_number
    return headers


def ensure_data_workbook_headers(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, int]:
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        for column_number, header in enumerate(DATA_WORKBOOK_HEADERS, start=1):
            worksheet.cell(row=1, column=column_number).value = header
        return {header: index for index, header in enumerate(DATA_WORKBOOK_HEADERS, start=1)}

    headers = _header_map(worksheet)
    next_column = worksheet.max_column + 1
    for header in DATA_WORKBOOK_HEADERS:
        if header not in headers:
            worksheet.cell(row=1, column=next_column).value = header
            headers[header] = next_column
            next_column += 1
    return headers


def append_instamart_rows(
    rows: list[dict[str, str]],
    excel_path: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> dict[str, int]:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

    headers = ensure_data_workbook_headers(worksheet)
    existing_po_numbers: set[str] = set()
    seen_incoming: set[str] = set()

    for row_number in range(2, worksheet.max_row + 1):
        po_number = normalize_po_number(worksheet.cell(row=row_number, column=headers["PO Number"]).value)
        if po_number:
            existing_po_numbers.add(po_number)

    rows_to_append: list[dict[str, str]] = []
    skipped_duplicate = 0
    skipped_invalid = 0

    for row in rows:
        po_number = normalize_po_number(row.get("PO Number"))
        edd = normalize_sheet_text(row.get("EDD"))

        if not po_number or not edd:
            skipped_invalid += 1
            logger.warning("Skipping invalid Instamart row: %s", row)
            continue

        if po_number in existing_po_numbers or po_number in seen_incoming:
            skipped_duplicate += 1
            continue

        rows_to_append.append({"PO Number": po_number, "EDD": edd})
        existing_po_numbers.add(po_number)
        seen_incoming.add(po_number)

    if not dry_run:
        next_row = worksheet.max_row + 1
        for row in rows_to_append:
            worksheet.cell(row=next_row, column=headers["PO Number"]).value = row["PO Number"]
            worksheet.cell(row=next_row, column=headers["EDD"]).value = row["EDD"]
            next_row += 1

        if rows_to_append:
            workbook.save(excel_path)

    workbook.close()
    return {
        "added": len(rows_to_append) if not dry_run else 0,
        "skipped_duplicate": skipped_duplicate,
        "skipped_invalid": skipped_invalid,
    }


def parse_excel_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel_serial(value)
            if isinstance(parsed, datetime):
                return parsed
            if isinstance(parsed, date):
                return datetime.combine(parsed, time.min)
        except Exception:
            pass

    text = normalize_sheet_text(value).replace(",", " ")
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d %b %Y",
        "%d %b %y",
        "%d-%B-%Y",
        "%d-%B-%y",
        "%d %B %Y",
        "%d %B %y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def format_scheduled_date_for_tracker(value: Any) -> str:
    parsed = parse_excel_date(value)
    if not parsed:
        return normalize_sheet_text(value)
    return parsed.strftime(TRACKER_DATE_FORMAT)


def load_instamart_rows_ready_for_push(
    excel_path: Path,
    logger: logging.Logger,
) -> list[dict[str, Any]]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_path}")

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    worksheet = workbook.active
    headers = _header_map(worksheet)

    if "PO Number" not in headers:
        workbook.close()
        raise KeyError(f"{excel_path.name} is missing required column: PO Number")

    po_column = headers["PO Number"]
    awb_column = headers.get("AWB No.")
    appointment_id_column = headers.get("Appointment ID")
    scheduled_date_column = headers.get("Scheduled Date")
    reporting_time_column = headers.get("Reporting Time")
    status_column = headers.get("Status")

    rows_by_po_number: dict[str, dict[str, Any]] = {}

    for row_number in range(2, worksheet.max_row + 1):
        po_number = normalize_po_number(worksheet.cell(row=row_number, column=po_column).value)
        if not po_number:
            continue

        awb_value = (
            normalize_sheet_text(worksheet.cell(row=row_number, column=awb_column).value)
            if awb_column
            else ""
        )
        if awb_value:
            continue

        appointment_id = (
            normalize_sheet_text(worksheet.cell(row=row_number, column=appointment_id_column).value)
            if appointment_id_column
            else ""
        )
        scheduled_date = (
            worksheet.cell(row=row_number, column=scheduled_date_column).value
            if scheduled_date_column
            else ""
        )
        reporting_time = (
            normalize_sheet_text(worksheet.cell(row=row_number, column=reporting_time_column).value)
            if reporting_time_column
            else ""
        )
        status = (
            normalize_sheet_text(worksheet.cell(row=row_number, column=status_column).value)
            if status_column
            else ""
        )

        if not appointment_id and not normalize_sheet_text(scheduled_date) and not reporting_time:
            continue

        rows_by_po_number[po_number] = {
            "row_number": row_number,
            "po_number": po_number,
            "appointment_id": appointment_id,
            "scheduled_date": scheduled_date,
            "reporting_time": reporting_time,
            "status": status,
        }

    workbook.close()
    logger.info(
        "Loaded %s Instamart row(s) with appointment data from %s",
        len(rows_by_po_number),
        excel_path,
    )
    return list(rows_by_po_number.values())


def build_row_number_map(column_values: list[Any]) -> dict[str, int]:
    row_map: dict[str, int] = {}
    for row_index, value in enumerate(column_values, start=1):
        if isinstance(value, list):
            value = value[0] if value else ""
        normalized_value = normalize_po_number(value)
        if normalized_value and normalized_value not in row_map:
            row_map[normalized_value] = row_index
    return row_map


def column_letter_to_index(column_letter: str) -> int:
    total = 0
    for char in str(column_letter).upper():
        total = (total * 26) + (ord(char) - 64)
    return total

"""
╔══════════════════════════════════════════════════════════════════╗
║  POD Uploading Data Feeder — Scenario P Pipeline                ║
║                                                                  ║
║  Automates task-list generation for POD-only uploads:           ║
║    1. Scans POD_FILES/ folder for available POD documents       ║
║    2. Extracts numeric Tracking ID from messy filenames         ║
║    3. Cross-references with Order-wise.csv to find PO Numbers   ║
║    4. Appends matched rows to Filflo_Tasks.xlsx (Scenario P)   ║
║                                                                  ║
║  Unlike data_feeder.py (Scenario D), this module:              ║
║    - Does NOT need Google Sheets (no delivery dates involved)  ║
║    - Works purely from local files (POD folder + CSV)          ║
║    - Creates rows with blank Delivery Date (POD-only upload)   ║
║                                                                  ║
║  Usage:                                                          ║
║    python pod_feeder.py                                         ║
║    python pod_feeder.py --dry-run   (preview without writing)  ║
║    python pod_feeder.py --csv "path/to/Order-wise.csv"        ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from filflo_monitor_bus import attach_monitor_handler


# ═══════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════

BOT_FOLDER       = Path(__file__).resolve().parent
POD_FOLDER       = BOT_FOLDER / "POD_FILES"
UPLOADED_SUBDIR  = "_uploaded"                    # subdirectory to skip
DEFAULT_CSV_PATH = BOT_FOLDER / "Order-wise.csv"
EXCEL_PATH       = BOT_FOLDER / "Filflo_Tasks.xlsx"
LOG_DIR          = BOT_FOLDER / "logs"

# Valid POD file extensions
VALID_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".jfif", ".png", ".csv", ".xlsx", ".xls"}

# CSV column names (verified from actual CSV headers)
CSV_COL_PO_NUMBER    = "PO Number/Order No"     # Column L  (12th)
CSV_COL_ORDER_TYPE   = "Platform/Order Type"     # Column O  (15th)
CSV_COL_TRACKING     = "Tracking number"         # Column BC (55th)

# Excel output structure (must match Filflo_Tasks.xlsx)
EXCEL_HEADERS = ["PO Number", "Order Type", "Delivery Date", "Tracking ID", "Status"]


# ═══════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════

def setup_pod_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"pod_feeder_{datetime.now():%Y%m%d}.log"

    logger = logging.getLogger("PODFeeder")
    logger.setLevel(logging.DEBUG)

    if logger.handlers:
        attach_monitor_handler(logger, source="pod_feeder")
        return logger

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%H:%M:%S"
    ))

    logger.addHandler(fh)
    logger.addHandler(ch)
    attach_monitor_handler(logger, source="pod_feeder")
    return logger


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1: SCAN — Read POD_FILES/ folder & extract Tracking IDs
# ═══════════════════════════════════════════════════════════════════════

def extract_tracking_from_filename(filename: str) -> str:
    """
    Extract the numeric Tracking ID from a messy POD filename.

    Filename patterns observed in the real POD_FILES/ folder:
    ────────────────────────────────────────────────────────────
      "300078320 a.jpg"     → 300078320   (space + letter suffix)
      "300085360a.pdf"      → 300085360   (letter glued to number)
      "296237047.pdf"       → 296237047   (clean numeric)
      "3002215123.jpeg"     → 3002215123  (clean numeric)
      "2576310012746.pdf"   → 2576310012746 (long numeric)
      "123456 - copy.pdf"   → 123456      (copy suffix with dash)
      "2PH8B4YR.jpg"        → None        (alphanumeric — NOT a valid Tracking ID)

    Regex Strategy:
    ────────────────
    1. Strip the file extension to get the stem.
    2. Match the LEADING numeric sequence from the start of the stem.
       Pattern: ^(\d{6,})
       - \d{6,} = at least 6 consecutive digits starting from position 0.
       - This captures the Tracking ID and ignores any trailing junk
         (spaces, letters like 'a', ' - copy', etc.).
    3. Minimum 6 digits ensures we don't match single random numbers
       while catching all real Tracking IDs (typically 9-13 digits).
    4. Files starting with letters (like '2PH8B4YR') won't match
       because the first char must be a digit followed by 5+ more digits.

    Returns:
        The extracted numeric Tracking ID string, or "" if no valid ID found.
    """
    # Get the filename stem (without extension)
    stem = Path(filename).stem.strip()

    if not stem:
        return ""

    # ── Primary regex: leading numeric sequence (minimum 6 digits) ───
    # ^(\d{6,})  →  starts at beginning, captures 6+ digits
    # This handles:
    #   "300078320 a"  → captures "300078320"  (stops at space)
    #   "300085360a"   → captures "300085360"  (stops at letter 'a')
    #   "123456 - copy"→ captures "123456"     (stops at space)
    #   "296237047"    → captures "296237047"  (full match)
    match = re.match(r"^(\d{6,})", stem)

    if match:
        return match.group(1)

    # No valid numeric Tracking ID found
    return ""


def scan_pod_folder(pod_folder: Path, logger: logging.Logger) -> pd.DataFrame:
    """
    Scan the POD_FILES/ directory and extract Tracking IDs from filenames.

    Rules:
    - Only process files with valid extensions (.pdf, .jpg, .jpeg, .png, etc.)
    - Skip the _uploaded/ subdirectory entirely
    - Skip hidden files (starting with '.')
    - Extract numeric Tracking ID from each filename using regex
    - Log warnings for files that don't yield a valid Tracking ID

    Returns:
        DataFrame with columns: [tracking_id, filename, filepath]
    """
    logger.info(f"[SCAN] Scanning POD folder: {pod_folder}")

    if not pod_folder.exists():
        raise FileNotFoundError(f"POD folder not found: {pod_folder}")

    records = []
    skipped_invalid = []
    skipped_extension = []

    for entry in pod_folder.iterdir():
        # ── Skip subdirectories (especially _uploaded/) ──────────────
        if entry.is_dir():
            logger.debug(f"[SCAN] Skipping directory: {entry.name}")
            continue

        # ── Skip hidden files ────────────────────────────────────────
        if entry.name.startswith("."):
            continue

        # ── Validate file extension ──────────────────────────────────
        if entry.suffix.lower() not in VALID_EXTENSIONS:
            skipped_extension.append(entry.name)
            continue

        # ── Extract Tracking ID from filename ────────────────────────
        tracking_id = extract_tracking_from_filename(entry.name)

        if tracking_id:
            records.append({
                "tracking_id": tracking_id,
                "filename": entry.name,
                "filepath": str(entry),
            })
        else:
            skipped_invalid.append(entry.name)

    # ── Logging ──────────────────────────────────────────────────────
    logger.info(f"[SCAN] Total files found: {len(records) + len(skipped_invalid) + len(skipped_extension)}")
    logger.info(f"[SCAN] Valid Tracking IDs extracted: {len(records)}")

    if skipped_invalid:
        logger.warning(
            f"[SCAN] {len(skipped_invalid)} file(s) could not yield a Tracking ID "
            f"(non-numeric or too short): {skipped_invalid[:15]}"
            f"{'...' if len(skipped_invalid) > 15 else ''}"
        )

    if skipped_extension:
        logger.debug(f"[SCAN] {len(skipped_extension)} file(s) with unsupported extension skipped.")

    # ── Handle duplicate Tracking IDs (same ID, multiple files) ──────
    df = pd.DataFrame(records)
    if not df.empty:
        dupes = df[df.duplicated(subset="tracking_id", keep="first")]
        if len(dupes) > 0:
            logger.warning(
                f"[SCAN] {len(dupes)} duplicate Tracking ID(s) found in POD folder. "
                f"Keeping first occurrence only. Duplicates: "
                f"{dupes['filename'].tolist()[:10]}"
            )
            df = df.drop_duplicates(subset="tracking_id", keep="first")

    logger.info(f"[SCAN] Final unique Tracking IDs from folder: {len(df)}")
    return df


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2: CROSS-REFERENCE — Match Tracking IDs with CSV
# ═══════════════════════════════════════════════════════════════════════

def _clean_csv_tracking(val) -> str:
    """
    Clean tracking ID from CSV: handle NaN, float→int conversion, whitespace.

    CSV/Excel often reads pure-numeric strings as floats (e.g., 6001029148.0).
    This converts them back to clean integer strings.
    """
    if pd.isna(val) or str(val).strip().lower() in ("nan", "", "none", "#n/a"):
        return ""
    s = str(val).strip()
    # Float to int: "6001029148.0" → "6001029148"
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    return s


def cross_reference_csv(
    pod_data: pd.DataFrame,
    csv_path: Path,
    logger: logging.Logger,
) -> pd.DataFrame:
    """
    Match POD folder Tracking IDs against the Order-wise CSV's Tracking Number
    column to find the corresponding PO Number and Order Type.

    Merge Strategy:
    ────────────────
    - INNER JOIN: pod_data.tracking_id == csv.tracking_number
    - Only rows that exist in BOTH the folder AND the CSV are kept.
    - Unmatched files (in folder but NOT in CSV) are logged as warnings.

    Returns:
        DataFrame with columns: [tracking_id, filename, filepath, po_number, order_type]
    """
    logger.info(f"[XREF] Loading CSV: {csv_path}")

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    if pod_data.empty:
        logger.warning("[XREF] No Tracking IDs from POD folder. Nothing to cross-reference.")
        return pd.DataFrame(columns=["tracking_id", "filename", "filepath", "po_number", "order_type"])

    # ── Read CSV ─────────────────────────────────────────────────────
    df_csv = pd.read_csv(csv_path, encoding="utf-8-sig", on_bad_lines="skip")

    # Drop unnamed/empty columns (CSV often has trailing commas)
    df_csv = df_csv.loc[:, ~df_csv.columns.str.startswith("Unnamed")]
    df_csv = df_csv.loc[:, df_csv.columns.str.strip() != ""]

    logger.info(f"[XREF] Total rows in CSV: {len(df_csv)}")

    # ── Validate required columns ────────────────────────────────────
    required = [CSV_COL_PO_NUMBER, CSV_COL_TRACKING]
    missing = [c for c in required if c not in df_csv.columns]
    if missing:
        raise KeyError(f"CSV missing required columns: {missing}. Available: {list(df_csv.columns)}")

    # ── Extract & clean relevant columns ─────────────────────────────
    csv_subset = df_csv[[CSV_COL_PO_NUMBER, CSV_COL_ORDER_TYPE, CSV_COL_TRACKING]].copy()
    csv_subset.columns = ["po_number", "order_type", "tracking_number"]

    csv_subset["tracking_number"] = csv_subset["tracking_number"].apply(_clean_csv_tracking)
    csv_subset["po_number"] = csv_subset["po_number"].astype(str).str.strip()
    csv_subset["order_type"] = csv_subset["order_type"].astype(str).str.strip()

    # Drop rows with empty tracking
    csv_subset = csv_subset[csv_subset["tracking_number"] != ""].copy()

    # Drop duplicate tracking numbers in CSV (keep first occurrence)
    csv_subset = csv_subset.drop_duplicates(subset="tracking_number", keep="first")

    logger.info(f"[XREF] CSV rows with valid tracking numbers: {len(csv_subset)}")

    # ── INNER JOIN: folder tracking_id ↔ CSV tracking_number ────────
    #
    # This is the core cross-reference:
    #   POD folder has:  tracking_id (extracted from filename)
    #   CSV has:         tracking_number, po_number, order_type
    #
    # INNER JOIN ensures we ONLY keep rows that exist in BOTH sources.
    # Files without a CSV match → not included → logged as warnings.
    #
    merged = pod_data.merge(
        csv_subset,
        left_on="tracking_id",
        right_on="tracking_number",
        how="inner",
    )

    # ── Identify & log unmatched files ───────────────────────────────
    matched_ids = set(merged["tracking_id"])
    all_pod_ids = set(pod_data["tracking_id"])
    unmatched_ids = all_pod_ids - matched_ids

    logger.info(f"[XREF] Matched: {len(merged)} | Unmatched: {len(unmatched_ids)}")

    if unmatched_ids:
        # Get filenames for unmatched IDs (for helpful logging)
        unmatched_files = pod_data[pod_data["tracking_id"].isin(unmatched_ids)][["tracking_id", "filename"]]
        unmatched_list = unmatched_files.to_dict("records")

        logger.warning(
            f"[XREF] {len(unmatched_ids)} POD file(s) have NO matching Tracking Number in CSV. "
            f"These files will be SKIPPED:"
        )
        for item in unmatched_list[:20]:
            logger.warning(f"  ⚠ File: {item['filename']} (Tracking ID: {item['tracking_id']})")
        if len(unmatched_list) > 20:
            logger.warning(f"  ... and {len(unmatched_list) - 20} more.")

    # ── Clean up columns ─────────────────────────────────────────────
    result = merged[["tracking_id", "filename", "filepath", "po_number", "order_type"]].copy()

    return result


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3: PREPARE & EXPORT — Format for Filflo_Tasks.xlsx
# ═══════════════════════════════════════════════════════════════════════

def prepare_and_export(
    matched_data: pd.DataFrame,
    excel_path: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> dict:
    """
    Format matched rows for Filflo_Tasks.xlsx and append them.

    Scenario P format:
    ──────────────────
      Col 1 (PO Number):     from CSV match
      Col 2 (Order Type):    from CSV match
      Col 3 (Delivery Date): BLANK (Scenario P = POD-only, no delivery date)
      Col 4 (Tracking ID):   from POD filename / CSV match
      Col 5 (Status):        BLANK (pending for bot to process)

    Deduplication:
      - Skips any PO Number that already exists in the Excel file.
      - This prevents re-feeding the same POD data on repeated runs.

    Returns:
        Summary dict with counts: added, skipped_duplicate, skipped_unmatched, would_add
    """
    logger.info(f"[EXPORT] Preparing data for Filflo_Tasks.xlsx")

    if matched_data.empty:
        logger.warning("[EXPORT] No matched data to export.")
        return {"added": 0, "skipped_duplicate": 0, "skipped_unmatched": 0}

    # ── Build output DataFrame matching Excel structure ──────────────
    output = pd.DataFrame({
        "PO Number":     matched_data["po_number"],
        "Order Type":    matched_data["order_type"],
        "Delivery Date": "",             # Blank — this is Scenario P (POD upload only)
        "Tracking ID":   matched_data["tracking_id"],
        "Status":        None,           # Blank = pending for bot
    })

    logger.info(f"[EXPORT] Rows prepared for export: {len(output)}")

    # ── Read existing POs from Excel (deduplication) ─────────────────
    existing_pos = set()
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            po = ws.cell(row=row_idx, column=1).value
            if po:
                existing_pos.add(str(po).strip())
        wb.close()
    else:
        # Create new Excel with headers
        logger.info("[EXPORT] Excel file not found. Creating new one with headers.")
        wb = openpyxl.Workbook()
        ws = wb.active
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(excel_path)
        wb.close()

    logger.info(f"[EXPORT] Existing POs in Excel: {len(existing_pos)}")

    # ── Deduplicate — skip POs already in Excel ──────────────────────
    output["_is_dup"] = output["PO Number"].astype(str).str.strip().isin(existing_pos)
    duplicates = int(output["_is_dup"].sum())
    to_add = output[~output["_is_dup"]].drop(columns=["_is_dup"]).copy()

    logger.info(f"[EXPORT] New unique rows: {len(to_add)} | Duplicates skipped: {duplicates}")

    if to_add.empty:
        logger.info("[EXPORT] All rows already exist in Excel. Nothing new to add.")
        return {"added": 0, "skipped_duplicate": duplicates}

    # ── Dry run — preview only ───────────────────────────────────────
    if dry_run:
        logger.info("[EXPORT] DRY RUN — preview (not writing to Excel):")
        for _, row in to_add.head(20).iterrows():
            logger.info(
                f"  → PO: {row['PO Number']} | Type: {row['Order Type']} | "
                f"Tracking: {row['Tracking ID']} | Date: (blank — Scenario P)"
            )
        if len(to_add) > 20:
            logger.info(f"  ... and {len(to_add) - 20} more rows.")
        return {"added": 0, "skipped_duplicate": duplicates, "would_add": len(to_add)}

    # ── Append to Excel ──────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        start_row = ws.max_row + 1

        for idx, (_, row) in enumerate(to_add.iterrows()):
            r = start_row + idx
            ws.cell(row=r, column=1, value=str(row["PO Number"]).strip())
            ws.cell(row=r, column=2, value=str(row["Order Type"]).strip())
            ws.cell(row=r, column=3, value="")                              # Blank — Scenario P
            ws.cell(row=r, column=4, value=str(row["Tracking ID"]).strip())
            # Column 5 (Status) left blank = pending for bot

        wb.save(excel_path)
        wb.close()

        logger.info(
            f"[EXPORT] ✅ Appended {len(to_add)} rows to Excel "
            f"(rows {start_row} to {start_row + len(to_add) - 1})"
        )

    except PermissionError:
        logger.error("[EXPORT] ❌ Excel file is open in another program. Please close it and retry.")
        raise

    return {"added": len(to_add), "skipped_duplicate": duplicates}


# ═══════════════════════════════════════════════════════════════════════
#  MAIN PIPELINE ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════

def run_pod_feeder(
    pod_folder: Path = None,
    csv_path: Path = None,
    excel_path: Path = None,
    dry_run: bool = False,
    logger: logging.Logger = None,
) -> dict:
    """
    Run the full POD Data Feeder pipeline:

      POD_FILES/ folder (scan filenames)
        → extract numeric Tracking IDs (regex)
        → cross-reference with Order-wise.csv (inner join on Tracking Number)
        → append matched rows to Filflo_Tasks.xlsx (Scenario P format)

    Returns a summary dict.
    """
    if logger is None:
        logger = setup_pod_logging()
    if pod_folder is None:
        pod_folder = POD_FOLDER
    if csv_path is None:
        csv_path = DEFAULT_CSV_PATH
    if excel_path is None:
        excel_path = EXCEL_PATH

    logger.info("=" * 60)
    logger.info("[POD-FEEDER] Starting POD Uploading Data Feeder pipeline")
    logger.info(f"[POD-FEEDER] POD folder: {pod_folder}")
    logger.info(f"[POD-FEEDER] CSV: {csv_path}")
    logger.info(f"[POD-FEEDER] Excel: {excel_path}")
    logger.info(f"[POD-FEEDER] Dry run: {dry_run}")

    try:
        # Step 1: Scan POD folder & extract Tracking IDs
        pod_data = scan_pod_folder(pod_folder, logger)

        if pod_data.empty:
            logger.warning("[POD-FEEDER] No valid POD files found. Pipeline complete (nothing to do).")
            return {"added": 0, "skipped_duplicate": 0, "scanned_files": 0, "matched": 0, "unmatched": 0}

        # Step 2: Cross-reference with CSV
        matched = cross_reference_csv(pod_data, csv_path, logger)

        unmatched_count = len(pod_data) - len(matched)

        # Step 3: Prepare & export to Excel
        summary = prepare_and_export(matched, excel_path, logger, dry_run=dry_run)

        # Add extra stats to summary
        summary["scanned_files"] = len(pod_data)
        summary["matched"] = len(matched)
        summary["unmatched"] = unmatched_count

        logger.info(f"[POD-FEEDER] ✅ Pipeline complete: {summary}")
        return summary

    except Exception as e:
        logger.error(f"[POD-FEEDER] ❌ Pipeline failed: {e}", exc_info=True)
        raise


# ═══════════════════════════════════════════════════════════════════════
#  CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="POD Uploading Data Feeder — Scenario P Pipeline")
    parser.add_argument("--pod-folder", type=str, default=str(POD_FOLDER),
                        help="Path to POD_FILES/ folder")
    parser.add_argument("--csv", type=str, default=str(DEFAULT_CSV_PATH),
                        help="Path to Order-wise.csv")
    parser.add_argument("--excel", type=str, default=str(EXCEL_PATH),
                        help="Path to Filflo_Tasks.xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview data without writing to Excel")
    args = parser.parse_args()

    logger = setup_pod_logging()

    summary = run_pod_feeder(
        pod_folder=Path(args.pod_folder),
        csv_path=Path(args.csv),
        excel_path=Path(args.excel),
        dry_run=args.dry_run,
        logger=logger,
    )

    print(f"\n{'═' * 55}")
    print(f"  POD Data Feeder Summary:")
    print(f"  POD files scanned:    {summary.get('scanned_files', 0)}")
    print(f"  Matched with CSV:     {summary.get('matched', 0)}")
    print(f"  Unmatched (skipped):  {summary.get('unmatched', 0)}")
    print(f"  Added to Excel:       {summary.get('added', 0)}")
    print(f"  Duplicates skipped:   {summary.get('skipped_duplicate', 0)}")
    if "would_add" in summary:
        print(f"  Would add (dry run):  {summary['would_add']}")
    print(f"{'═' * 55}\n")


if __name__ == "__main__":
    main()

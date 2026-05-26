"""Create a sample Filflo tasks workbook with the expected headers."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = ["PO Number", "Order Type", "Delivery Date", "Tracking ID", "Status"]
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "Filflo_Tasks_sample.xlsx"


def create_sample_excel(output_path: str | Path | None = None, overwrite: bool = False) -> Path:
    """
    Create a sample Excel file with the standard Filflo task headers.

    Safe default: writes to ``Filflo_Tasks_sample.xlsx`` instead of overwriting
    the live workbook.
    """
    output = Path(output_path) if output_path else DEFAULT_OUTPUT
    if output.exists() and not overwrite:
        raise FileExistsError(
            f"File already exists: {output}. Pass overwrite=True to replace it."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )

    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 40

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    wb.close()
    return output


def main():
    output = create_sample_excel()
    print(f"Created: {output}")
    print("Add your PO rows and run the bot!")


if __name__ == "__main__":
    main()

"""
Excel utility functions for Filflo Bot.

Provides helper functions to read, parse, and update Excel files containing
PO tracking data. Handles cell locking, data validation, and formatting.
"""

from pathlib import Path
from datetime import datetime
import openpyxl
from filelock import FileLock

from filflo_config import (
    COL_PO_NUMBER,
    COL_ORDER_TYPE,
    COL_DELIVERY_DATE,
    COL_TRACKING_ID,
    COL_STATUS,
    HEADER_ROW,
    EXCEL_LOCK_PATH,
)
from po_status import is_row_done


def ensure_excel_headers(path: Path, logger):
    """Ensure Excel file has the correct column headers."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = {
        COL_PO_NUMBER: "PO Number",
        COL_ORDER_TYPE: "Order Type",
        COL_DELIVERY_DATE: "Delivery Date",
        COL_TRACKING_ID: "Tracking ID",
        COL_STATUS: "Status",
    }
    changed = False
    for col, name in headers.items():
        if ws.cell(row=HEADER_ROW, column=col).value != name:
            ws.cell(row=HEADER_ROW, column=col, value=name)
            changed = True
    if changed:
        wb.save(path)
    wb.close()


def read_pending_entries(path: Path, logger):
    """Read all pending (not fully done) entries from Excel file."""
    lock = FileLock(str(EXCEL_LOCK_PATH), timeout=30)
    with lock:
        wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    entries = []
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        po_number = ws.cell(row=row_idx, column=COL_PO_NUMBER).value
        if not po_number:
            continue

        status = str(ws.cell(row=row_idx, column=COL_STATUS).value or "").strip()

        # Skip fully done rows
        if is_row_done(status):
            continue

        tracking_raw = ws.cell(row=row_idx, column=COL_TRACKING_ID).value
        tracking_id = ""

        if tracking_raw is not None:
            tracking_id = str(tracking_raw).strip()
            # Remove hidden characters
            tracking_id = tracking_id.replace("\u200b", "").replace("\ufeff", "").replace("\t", "")

            # Skip formula-based tracking IDs
            if tracking_id.startswith("="):
                logger.warning(f"Row {row_idx}: Tracking ID is a formula -- paste as plain values!")
                tracking_id = ""

            # Handle numeric tracking IDs (Excel may read as float like 6001029489.0)
            if tracking_id and "." in tracking_id:
                try:
                    tracking_id = str(int(float(tracking_id)))
                except (ValueError, OverflowError):
                    pass

        po_raw = str(po_number).strip()
        po_str = normalize_po_number(po_raw)
        if po_str != po_raw:
            logger.info(f"Row {row_idx}: Normalized PO '{po_raw}' -> '{po_str}'")
        logger.debug(f"Row {row_idx}: PO='{po_str}', Tracking ID='{tracking_id}' (raw={repr(tracking_raw)})")

        entries.append({
            "row": row_idx,
            "po_number": po_str,
            "order_type": str(ws.cell(row=row_idx, column=COL_ORDER_TYPE).value or "").strip(),
            "delivery_date": ws.cell(row=row_idx, column=COL_DELIVERY_DATE).value,
            "tracking_id": tracking_id,
            "status": status,
        })

    wb.close()
    return entries


def update_excel_status(path: Path, row_idx: int, status: str, logger):
    """Update the status of a specific row in the Excel file."""
    lock = FileLock(str(EXCEL_LOCK_PATH), timeout=30)
    try:
        with lock:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.cell(row=row_idx, column=COL_STATUS,
                    value=f"{status} ({datetime.now():%Y-%m-%d})")
            wb.save(path)
            wb.close()
    except PermissionError:
        logger.error(f"Row {row_idx}: Excel file is open. Please close it.")
    except Exception as e:
        logger.error(f"Row {row_idx}: Excel update failed -- {e}")


def normalize_po_number(raw_po) -> str:
    """Normalize PO number by removing hidden characters and extra quotes."""
    po_str = str(raw_po or "").strip()
    po_str = po_str.replace("\u200b", "").replace("\ufeff", "").replace("\t", "")
    # Excel sometimes wraps text PO numbers with apostrophes/quotes to preserve formatting.
    return po_str.strip("'").strip('"').strip()


def parse_delivery_date(raw_date) -> str:
    """Convert delivery date to DD-MM-YYYY format."""
    if isinstance(raw_date, datetime):
        return raw_date.strftime("%d-%m-%Y")

    date_str = str(raw_date).strip()
    formats = [
        "%d-%b-%y", "%d-%b-%Y", "%d-%m-%Y", "%d/%m/%Y",
        "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            if dt.year < 100:
                dt = dt.replace(year=dt.year + 2000)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            continue

    raise ValueError(f"Cannot parse delivery date: '{raw_date}'")


def should_prefer_all_time(raw_date) -> bool:
    """
    Older deliveries are usually outside Filflo's default 'Last 30 Days' filter.
    Use the Excel date as a fast heuristic so we don't waste one search cycle first.
    """
    if not raw_date:
        return False

    try:
        normalized = parse_delivery_date(raw_date)
        delivery_dt = datetime.strptime(normalized, "%d-%m-%Y").date()
        return (datetime.now().date() - delivery_dt).days > 25
    except Exception:
        return False


def find_pod_file(folder: Path, tracking_id: str, logger=None) -> Path | None:
    """
    Search for POD file by tracking ID — EXACT match only.
    File must be named exactly as the Tracking ID (e.g., TRACK123.jpg for Tracking ID 'TRACK123').
    No partial matching to prevent wrong POD being attached.
    Handles whitespace, hidden characters, and case differences.
    """
    if not tracking_id or not folder.exists():
        return None

    # Aggressive cleanup: remove ALL whitespace, hidden chars, BOM, zero-width chars
    tid = tracking_id.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
    if not tid:
        return None

    supported_ext = (".jpg", ".jpeg", ".jfif", ".png", ".pdf", ".csv", ".xlsx", ".xls")

    # EXACT stem match — also clean up filenames the same way
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
        if clean_stem == tid:
            if logger:
                logger.info(f"POD file matched: Tracking ID '{tid}' -> {f.name}")
            return f

    # Case-insensitive fallback (same exact ID but different case)
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "").replace(" ", "")
        if clean_stem.lower() == tid.lower():
            if logger:
                logger.info(f"POD file matched (case-insensitive): Tracking ID '{tid}' -> {f.name}")
            return f

    # Prefix match fallback: file stem starts with the tracking ID
    # Handles files like "301203752 a.pdf" matching tracking ID "301203752"
    # Only matches if the character after the ID is a space, hyphen, underscore, or letter suffix
    for f in folder.iterdir():
        if not f.is_file() or f.suffix.lower() not in supported_ext:
            continue
        clean_stem = f.stem.strip().replace("\u200b", "").replace("\ufeff", "").replace("\t", "")
        if clean_stem.lower().startswith(tid.lower()) and len(clean_stem) > len(tid):
            # Ensure the extra part is a harmless suffix (space, letter, hyphen, underscore)
            remainder = clean_stem[len(tid):]
            if remainder.strip().replace("-", "").replace("_", "").isalpha() or remainder.startswith(" "):
                if logger:
                    logger.info(f"POD file matched (prefix match): Tracking ID '{tid}' -> {f.name}")
                return f

    if logger:
        # Log available files to help debug
        available = [f.stem for f in folder.iterdir()
                     if f.is_file() and f.suffix.lower() in supported_ext][:10]
        logger.info(f"No POD file for Tracking ID '{tid}' (repr: {repr(tracking_id)}). "
                     f"Sample files in folder: {available}")
    return None

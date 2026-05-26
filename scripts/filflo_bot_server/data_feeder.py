"""
╔══════════════════════════════════════════════════════════════════╗
║  Data Feeder — ETL Pipeline for Filflo Bot                      ║
║                                                                  ║
║  PHASE 1: Portal Orders -> Filflo_Tasks.xlsx                    ║
║  PHASE 2: Courier Trackers -> Data.xlsx                         ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import logging
import argparse
import time
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import gspread
from gspread.exceptions import APIError
from requests.exceptions import RequestException
from filflo_monitor_bus import attach_monitor_handler


# ═══════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════

BOT_FOLDER = Path(__file__).resolve().parent
LOG_DIR    = BOT_FOLDER / "logs"

# Path to Google OAuth2 credentials
GOOGLE_CLIENT_SECRET = BOT_FOLDER / "client_secret.json"
GOOGLE_AUTH_TOKEN    = BOT_FOLDER / "authorized_user.json"

# ── PHASE 1 CONFIG (Portal) ─────────────────────────────────────────────
DEFAULT_CSV_PATH      = BOT_FOLDER / "Order-wise.csv"
EXCEL_PATH_SCENARIO_D = BOT_FOLDER / "Filflo_Tasks.xlsx"
PHASE1_TRACKER_CACHE_PATH = BOT_FOLDER / "phase1_tracker_cache.csv"

CSV_COL_PO_NUMBER     = "PO Number/Order No"
CSV_COL_ORDER_STATUS  = "Order Status"
CSV_COL_ORDER_TYPE    = "Platform/Order Type"
CSV_COL_TRACKING      = "Tracking number"
GS_COL_AWB            = "AWB No."
GS_COL_CONSIGNEE      = "Consignee"
GS_COL_DELIVERY_DATE  = "Delivery Date"
GS_COL_STATUS         = "Status"
GRACIOUS_DELIVERY_DATE_COL = "D. Date"
PHASE1_ALLOWED_EXACT_STATUSES = {"delivered", "partial delivered"}
INVALID_DELIVERY_DATES = {"", "nan", "na", "n/a", "none", "nat", "null", "-"}

EXCEL_HEADERS_SCENARIO_D = ["PO Number", "Order Type", "Delivery Date", "Tracking ID", "Status"]


# ── PHASE 2 CONFIG (Courier Trackers) ───────────────────────────────────
EXCEL_PATH_COURIER = BOT_FOLDER / "Data.xlsx"
EXCEL_HEADERS_COURIER = ["PO Number", "AWB No.", "EDD", "Vendor", "Appointment ID", "Scheduled Date", "Reporting Time", "Status"]

OMKARA_SHEET_ID = "1VcGBpoD_ev1p_1XpZ4yMMVv_HdBP19Gqo-2sPFMrCho"
SKYLARK_SHEET_ID = "1zL7fue0UmtHRinhsO7tp5F3G97mWpZOgWYzQVGbKrHg"
GRACIOUS_SHEET_ID = "1ZlB2B1UhpSFtVkxRRoQG7uU56iBZTwZ6GGd7Mudl0z8"
PHASE2_INVALID_APPOINTMENT_VALUES = {"na", "no slot available", "n/a", "none", ""}
PHASE2_SKYLARK_TRACKER_TAB = "Tracker"
PHASE2_SKYLARK_CONSIGNEE_ALIASES = {"hands on traders", "hands on trade", "hot", "blinkit"}
PHASE2_SKYLARK_PENDING_STATUSES = {"intransit", "pending", "ofd"}
PHASE2_MISSING_APPOINTMENT_ID_VALUES = {"", "na", "n/a", "none"}
PHASE2_OMKARA_PENDING_STATUSES = {"intransit", "pending", "ofd"}
PHASE2_GRACIOUS_PENDING_STATUSES = {"intransit", "pending", "ofd", "podpending"}

GOOGLE_API_RETRYABLE_CODES = {429, 500, 502, 503, 504}
GOOGLE_API_MAX_ATTEMPTS = 5
GOOGLE_API_INITIAL_DELAY_SECONDS = 2.0
GOOGLE_API_REQUEST_TIMEOUT_SECONDS = (10, 45)

# ═══════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════

def setup_etl_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"data_feeder_{datetime.now():%Y%m%d}.log"

    logger = logging.getLogger("DataFeeder")
    logger.setLevel(logging.DEBUG)
    if logger.handlers:
        attach_monitor_handler(logger, source="data_feeder")
        return logger

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(errors="replace")
    except Exception:
        pass
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%H:%M:%S"))

    logger.addHandler(fh)
    logger.addHandler(ch)
    attach_monitor_handler(logger, source="data_feeder")
    return logger


def _clean_tracking_id(val) -> str:
    if pd.isna(val) or str(val).strip().lower() in ("nan", "", "none", "#n/a"):
        return ""
    s = str(val).strip()
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, TypeError, OverflowError):
            pass
    return s


def _is_retryable_google_error(exc: Exception) -> bool:
    if isinstance(exc, APIError):
        response_status = getattr(exc.response, "status_code", None)
        error_code = getattr(exc, "code", response_status)
        return bool(
            error_code in GOOGLE_API_RETRYABLE_CODES
            or response_status in GOOGLE_API_RETRYABLE_CODES
            or (response_status is not None and response_status >= 500)
        )
    return isinstance(exc, RequestException)


def _run_google_api_call(action: str, logger: logging.Logger, func):
    last_error = None
    for attempt in range(1, GOOGLE_API_MAX_ATTEMPTS + 1):
        try:
            result = func()
            if attempt > 1:
                logger.info(f"{action} succeeded on retry {attempt}/{GOOGLE_API_MAX_ATTEMPTS}")
            return result
        except Exception as exc:
            last_error = exc
            if not _is_retryable_google_error(exc) or attempt == GOOGLE_API_MAX_ATTEMPTS:
                raise
            delay = GOOGLE_API_INITIAL_DELAY_SECONDS * (2 ** (attempt - 1))
            logger.warning(
                f"{action} failed with transient error ({exc}). "
                f"Retrying in {delay:.1f}s [{attempt}/{GOOGLE_API_MAX_ATTEMPTS}]"
            )
            time.sleep(delay)
    raise last_error


def _get_gspread_client(logger: logging.Logger):
    client = _run_google_api_call(
        "[Google Sheets] Authorizing client",
        logger,
        lambda: gspread.oauth(
            credentials_filename=str(GOOGLE_CLIENT_SECRET),
            authorized_user_filename=str(GOOGLE_AUTH_TOKEN),
        ),
    )
    client.http_client.timeout = GOOGLE_API_REQUEST_TIMEOUT_SECONDS
    return client


def _get_worksheet(logger: logging.Logger, gc, sheet_id: str, worksheet_name: str):
    return _run_google_api_call(
        f"[Google Sheets] Opening worksheet '{worksheet_name}'",
        logger,
        lambda: gc.open_by_key(sheet_id).worksheet(worksheet_name),
    )


def _get_all_values(logger: logging.Logger, worksheet, worksheet_name: str):
    return _run_google_api_call(
        f"[Google Sheets] Reading worksheet '{worksheet_name}'",
        logger,
        worksheet.get_all_values,
    )


def _get_row_values(logger: logging.Logger, worksheet, worksheet_name: str, row_number: int):
    return _run_google_api_call(
        f"[Google Sheets] Reading row {row_number} from worksheet '{worksheet_name}'",
        logger,
        lambda: worksheet.row_values(row_number),
    )


def _get_col_values(
    logger: logging.Logger,
    worksheet,
    worksheet_name: str,
    col_number: int,
    column_name: str,
):
    return _run_google_api_call(
        f"[Google Sheets] Reading column '{column_name}' from worksheet '{worksheet_name}'",
        logger,
        lambda: worksheet.col_values(col_number),
    )


def _normalize_google_headers(headers: list[str]) -> dict[str, int]:
    normalized_headers = {}
    for idx, header in enumerate(headers, start=1):
        clean_header = header.strip() if header.strip() else f"_unnamed_{idx - 1}"
        normalized_headers.setdefault(clean_header, idx)
    return normalized_headers


def _read_selected_sheet_columns(
    logger: logging.Logger,
    gc,
    sheet_id: str,
    worksheet_name: str,
    required_cols: list[str],
) -> pd.DataFrame:
    worksheet = _get_worksheet(logger, gc, sheet_id, worksheet_name)
    headers = _get_row_values(logger, worksheet, worksheet_name, 1)
    if not headers:
        return pd.DataFrame(columns=required_cols)

    normalized_headers = _normalize_google_headers(headers)
    missing_cols = [col for col in required_cols if col not in normalized_headers]
    if missing_cols:
        raise KeyError(f"Worksheet '{worksheet_name}' is missing required column(s): {', '.join(missing_cols)}")

    extracted_columns = {}
    max_rows = 0
    for col_name in required_cols:
        col_values = _get_col_values(
            logger,
            worksheet,
            worksheet_name,
            normalized_headers[col_name],
            col_name,
        )
        data_values = col_values[1:] if col_values else []
        extracted_columns[col_name] = data_values
        max_rows = max(max_rows, len(data_values))

    for col_name in required_cols:
        values = extracted_columns[col_name]
        if len(values) < max_rows:
            extracted_columns[col_name] = values + [""] * (max_rows - len(values))

    return pd.DataFrame(extracted_columns)


def _is_phase1_allowed_status(status: str) -> bool:
    normalized = str(status).strip().lower()
    return normalized in PHASE1_ALLOWED_EXACT_STATUSES or ("pod" in normalized and "pending" in normalized)


def _has_valid_delivery_date(value) -> bool:
    return str(value).strip().lower() not in INVALID_DELIVERY_DATES


def _get_phase1_source_specs() -> list[dict]:
    now = datetime.now()
    previous_month = now.replace(day=1) - timedelta(days=1)
    return [
        {
            "source": "omkara",
            "sheet_id": OMKARA_SHEET_ID,
            "worksheet_name": now.strftime("%B"),
            "awb_col": GS_COL_AWB,
            "delivery_date_col": GS_COL_DELIVERY_DATE,
            "status_col": GS_COL_STATUS,
            "priority": 1,
        },
        {
            "source": "gracious_current",
            "sheet_id": GRACIOUS_SHEET_ID,
            "worksheet_name": now.strftime("%b. %y"),
            "awb_col": GS_COL_AWB,
            "delivery_date_col": GRACIOUS_DELIVERY_DATE_COL,
            "status_col": GS_COL_STATUS,
            "priority": 2,
        },
        {
            "source": "gracious_previous",
            "sheet_id": GRACIOUS_SHEET_ID,
            "worksheet_name": previous_month.strftime("%b. %y"),
            "awb_col": GS_COL_AWB,
            "delivery_date_col": GRACIOUS_DELIVERY_DATE_COL,
            "status_col": GS_COL_STATUS,
            "priority": 3,
        },
        {
            "source": "skylark",
            "sheet_id": SKYLARK_SHEET_ID,
            "worksheet_name": "Tracker",
            "awb_col": GS_COL_AWB,
            "delivery_date_col": GS_COL_DELIVERY_DATE,
            "status_col": GS_COL_STATUS,
            "priority": 4,
        },
    ]


# ═══════════════════════════════════════════════════════════════════════
#  PHASE 1: PORTAL EXTRACTION & MERGE
# ═══════════════════════════════════════════════════════════════════════

def extract_from_csv(csv_path: Path, logger: logging.Logger) -> pd.DataFrame:
    logger.info(f"[PHASE 1] Reading CSV: {csv_path}")
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")
    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", on_bad_lines="skip")
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, encoding="latin1", on_bad_lines="skip")
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    
    df[CSV_COL_ORDER_STATUS] = df[CSV_COL_ORDER_STATUS].astype(str).str.strip().str.lower()
    df_filtered = df[df[CSV_COL_ORDER_STATUS] == "in_transit"].copy()
    
    result = df_filtered[[CSV_COL_PO_NUMBER, CSV_COL_ORDER_TYPE, CSV_COL_TRACKING]].copy()
    result.columns = ["po_number", "order_type", "tracking_number"]
    result["tracking_number"] = result["tracking_number"].apply(_clean_tracking_id)
    return result[(result["tracking_number"] != "") & (result["po_number"] != "")]

def extract_from_google_sheet(logger: logging.Logger) -> pd.DataFrame:
    tracker_cols = ["awb_no", "delivery_date", "status", "source", "priority"]
    source_specs = _get_phase1_source_specs()
    try:
        gc = _get_gspread_client(logger)
        frames = []
        for spec in source_specs:
            source_df = _read_selected_sheet_columns(
                logger,
                gc,
                spec["sheet_id"],
                spec["worksheet_name"],
                [spec["awb_col"], spec["delivery_date_col"], spec["status_col"]],
            )
            source_df = source_df.rename(
                columns={
                    spec["awb_col"]: "awb_no",
                    spec["delivery_date_col"]: "delivery_date",
                    spec["status_col"]: "status",
                }
            )
            source_df["awb_no"] = source_df["awb_no"].apply(_clean_tracking_id)
            source_df["delivery_date"] = source_df["delivery_date"].astype(str).str.strip()
            source_df["status"] = source_df["status"].astype(str).str.strip()
            source_df["source"] = spec["source"]
            source_df["priority"] = spec["priority"]

            valid_rows = source_df[
                (source_df["awb_no"] != "")
                & source_df["delivery_date"].apply(_has_valid_delivery_date)
                & source_df["status"].apply(_is_phase1_allowed_status)
            ].copy()
            frames.append(valid_rows[tracker_cols])
            logger.info(
                f"[PHASE 1] Source '{spec['source']}' ({spec['worksheet_name']}) "
                f"rows after status filter: {len(valid_rows)}"
            )

        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=tracker_cols)
        combined.sort_values(["priority"], inplace=True, kind="stable")
        combined = combined.drop_duplicates(subset=["awb_no"], keep="first")
        combined.to_csv(PHASE1_TRACKER_CACHE_PATH, index=False, encoding="utf-8-sig")
    except Exception as exc:
        if not PHASE1_TRACKER_CACHE_PATH.exists():
            raise
        logger.warning(
            f"[PHASE 1] Live tracker sheet read failed ({exc}). "
            f"Using cached tracker data from {PHASE1_TRACKER_CACHE_PATH.name}"
        )
        combined = pd.read_csv(PHASE1_TRACKER_CACHE_PATH, dtype=str, encoding="utf-8-sig").fillna("")
        missing_cols = [col for col in tracker_cols if col not in combined.columns]
        if missing_cols:
            raise KeyError(
                f"Cached tracker data is missing required column(s): {', '.join(missing_cols)}"
            )

    logger.info(f"[PHASE 1] Combined tracker rows after source-priority dedupe: {len(combined)}")
    return combined[["awb_no", "delivery_date"]].copy()

def transform_merge(csv_data: pd.DataFrame, gsheet_data: pd.DataFrame, logger: logging.Logger) -> pd.DataFrame:
    merged = csv_data.merge(gsheet_data, left_on="tracking_number", right_on="awb_no", how="left")
    output = pd.DataFrame({
        "PO Number":     merged["po_number"],
        "Order Type":    merged["order_type"],
        "Delivery Date": merged["delivery_date"].fillna(""),
        "Tracking ID":   "",
        "Status":        None,
    })
    return output[output["Delivery Date"].apply(_has_valid_delivery_date)].copy()

def load_to_filflo_tasks(new_data: pd.DataFrame, excel_path: Path, logger: logging.Logger, dry_run: bool) -> dict:
    if new_data.empty: return {"added": 0, "skipped_duplicate": 0}
    
    # Open or Create the workbook EXACTLY ONCE
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS_SCENARIO_D)

    existing_pos = set()
    for row_idx in range(2, ws.max_row + 1):
        po = ws.cell(row=row_idx, column=1).value
        if po: existing_pos.add(str(po).strip())

    new_data["_is_duplicate"] = new_data["PO Number"].astype(str).str.strip().isin(existing_pos)
    duplicates = new_data["_is_duplicate"].sum()
    to_add = new_data[~new_data["_is_duplicate"]].drop(columns=["_is_duplicate"]).copy()[EXCEL_HEADERS_SCENARIO_D]

    if dry_run or to_add.empty:
        wb.close()
        return {"added": 0, "skipped_duplicate": int(duplicates)}

    for row in dataframe_to_rows(to_add, index=False, header=False):
        ws.append(row)
        
    wb.save(excel_path)
    wb.close()
    
    logger.info(f"[PHASE 1] Appended {len(to_add)} rows to {excel_path.name}")
    return {"added": len(to_add), "skipped_duplicate": int(duplicates)}


# ═══════════════════════════════════════════════════════════════════════
#  PHASE 2: COURIER EXTRACTION
# ═══════════════════════════════════════════════════════════════════════

def extract_courier_data(logger: logging.Logger) -> pd.DataFrame:
    logger.info("[PHASE 2] Starting Courier data extraction...")
    gc = _get_gspread_client(logger)

    now = datetime.now()
    omkara_tab_name = now.strftime("%B")
    gracious_tab_name = now.strftime("%b. %y")

    all_data = []

    # ── 1. Omkara Tracker ──
    try:
        logger.info(f"[PHASE 2] Reading Omkara: Tab '{omkara_tab_name}'")
        omkara_ws = _get_worksheet(logger, gc, OMKARA_SHEET_ID, omkara_tab_name)
        omkara_raw = _get_all_values(logger, omkara_ws, omkara_tab_name)
        omkara_added = 0

        for row in omkara_raw[1:]: 
            row += [""] * (30 - len(row))
            filter_val = str(row[5]).strip()
            status_val = str(row[10]).strip().lower().replace(" ", "") 
            app_id_val = str(row[15]).strip().lower()
            awb = _clean_tracking_id(row[2])
            po = str(row[22]).strip()
            
            if (
                filter_val == "Hands on Traders"
                and status_val in PHASE2_OMKARA_PENDING_STATUSES
                and app_id_val in PHASE2_MISSING_APPOINTMENT_ID_VALUES
                and awb
                and po
            ):
                all_data.append([po, awb, str(row[11]).strip(), "Delhivery", "", "", "", ""])
                omkara_added += 1
        logger.info(f"[PHASE 2] Finished Omkara processing. Added {omkara_added} rows.")
    except Exception as e:
        logger.error(f"[PHASE 2] Failed processing Omkara Tracker: {e}")

    # ── 2. Gracious Tracker ──
    try:
        logger.info(f"[PHASE 2] Reading Gracious: Tab '{gracious_tab_name}'")
        gracious_ws = _get_worksheet(logger, gc, GRACIOUS_SHEET_ID, gracious_tab_name)
        gracious_raw = _get_all_values(logger, gracious_ws, gracious_tab_name)
        gracious_added = 0

        for row in gracious_raw[1:]:
            row += [""] * (40 - len(row))
            filter_val = str(row[6]).strip()
            status_val = str(row[13]).strip().lower().replace(" ", "")
            app_id_val = str(row[27]).strip().lower()
            awb = _clean_tracking_id(row[3])
            po = str(row[30]).strip()
            
            if (
                filter_val == "HOT"
                and status_val in PHASE2_GRACIOUS_PENDING_STATUSES
                and app_id_val in PHASE2_MISSING_APPOINTMENT_ID_VALUES
                and awb
                and po
            ):
                all_data.append([po, awb, str(row[22]).strip(), "Delhivery", "", "", "", ""])
                gracious_added += 1
        logger.info(f"[PHASE 2] Finished Gracious processing. Added {gracious_added} rows.")
    except Exception as e:
        logger.error(f"[PHASE 2] Failed processing Gracious Tracker: {e}")

    # ?? 3. Skylark Tracker ??
    try:
        logger.info(f"[PHASE 2] Reading Skylark: Tab '{PHASE2_SKYLARK_TRACKER_TAB}'")
        skylark_ws = _get_worksheet(logger, gc, SKYLARK_SHEET_ID, PHASE2_SKYLARK_TRACKER_TAB)
        skylark_raw = _get_all_values(logger, skylark_ws, PHASE2_SKYLARK_TRACKER_TAB)
        skylark_added = 0

        for row in skylark_raw[1:]:
            row += [""] * (30 - len(row))
            consignee_val = str(row[5]).strip().lower()
            status_val = str(row[10]).strip().lower().replace(" ", "")
            app_id_val = str(row[13]).strip().lower()
            awb = _clean_tracking_id(row[2])
            po = str(row[20]).strip()
            vendor = str(row[3]).strip() or "Skylark Logistics"

            if (
                consignee_val in PHASE2_SKYLARK_CONSIGNEE_ALIASES
                and status_val in PHASE2_SKYLARK_PENDING_STATUSES
                and app_id_val in PHASE2_MISSING_APPOINTMENT_ID_VALUES
                and awb
                and po
            ):
                all_data.append([po, awb, str(row[11]).strip(), vendor, "", "", "", ""])
                skylark_added += 1
        logger.info(f"[PHASE 2] Finished Skylark processing. Added {skylark_added} rows.")
    except Exception as e:
        logger.error(f"[PHASE 2] Failed processing Skylark Tracker: {e}")

    df = pd.DataFrame(all_data, columns=EXCEL_HEADERS_COURIER)
    logger.info(f"[PHASE 2] Total valid courier rows ready: {len(df)}")
    return df


def load_to_data_xlsx(new_data: pd.DataFrame, excel_path: Path, logger: logging.Logger, dry_run: bool) -> dict:
    if new_data.empty: return {"added": 0, "skipped_duplicate": 0}
    
    # Open or Create the workbook EXACTLY ONCE
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS_COURIER)

    existing_awbs = set()
    for row_idx in range(2, ws.max_row + 1):
        awb = ws.cell(row=row_idx, column=2).value
        if awb: existing_awbs.add(str(awb).strip())

    new_data["_is_duplicate"] = new_data["AWB No."].astype(str).str.strip().isin(existing_awbs)
    duplicates = new_data["_is_duplicate"].sum()
    to_add = new_data[~new_data["_is_duplicate"]].drop(columns=["_is_duplicate"]).copy()[EXCEL_HEADERS_COURIER]

    if dry_run or to_add.empty:
        wb.close()
        return {"added": 0, "skipped_duplicate": int(duplicates)}

    for row in dataframe_to_rows(to_add, index=False, header=False):
        ws.append(row)
        
    wb.save(excel_path)
    wb.close()
    
    logger.info(f"[PHASE 2] Appended {len(to_add)} rows to {excel_path.name}")
    return {"added": len(to_add), "skipped_duplicate": int(duplicates)}


# ═══════════════════════════════════════════════════════════════════════
#  MAIN ETL ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════

def run_etl(run_portal: bool = True, run_courier: bool = True, dry_run: bool = False):
    logger = setup_etl_logging()
    summary = {
        "success": True,
        "dry_run": dry_run,
        "phases": {
            "portal": {"skipped": not run_portal},
            "courier": {"skipped": not run_courier},
        },
    }
    logger.info("=" * 60)
    logger.info("Starting ETL Pipeline")
    logger.info(f"Dry run mode: {dry_run}")
    
    # --- PHASE 1 ---
    if run_portal:
        try:
            logger.info("\n" + "-"*20 + " PHASE 1: Portal Orders " + "-"*20)
            csv_data = extract_from_csv(DEFAULT_CSV_PATH, logger)
            gsheet_data = extract_from_google_sheet(logger)
            merged_data = transform_merge(csv_data, gsheet_data, logger)
            portal_result = load_to_filflo_tasks(merged_data, EXCEL_PATH_SCENARIO_D, logger, dry_run)
            summary["phases"]["portal"] = {"success": True, **portal_result}
        except Exception as e:
            logger.error(f"[PHASE 1] Failed: {e}", exc_info=True)
            summary["success"] = False
            summary["phases"]["portal"] = {"success": False, "error": str(e)}
    else:
        logger.info("\n[SKIP] Skipping Phase 1 (Portal Orders)")

    # --- PHASE 2 ---
    if run_courier:
        try:
            logger.info("\n" + "-"*20 + " PHASE 2: Courier Trackers " + "-"*20)
            courier_data = extract_courier_data(logger)
            courier_result = load_to_data_xlsx(courier_data, EXCEL_PATH_COURIER, logger, dry_run)
            summary["phases"]["courier"] = {"success": True, **courier_result}
        except Exception as e:
            logger.error(f"[PHASE 2] Failed: {e}", exc_info=True)
            summary["success"] = False
            summary["phases"]["courier"] = {"success": False, "error": str(e)}
    else:
        logger.info("\n[SKIP] Skipping Phase 2 (Courier Trackers)")

    logger.info("=" * 60)
    logger.info("Pipeline Complete!")
    return summary

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filflo + Courier Data Feeder")
    parser.add_argument("--dry-run", action="store_true", help="Preview data without writing to Excel")
    parser.add_argument("--portal", action="store_true", help="Run ONLY the Portal Orders module")
    parser.add_argument("--courier", action="store_true", help="Run ONLY the Courier Trackers module")
    args = parser.parse_args()
    
    do_portal = True
    do_courier = True

    if args.portal and not args.courier:
        do_courier = False
    elif args.courier and not args.portal:
        do_portal = False

    run_etl(run_portal=do_portal, run_courier=do_courier, dry_run=args.dry_run)
